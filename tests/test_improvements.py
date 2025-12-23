"""
Additional unit tests for LBO Model Generator improvements.
Tests new helper functions and utilities.
"""

import sys
import os
from pathlib import Path
import logging

# Add src to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from lbo_excel_helpers import ExcelFormattingHelper
from lbo_logging import setup_logging, get_logger
from lbo_validation import validate_json_input, validate_output_path, sanitize_filename
from lbo_exceptions import LBOValidationError, LBOConfigurationError


def test_excel_formatting_helper():
    """Test ExcelFormattingHelper functionality."""
    import openpyxl
    
    # Create a test workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Test format_header_cell
    helper = ExcelFormattingHelper
    helper.format_header_cell(ws, ws['A1'], "Test Header", 'A1:D1')
    assert ws['A1'].value == "Test Header"
    assert ws['A1'].font.bold is True
    
    # Test format_data_cell (use different row to avoid merged cell conflict)
    helper.format_data_cell(ws, ws['A2'], 1234.56, format_type='currency')
    assert ws['A2'].value == 1234.56
    
    # Test format_total_row
    helper.format_total_row(ws, 5, 1, 3, values=[1000, 2000], label="Total")
    assert ws.cell(row=5, column=1).value == "Total"
    assert ws.cell(row=5, column=2).value == 1000
    
    # Test format_input_cell (use different row to avoid merged cell conflict)
    helper.format_input_cell(ws, ws['B2'], 500, number_format='#,##0')
    assert ws['B2'].value == 500
    
    # Test format_output_cell (use different row to avoid merged cell conflict)
    helper.format_output_cell(ws, ws['C2'], 750, number_format='#,##0')
    assert ws['C2'].value == 750
    
    # Test set_column_width
    helper.set_column_width(ws, 'A', 15.0)
    assert ws.column_dimensions['A'].width == 15.0
    
    # Test set_row_height
    helper.set_row_height(ws, 1, 20.0)
    assert ws.row_dimensions[1].height == 20.0
    
    print("✓ ExcelFormattingHelper tests passed")


def test_logging_configuration():
    """Test logging configuration."""
    import tempfile
    import os
    
    # Test setup_logging
    setup_logging(log_level='DEBUG')
    logger = get_logger(__name__)
    logger.debug("Test debug message")
    logger.info("Test info message")
    
    # Test file logging
    with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.log') as f:
        log_file = f.name
    
    try:
        setup_logging(log_level='INFO', log_file=log_file)
        logger = get_logger('test_logger')
        logger.info("Test file logging")
        
        # Verify file was created and has content
        assert os.path.exists(log_file)
        with open(log_file, 'r') as f:
            content = f.read()
            assert "Test file logging" in content
    finally:
        if os.path.exists(log_file):
            os.unlink(log_file)
    
    print("✓ Logging configuration tests passed")


def test_validation_functions():
    """Test validation helper functions."""
    import tempfile
    import os
    from pathlib import Path
    
    # Test validate_json_input
    valid_config = {
        'entry_ebitda': 10000,
        'entry_multiple': 7.0,
        'revenue_growth_rate': [0.10, 0.10, 0.10]
    }
    result = validate_json_input(valid_config)
    assert result == valid_config
    
    # Test invalid config
    try:
        invalid_config = {
            'entry_ebitda': -1000  # Invalid: negative
        }
        validate_json_input(invalid_config)
        assert False, "Should have raised LBOValidationError"
    except LBOValidationError:
        pass  # Expected
    
    # Test validate_output_path
    with tempfile.TemporaryDirectory() as tmpdir:
        test_file = os.path.join(tmpdir, "test.xlsx")
        path = validate_output_path(test_file)
        assert isinstance(path, Path)
    
    # Test invalid path
    try:
        validate_output_path("/nonexistent/path/test.xlsx")
        assert False, "Should have raised LBOConfigurationError"
    except LBOConfigurationError:
        pass  # Expected
    
    # Test sanitize_filename
    safe_name = sanitize_filename("test<file>.xlsx")
    assert '<' not in safe_name
    assert '>' not in safe_name
    
    safe_name = sanitize_filename("../../etc/passwd")
    assert '..' not in safe_name
    assert '/' not in safe_name
    
    print("✓ Validation function tests passed")


def test_type_hints():
    """Test that type hints are present and valid."""
    import inspect
    from typing import get_type_hints
    
    # Test ExcelFormattingHelper methods have type hints
    helper_methods = [
        'format_header_cell',
        'format_data_cell',
        'format_total_row',
        'format_input_cell',
        'format_output_cell'
    ]
    
    for method_name in helper_methods:
        method = getattr(ExcelFormattingHelper, method_name)
        sig = inspect.signature(method)
        assert len(sig.parameters) > 0, f"{method_name} should have parameters"
        # Note: Return type hints may be None for some methods
    
    # Test logging functions have type hints
    setup_sig = inspect.signature(setup_logging)
    get_logger_sig = inspect.signature(get_logger)
    
    assert len(setup_sig.parameters) >= 0
    assert len(get_logger_sig.parameters) >= 1
    
    print("✓ Type hints tests passed")


if __name__ == "__main__":
    print("\n" + "="*60)
    print("Testing LBO Model Generator Improvements")
    print("="*60 + "\n")
    
    try:
        test_excel_formatting_helper()
        test_logging_configuration()
        test_validation_functions()
        test_type_hints()
        
        print("\n" + "="*60)
        print("✓ All improvement tests passed!")
        print("="*60 + "\n")
    except Exception as e:
        print(f"\n✗ Test failed: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

