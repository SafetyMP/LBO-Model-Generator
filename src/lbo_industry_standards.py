"""
Industry-Standard LBO Model Template
Based on investment banking best practices from top-tier banks.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Optional, Any
import openpyxl


class IndustryStandardTemplate:
    """Industry-standard Excel template following investment banking best practices."""

    # Color coding standards (from AI recommendations)
    COLOR_INPUT = "FFC7CE"  # Light red - Input cells
    COLOR_CALCULATION = "C6EFCE"  # Light green - Calculation cells
    COLOR_OUTPUT = "FFEB9C"  # Light yellow - Output cells
    COLOR_HEADER = "4472C4"  # Blue - Headers (industry standard)
    COLOR_TOTAL = "D9E1F2"  # Light blue - Total rows
    COLOR_BORDER = "000000"  # Black borders

    # Font standards
    FONT_NAME = "Calibri"
    FONT_SIZE_BODY = 11
    FONT_SIZE_HEADER = 14
    FONT_SIZE_SECTION = 12
    FONT_SIZE_SUBTOTAL = 11

    # Fonts
    FONT_HEADER = Font(name=FONT_NAME, size=FONT_SIZE_HEADER, bold=True, color="FFFFFF")
    FONT_SECTION = Font(name=FONT_NAME, size=FONT_SIZE_SECTION, bold=True)
    FONT_BODY = Font(name=FONT_NAME, size=FONT_SIZE_BODY)
    FONT_INPUT = Font(name=FONT_NAME, size=FONT_SIZE_BODY, italic=False)
    FONT_CALCULATION = Font(name=FONT_NAME, size=FONT_SIZE_BODY)
    FONT_OUTPUT = Font(name=FONT_NAME, size=FONT_SIZE_BODY, bold=True)
    FONT_TOTAL = Font(name=FONT_NAME, size=FONT_SIZE_BODY, bold=True)
    FONT_SUBTOTAL = Font(name=FONT_NAME, size=FONT_SIZE_SUBTOTAL, italic=True)

    # Fills
    FILL_INPUT = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    FILL_CALCULATION = PatternFill(
        start_color=COLOR_CALCULATION, end_color=COLOR_CALCULATION, fill_type="solid"
    )
    FILL_OUTPUT = PatternFill(start_color=COLOR_OUTPUT, end_color=COLOR_OUTPUT, fill_type="solid")
    FILL_HEADER = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    FILL_TOTAL = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type="solid")

    # Borders
    BORDER_THIN = Border(
        left=Side(style="thin", color=COLOR_BORDER),
        right=Side(style="thin", color=COLOR_BORDER),
        top=Side(style="thin", color=COLOR_BORDER),
        bottom=Side(style="thin", color=COLOR_BORDER),
    )
    BORDER_THICK = Border(
        left=Side(style="thick", color=COLOR_BORDER),
        right=Side(style="thick", color=COLOR_BORDER),
        top=Side(style="thick", color=COLOR_BORDER),
        bottom=Side(style="thick", color=COLOR_BORDER),
    )
    BORDER_DOUBLE_BOTTOM = Border(
        left=Side(style="thin", color=COLOR_BORDER),
        right=Side(style="thin", color=COLOR_BORDER),
        top=Side(style="thin", color=COLOR_BORDER),
        bottom=Side(style="double", color=COLOR_BORDER),
    )

    # Sheet names (industry standard order)
    SHEET_COVER = "Cover"
    SHEET_SUMMARY = "Summary"
    SHEET_SOURCES_USES = "Sources & Uses"
    SHEET_ASSUMPTIONS = "Assumptions"
    SHEET_INCOME_STATEMENT = "Income Statement"
    SHEET_CASH_FLOW = "Cash Flow"
    SHEET_BALANCE_SHEET = "Balance Sheet"
    SHEET_DEBT_SCHEDULE = "Debt Schedule"
    SHEET_RETURNS = "Returns Analysis"
    SHEET_SENSITIVITY = "Sensitivity Analysis"

    @staticmethod
    def format_input_cell(
        cell: openpyxl.cell.cell.Cell, value: Optional[Any] = None
    ) -> openpyxl.cell.cell.Cell:
        """Format cell as input (light red background)."""
        if value is not None:
            # Round to 2 decimal places
            from lbo_constants import LBOConstants

            if isinstance(value, (int, float)):
                cell.value = round(value, LBOConstants.DECIMAL_PLACES)
            else:
                cell.value = value
        cell.font = IndustryStandardTemplate.FONT_INPUT
        cell.fill = IndustryStandardTemplate.FILL_INPUT
        cell.border = IndustryStandardTemplate.BORDER_THIN
        cell.alignment = Alignment(horizontal="right", vertical="center")
        # Set number format to 2 decimal places
        if isinstance(cell.value, (int, float)):
            cell.number_format = "#,##0.00"
        return cell

    @staticmethod
    def format_calculation_cell(
        cell: openpyxl.cell.cell.Cell, formula: Optional[str] = None, value: Optional[Any] = None
    ) -> openpyxl.cell.cell.Cell:
        """Format cell as calculation (light green background)."""
        if formula:
            cell.value = formula
        elif value is not None:
            # Round to 2 decimal places
            from lbo_constants import LBOConstants

            if isinstance(value, (int, float)):
                cell.value = round(value, LBOConstants.DECIMAL_PLACES)
            else:
                cell.value = value
        cell.font = IndustryStandardTemplate.FONT_CALCULATION
        cell.fill = IndustryStandardTemplate.FILL_CALCULATION
        cell.border = IndustryStandardTemplate.BORDER_THIN
        cell.alignment = Alignment(horizontal="right", vertical="center")
        # Set number format to 2 decimal places
        if isinstance(cell.value, (int, float)):
            cell.number_format = "#,##0.00"
        return cell

    @staticmethod
    def format_output_cell(
        cell: openpyxl.cell.cell.Cell, value: Optional[Any] = None
    ) -> openpyxl.cell.cell.Cell:
        """Format cell as output (light yellow background)."""
        if value is not None:
            # Round to 2 decimal places
            from lbo_constants import LBOConstants

            if isinstance(value, (int, float)):
                cell.value = round(value, LBOConstants.DECIMAL_PLACES)
            else:
                cell.value = value
        cell.font = IndustryStandardTemplate.FONT_OUTPUT
        cell.fill = IndustryStandardTemplate.FILL_OUTPUT
        cell.border = IndustryStandardTemplate.BORDER_THIN
        cell.alignment = Alignment(horizontal="right", vertical="center")
        # Set number format to 2 decimal places
        if isinstance(cell.value, (int, float)):
            cell.number_format = "#,##0.00"
        return cell

    @staticmethod
    def format_header_cell(
        cell: openpyxl.cell.cell.Cell, text: str, merge_range: Optional[str] = None
    ) -> openpyxl.cell.cell.Cell:
        """Format section header."""
        if merge_range:
            cell.parent.merge_cells(merge_range)
        cell.value = text
        cell.font = IndustryStandardTemplate.FONT_HEADER
        cell.fill = IndustryStandardTemplate.FILL_HEADER
        cell.border = IndustryStandardTemplate.BORDER_THICK
        cell.alignment = Alignment(horizontal="center", vertical="center")
        return cell

    @staticmethod
    def format_section_header(cell: openpyxl.cell.cell.Cell, text: str) -> openpyxl.cell.cell.Cell:
        """Format subsection header."""
        cell.value = text
        cell.font = IndustryStandardTemplate.FONT_SECTION
        cell.border = IndustryStandardTemplate.BORDER_THICK
        cell.alignment = Alignment(horizontal="left", vertical="center")
        return cell

    @staticmethod
    def format_total_row(
        cell: openpyxl.cell.cell.Cell, value: Optional[Any] = None
    ) -> openpyxl.cell.cell.Cell:
        """Format total row (bold, double underline)."""
        if value is not None:
            # Round to 2 decimal places
            from lbo_constants import LBOConstants

            if isinstance(value, (int, float)):
                cell.value = round(value, LBOConstants.DECIMAL_PLACES)
            else:
                cell.value = value
        cell.font = IndustryStandardTemplate.FONT_TOTAL
        cell.fill = IndustryStandardTemplate.FILL_TOTAL
        cell.border = IndustryStandardTemplate.BORDER_DOUBLE_BOTTOM
        cell.alignment = Alignment(horizontal="right", vertical="center")
        # Set number format to 2 decimal places
        if isinstance(cell.value, (int, float)):
            cell.number_format = "#,##0.00"
        return cell

    @staticmethod
    def format_subtotal_row(
        cell: openpyxl.cell.cell.Cell, value: Optional[Any] = None
    ) -> openpyxl.cell.cell.Cell:
        """Format subtotal row (italic)."""
        if value is not None:
            cell.value = value
        cell.font = IndustryStandardTemplate.FONT_SUBTOTAL
        cell.border = IndustryStandardTemplate.BORDER_THIN
        cell.alignment = Alignment(horizontal="right", vertical="center")
        return cell

    @staticmethod
    def create_formula_with_error_handling(formula: str) -> str:
        """Wrap formula in IFERROR for error handling."""
        return f"=IFERROR({formula},0)"

    @staticmethod
    def get_column_structure(num_projected_years: int = 5) -> Dict[str, int]:
        """Get column structure: 3 historical, 1 transaction, N projected."""
        return {
            "label_col": 1,  # Column A
            "historical_start": 2,  # Column B (Year -3)
            "historical_mid": 3,  # Column C (Year -2)
            "historical_end": 4,  # Column D (Year -1)
            "transaction": 5,  # Column E (Transaction adjustments)
            "projected_start": 6,  # Column F (Year 1)
            "projected_end": 6 + num_projected_years - 1,  # Last projected year
        }
