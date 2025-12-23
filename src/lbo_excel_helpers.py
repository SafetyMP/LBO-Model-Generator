"""
Excel formatting helper utilities for LBO Model Generator.
Extracts common formatting patterns to reduce code duplication.
"""

from typing import Optional, Union
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from .lbo_industry_standards import IndustryStandardTemplate
except ImportError:
    from lbo_industry_standards import IndustryStandardTemplate


class ExcelFormattingHelper:
    """Helper class for common Excel formatting operations."""

    @staticmethod
    def format_header_cell(
        ws: openpyxl.worksheet.worksheet.Worksheet,
        cell: openpyxl.cell.cell.Cell,
        text: str,
        merge_range: Optional[str] = None,
    ) -> None:
        """Format a header cell with standard styling.

        Args:
            ws: Worksheet object
            cell: Cell to format
            text: Text to set in cell
            merge_range: Optional merge range (e.g., 'A1:D1')
        """
        if merge_range:
            ws.merge_cells(merge_range)

        cell.value = text
        cell.font = IndustryStandardTemplate.FONT_HEADER
        cell.fill = IndustryStandardTemplate.FILL_HEADER
        cell.border = IndustryStandardTemplate.BORDER_THICK
        cell.alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def format_data_cell(
        ws: openpyxl.worksheet.worksheet.Worksheet,
        cell: openpyxl.cell.cell.Cell,
        value: Union[str, int, float],
        format_type: str = "general",
        number_format: Optional[str] = None,
        font_style: Optional[Font] = None,
        fill_style: Optional[PatternFill] = None,
    ) -> None:
        """Format a data cell with standard styling.

        Args:
            ws: Worksheet object
            cell: Cell to format
            value: Value to set in cell
            format_type: Format type ('general', 'currency', 'percentage', 'thousands')
            number_format: Custom number format string (overrides format_type)
            font_style: Optional custom font (defaults to FONT_BODY)
            fill_style: Optional custom fill (defaults to none)
        """
        cell.value = value

        # Set font
        if font_style:
            cell.font = font_style
        else:
            cell.font = IndustryStandardTemplate.FONT_BODY

        # Set fill if provided
        if fill_style:
            cell.fill = fill_style

        # Set number format
        if number_format:
            cell.number_format = number_format
        elif format_type == "currency":
            cell.number_format = "#,##0.00"
        elif format_type == "percentage":
            cell.number_format = "0.0%"
        elif format_type == "thousands":
            cell.number_format = "#,##0"
        elif format_type == "general":
            # No specific format
            pass

        # Standard border and alignment
        cell.border = IndustryStandardTemplate.BORDER_THIN
        cell.alignment = Alignment(horizontal="right", vertical="center")

    @staticmethod
    def format_total_row(
        ws: openpyxl.worksheet.worksheet.Worksheet,
        row: int,
        start_col: int,
        end_col: int,
        values: Optional[list] = None,
        label: str = "Total",
    ) -> None:
        """Format a total row with standard styling.

        Args:
            ws: Worksheet object
            row: Row number
            start_col: Starting column number
            end_col: Ending column number (inclusive)
            values: Optional list of values for each column
            label: Label for the first column (default: "Total")
        """
        # Format label cell
        label_cell = ws.cell(row=row, column=start_col)
        label_cell.value = label
        label_cell.font = IndustryStandardTemplate.FONT_TOTAL
        label_cell.fill = IndustryStandardTemplate.FILL_TOTAL
        label_cell.border = IndustryStandardTemplate.BORDER_THICK
        label_cell.alignment = Alignment(horizontal="left", vertical="center")

        # Format data cells
        for col_idx, col in enumerate(range(start_col + 1, end_col + 1), start=0):
            cell = ws.cell(row=row, column=col)

            if values and col_idx < len(values):
                cell.value = values[col_idx]

            cell.font = IndustryStandardTemplate.FONT_TOTAL
            cell.fill = IndustryStandardTemplate.FILL_TOTAL
            cell.border = IndustryStandardTemplate.BORDER_THICK
            cell.alignment = Alignment(horizontal="right", vertical="center")

            # Apply number format if value is numeric
            if values and col_idx < len(values) and isinstance(values[col_idx], (int, float)):
                cell.number_format = "#,##0.00"

    @staticmethod
    def format_input_cell(
        ws: openpyxl.worksheet.worksheet.Worksheet,
        cell: openpyxl.cell.cell.Cell,
        value: Union[str, int, float],
        number_format: Optional[str] = None,
    ) -> None:
        """Format an input cell with standard styling (light red background).

        Args:
            ws: Worksheet object
            cell: Cell to format
            value: Value to set in cell
            number_format: Optional number format string
        """
        cell.value = value
        cell.font = IndustryStandardTemplate.FONT_INPUT
        cell.fill = IndustryStandardTemplate.FILL_INPUT
        cell.border = IndustryStandardTemplate.BORDER_THIN
        cell.alignment = Alignment(horizontal="right", vertical="center")

        if number_format:
            cell.number_format = number_format

    @staticmethod
    def format_output_cell(
        ws: openpyxl.worksheet.worksheet.Worksheet,
        cell: openpyxl.cell.cell.Cell,
        value: Union[str, int, float],
        number_format: Optional[str] = None,
    ) -> None:
        """Format an output cell with standard styling (light yellow background).

        Args:
            ws: Worksheet object
            cell: Cell to format
            value: Value to set in cell
            number_format: Optional number format string
        """
        cell.value = value
        cell.font = IndustryStandardTemplate.FONT_OUTPUT
        cell.fill = IndustryStandardTemplate.FILL_OUTPUT
        cell.border = IndustryStandardTemplate.BORDER_THIN
        cell.alignment = Alignment(horizontal="right", vertical="center")

        if number_format:
            cell.number_format = number_format

    @staticmethod
    def format_section_header(
        ws: openpyxl.worksheet.worksheet.Worksheet,
        cell: openpyxl.cell.cell.Cell,
        text: str,
        merge_range: Optional[str] = None,
    ) -> None:
        """Format a section header with standard styling.

        Args:
            ws: Worksheet object
            cell: Cell to format
            text: Text to set in cell
            merge_range: Optional merge range
        """
        if merge_range:
            ws.merge_cells(merge_range)

        cell.value = text
        cell.font = IndustryStandardTemplate.FONT_SECTION
        cell.border = IndustryStandardTemplate.BORDER_THICK
        cell.alignment = Alignment(horizontal="left", vertical="center")

    @staticmethod
    def set_column_width(
        ws: openpyxl.worksheet.worksheet.Worksheet, column: Union[str, int], width: float
    ) -> None:
        """Set column width.

        Args:
            ws: Worksheet object
            column: Column letter (e.g., 'A') or number (e.g., 1)
            width: Column width
        """
        if isinstance(column, int):
            column = get_column_letter(column)

        ws.column_dimensions[column].width = width

    @staticmethod
    def set_row_height(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, height: float) -> None:
        """Set row height.

        Args:
            ws: Worksheet object
            row: Row number
            height: Row height
        """
        ws.row_dimensions[row].height = height
