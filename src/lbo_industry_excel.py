"""
Industry-Standard LBO Excel Export
Following investment banking best practices from top-tier banks.
Based on OpenAI recommendations for optimal format.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, quote_sheetname
from typing import List, Tuple, Dict, Any, Optional
import pandas as pd
import logging

try:
    from .lbo_industry_standards import IndustryStandardTemplate
    from .lbo_constants import LBOConstants
except ImportError:
    from lbo_industry_standards import IndustryStandardTemplate
    from lbo_constants import LBOConstants

logger = logging.getLogger(__name__)


class IndustryStandardExcelExporter:
    """Industry-standard Excel exporter for LBO models."""
    
    def __init__(self, model):
        """Initialize exporter with LBO model."""
        self.model = model
        self.wb = None
        self.cols = IndustryStandardTemplate.get_column_structure(len(model.years))
    
    def export(self, filename: str, company_name: str = "LBO Model"):
        """Export model to Excel following industry standards."""
        # Create workbook
        self.wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
        
        # Create sheets in industry-standard order
        self._create_cover_sheet(company_name)
        self._create_summary_sheet(company_name)
        self._create_sources_uses_sheet()
        self._create_assumptions_sheet()
        self._create_income_statement_sheet()
        self._create_cash_flow_sheet()
        self._create_balance_sheet()
        self._create_debt_schedule_sheet()
        self._create_returns_sheet()
        self._create_sensitivity_sheet()
        
        # Add charts to relevant sheets
        self._add_charts()
        
        # Add navigation hyperlinks
        self._add_navigation_hyperlinks()
        
        # Set column widths
        self._set_column_widths()
        
        # Protect formula cells
        self._protect_formulas()
        
        # Save
        self.wb.save(filename)
    
    def _get_navigation_items(self):
        """Get navigation items mapping (label to sheet name)."""
        return [
            ("Summary", IndustryStandardTemplate.SHEET_SUMMARY),
            ("Sources & Uses", IndustryStandardTemplate.SHEET_SOURCES_USES),
            ("Assumptions", IndustryStandardTemplate.SHEET_ASSUMPTIONS),
            ("Income Statement", IndustryStandardTemplate.SHEET_INCOME_STATEMENT),
            ("Cash Flow", IndustryStandardTemplate.SHEET_CASH_FLOW),
            ("Balance Sheet", IndustryStandardTemplate.SHEET_BALANCE_SHEET),
            ("Debt Schedule", IndustryStandardTemplate.SHEET_DEBT_SCHEDULE),
            ("Returns Analysis", IndustryStandardTemplate.SHEET_RETURNS),
            ("Sensitivity Analysis", IndustryStandardTemplate.SHEET_SENSITIVITY),
        ]
    
    def _add_navigation_section(self, ws: openpyxl.worksheet.worksheet.Worksheet, start_row: int) -> int:
        """Add navigation section to worksheet."""
        row = self._add_section_header(ws, "NAVIGATION", start_row)
        nav_items = self._get_navigation_items()
        
        for label, sheet_name in nav_items:
            cell = ws.cell(row=row, column=1)
            cell.value = label
            cell.font = IndustryStandardTemplate.FONT_BODY
            row += 1
        
        return row
    
    def _calculate_cover_metrics(self) -> List[tuple]:
        """Calculate and format cover sheet metrics."""
        returns = self.model.calculate_returns()
        entry_ebitda_full = self.model.assumptions.entry_ebitda * LBOConstants.EXCEL_THOUSANDS_DIVISOR
        enterprise_value_full = self.model.enterprise_value * LBOConstants.EXCEL_THOUSANDS_DIVISOR
        calculated_ev = entry_ebitda_full * self.model.assumptions.entry_multiple
        
        if abs(calculated_ev - enterprise_value_full) > 1000.0:
            logger.warning(
                f"Cover sheet: Enterprise Value mismatch. "
                f"Calculated (${calculated_ev:,.0f}) != Model (${enterprise_value_full:,.0f})"
            )
        
        return [
            ("Entry EBITDA ($)", f"${entry_ebitda_full:,.0f}"),
            ("Entry Multiple (x)", f"{self.model.assumptions.entry_multiple:.1f}x"),
            ("Enterprise Value ($)", f"${enterprise_value_full:,.0f}"),
            ("Exit Year", self.model.assumptions.exit_year),
            ("Exit Multiple (x)", f"{self.model.assumptions.exit_multiple:.1f}x"),
            ("MOIC (x)", f"{returns['moic']:.2f}x"),
            ("IRR (%)", f"{returns['irr']*100:.2f}%"),
        ]
    
    def _create_cover_sheet(self, company_name: str):
        """Create Cover sheet with navigation."""
        ws = self.wb.create_sheet(IndustryStandardTemplate.SHEET_COVER, 0)
        
        ws.merge_cells('A1:D1')
        IndustryStandardTemplate.format_header_cell(ws['A1'], "LEVERAGED BUYOUT MODEL", 'A1:D1')
        
        ws.merge_cells('A2:D2')
        cell = ws['A2']
        cell.value = company_name
        cell.font = Font(name=IndustryStandardTemplate.FONT_NAME, size=16, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        from datetime import datetime
        ws['A4'] = f"Date: {datetime.now().strftime('%B %d, %Y')}"
        ws['A4'].font = IndustryStandardTemplate.FONT_BODY
        
        row = self._add_navigation_section(ws, 6)
        row = self._add_section_header(ws, "KEY METRICS", row + 1)
        metrics = self._calculate_cover_metrics()
        self._add_metrics_table(ws, metrics, row)
    
    def _create_summary_sheet(self, company_name: str) -> None:
        """Create Summary sheet with key highlights."""
        ws = self.wb.create_sheet(IndustryStandardTemplate.SHEET_SUMMARY, 1)
        
        row = self._create_sheet_header(ws, "EXECUTIVE SUMMARY", 'A1:D1', 1)
        
        returns = self.model.calculate_returns()
        row = 3
        
        # Investment Highlights
        row = self._add_section_header(ws, "INVESTMENT HIGHLIGHTS", row)
        
        highlights = [
            ("Equity Invested ($)", f"${returns['equity_invested']:,.0f}"),
            ("Exit Equity Value ($)", f"${returns['exit_equity_value']:,.0f}"),
            ("MOIC (x)", f"{returns['moic']:.2f}x"),
            ("IRR (%)", f"{returns['irr']*100:.2f}%"),
        ]
        
        row = self._add_metrics_table(ws, highlights, row)
        row += 1
        
        # Transaction Summary
        row = self._add_section_header(ws, "TRANSACTION SUMMARY", row)
        
        entry_ebitda_full = self.model.assumptions.entry_ebitda * LBOConstants.EXCEL_THOUSANDS_DIVISOR
        exit_ebitda_full = returns['exit_ebitda']
        
        calculated_exit_ev = exit_ebitda_full * self.model.assumptions.exit_multiple
        if abs(calculated_exit_ev - returns['exit_ev']) > 1.0:
            logger.warning(
                f"Summary sheet: Exit EV mismatch. "
                f"Calculated (${calculated_exit_ev:,.0f}) != Returns (${returns['exit_ev']:,.0f})"
            )
        
        calculated_exit_equity = returns['exit_ev'] - returns['exit_debt'] + returns['exit_cash']
        if abs(calculated_exit_equity - returns['exit_equity_value']) > 1.0:
            logger.warning(
                f"Summary sheet: Exit Equity Value mismatch. "
                f"Calculated (${calculated_exit_equity:,.0f}) != Returns (${returns['exit_equity_value']:,.0f})"
            )
        
        transaction = [
            ("Entry EBITDA ($)", f"${entry_ebitda_full:,.0f}"),
            ("Entry Multiple (x)", f"{self.model.assumptions.entry_multiple:.1f}x"),
            ("Enterprise Value ($)", f"${self.model.enterprise_value * LBOConstants.EXCEL_THOUSANDS_DIVISOR:,.0f}"),
            ("Exit EBITDA ($)", f"${exit_ebitda_full:,.0f}"),
            ("Exit Multiple (x)", f"{self.model.assumptions.exit_multiple:.1f}x"),
            ("Exit Enterprise Value ($)", f"${returns['exit_ev']:,.0f}"),
        ]
        
        self._add_metrics_table(ws, transaction, row)
    
    def _calculate_sources(self) -> Tuple[List[Tuple[str, float]], float]:
        """Calculate sources of funds."""
        total_debt = sum(d.amount for d in self.model.assumptions.debt_instruments)
        equity = self.model.assumptions.equity_amount
        total_sources = equity + total_debt
        
        sources = [("Equity", equity)]
        for debt in self.model.assumptions.debt_instruments:
            sources.append((debt.name, debt.amount))
        sources.append(("Total Sources", total_sources))
        
        return sources, total_sources
    
    def _calculate_uses(self, total_debt: float) -> Tuple[List[Tuple[str, float]], float]:
        """Calculate uses of funds."""
        total_uses = self.model.total_uses if hasattr(self.model, 'total_uses') else (
            self.model.equity_value + 
            self.model.assumptions.existing_debt + 
            self.model.transaction_expenses + 
            self.model.financing_fees
        )
        
        uses = [
            ("Purchase Price (Equity Value)", self.model.equity_value),
            ("Existing Debt Repayment", self.model.assumptions.existing_debt),
            ("Transaction Expenses", 
             self.model.transaction_expenses if hasattr(self.model, 'transaction_expenses') 
             else self.model.enterprise_value * self.model.assumptions.transaction_expenses_pct),
            ("Financing Fees", 
             self.model.financing_fees if hasattr(self.model, 'financing_fees') 
             else total_debt * self.model.assumptions.financing_fees_pct),
            ("Total Uses", total_uses),
        ]
        
        return uses, total_uses
    
    def _add_sources_uses_section(self, ws: openpyxl.worksheet.worksheet.Worksheet, 
                                  section_name: str, items: List[Tuple[str, float]],
                                  total: float, start_row: int) -> int:
        """Add a sources or uses section to the worksheet."""
        from openpyxl.styles import Font
        
        row = start_row
        ws.cell(row=row, column=1).value = section_name
        ws.cell(row=row, column=1).font = IndustryStandardTemplate.FONT_SECTION
        ws.cell(row=row, column=2).value = "Amount ($'000)"
        ws.cell(row=row, column=2).font = IndustryStandardTemplate.FONT_SECTION
        ws.cell(row=row, column=3).value = "% of Total"
        ws.cell(row=row, column=3).font = IndustryStandardTemplate.FONT_SECTION
        row += 1
        
        for label, amount in items:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font = IndustryStandardTemplate.FONT_BODY
            ws.cell(row=row, column=1).border = IndustryStandardTemplate.BORDER_THIN
            
            if amount is not None:
                is_total = label.startswith("Total")
                if is_total:
                    IndustryStandardTemplate.format_total_row(
                        ws.cell(row=row, column=2),
                        value=amount / LBOConstants.EXCEL_THOUSANDS_DIVISOR
                    )
                else:
                    IndustryStandardTemplate.format_calculation_cell(
                        ws.cell(row=row, column=2),
                        value=amount / LBOConstants.EXCEL_THOUSANDS_DIVISOR
                    )
                
                if total > 0 and not is_total:
                    pct = amount / total
                    IndustryStandardTemplate.format_calculation_cell(
                        ws.cell(row=row, column=3), value=pct
                    )
                    ws.cell(row=row, column=3).number_format = '0.0%'
                elif is_total:
                    ws.cell(row=row, column=3).value = 1.0
                    ws.cell(row=row, column=3).font = IndustryStandardTemplate.FONT_TOTAL
                    ws.cell(row=row, column=3).number_format = '0.0%'
            
            ws.cell(row=row, column=2).border = IndustryStandardTemplate.BORDER_THIN
            ws.cell(row=row, column=3).border = IndustryStandardTemplate.BORDER_THIN
            row += 1
        
        return row
    
    def _add_balance_check(self, ws: openpyxl.worksheet.worksheet.Worksheet,
                          total_sources: float, total_uses: float, row: int) -> None:
        """Add balance check to sources & uses sheet."""
        from openpyxl.styles import Font
        
        ws.cell(row=row, column=1).value = "Balance Check"
        ws.cell(row=row, column=1).font = IndustryStandardTemplate.FONT_BODY
        balance = total_sources - total_uses
        tolerance = 100
        
        if abs(balance) <= tolerance:
            ws.cell(row=row, column=2).value = "✓ Balanced"
            ws.cell(row=row, column=2).font = Font(
                name=IndustryStandardTemplate.FONT_NAME, size=11, bold=True, color="006100"
            )
            if abs(balance) > 0.01:
                ws.cell(row=row, column=3).value = f"Difference: ${abs(balance):,.0f}"
                ws.cell(row=row, column=3).font = IndustryStandardTemplate.FONT_BODY
        else:
            ws.cell(row=row, column=2).value = f"✗ Unbalanced: ${balance:,.0f}"
            ws.cell(row=row, column=2).font = Font(
                name=IndustryStandardTemplate.FONT_NAME, size=11, bold=True, color="C00000"
            )
            ws.cell(row=row, column=3).value = f"Sources: ${total_sources:,.0f} | Uses: ${total_uses:,.0f}"
            ws.cell(row=row, column=3).font = IndustryStandardTemplate.FONT_BODY
    
    def _create_sources_uses_sheet(self):
        """Create Sources & Uses sheet showing transaction financing."""
        ws = self.wb.create_sheet(IndustryStandardTemplate.SHEET_SOURCES_USES, 2)
        
        ws.merge_cells('A1:D1')
        IndustryStandardTemplate.format_header_cell(
            ws['A1'], "SOURCES & USES OF FUNDS", 'A1:D1'
        )
        
        sources, total_sources = self._calculate_sources()
        total_debt = sum(d.amount for d in self.model.assumptions.debt_instruments)
        uses, total_uses = self._calculate_uses(total_debt)
        
        row = self._add_sources_uses_section(ws, "SOURCES", sources, total_sources, 3)
        row = self._add_sources_uses_section(ws, "USES", uses, total_uses, row + 2)
        self._add_balance_check(ws, total_sources, total_uses, row + 1)
    
    def _create_assumptions_sheet(self):
        """Create Assumptions sheet with all inputs."""
        ws = self.wb.create_sheet(IndustryStandardTemplate.SHEET_ASSUMPTIONS, 2)
        
        # Title
        ws.merge_cells('A1:B1')
        IndustryStandardTemplate.format_header_cell(
            ws['A1'],
            "MODEL ASSUMPTIONS",
            'A1:B1'
        )
        
        # Get assumptions as dictionary
        assumptions = {
            'entry_ebitda': self.model.assumptions.entry_ebitda,
            'entry_multiple': self.model.assumptions.entry_multiple,
            'existing_debt': self.model.assumptions.existing_debt,
            'existing_cash': self.model.assumptions.existing_cash,
            'revenue_growth_rate': self.model.assumptions.revenue_growth_rate,
            'cogs_pct_of_revenue': self.model.assumptions.cogs_pct_of_revenue,
            'sganda_pct_of_revenue': self.model.assumptions.sganda_pct_of_revenue,
            'capex_pct_of_revenue': self.model.assumptions.capex_pct_of_revenue,
            'tax_rate': self.model.assumptions.tax_rate,
            'days_sales_outstanding': self.model.assumptions.days_sales_outstanding,
            'days_inventory_outstanding': self.model.assumptions.days_inventory_outstanding,
            'days_payable_outstanding': self.model.assumptions.days_payable_outstanding,
            'exit_year': self.model.assumptions.exit_year,
            'exit_multiple': self.model.assumptions.exit_multiple,
        }
        row = 3
        
        categories = {
            "Transaction Assumptions": [
                ("Entry EBITDA ($'000)", assumptions.get('entry_ebitda', 0)),
                ("Entry Multiple (x)", assumptions.get('entry_multiple', 0)),
                ("Existing Debt ($'000)", assumptions.get('existing_debt', 0)),
                ("Existing Cash ($'000)", assumptions.get('existing_cash', 0)),
            ],
            "Operating Assumptions": [
                ("Revenue Growth (Year 1) (%)", f"{assumptions.get('revenue_growth_rate', [0])[0]*100:.1f}%"),
                ("COGS % of Revenue", f"{assumptions.get('cogs_pct_of_revenue', 0)*100:.1f}%"),
                ("SG&A % of Revenue", f"{assumptions.get('sganda_pct_of_revenue', 0)*100:.1f}%"),
                ("CapEx % of Revenue", f"{assumptions.get('capex_pct_of_revenue', 0)*100:.1f}%"),
                ("Tax Rate (%)", f"{assumptions.get('tax_rate', 0)*100:.1f}%"),
            ],
            "Working Capital": [
                ("Days Sales Outstanding (days)", assumptions.get('days_sales_outstanding', 0)),
                ("Days Inventory Outstanding (days)", assumptions.get('days_inventory_outstanding', 0)),
                ("Days Payable Outstanding (days)", assumptions.get('days_payable_outstanding', 0)),
            ],
            "Exit Assumptions": [
                ("Exit Year", assumptions.get('exit_year', 5)),
                ("Exit Multiple (x)", assumptions.get('exit_multiple', 0)),
            ],
        }
        
        for category, items in categories.items():
            ws.cell(row=row, column=1).value = category
            ws.cell(row=row, column=1).font = IndustryStandardTemplate.FONT_SECTION
            row += 1
            
            for label, value in items:
                ws.cell(row=row, column=1).value = label
                ws.cell(row=row, column=1).font = IndustryStandardTemplate.FONT_BODY
                IndustryStandardTemplate.format_input_cell(ws.cell(row=row, column=2), value)
                row += 1
            
            row += 1
    
    def _validate_income_statement(self) -> List[tuple]:
        """Validate income statement calculations."""
        validation_issues = []
        for year in self.model.years:
            try:
                revenue = self.model.income_statement.loc['Revenue', year]
                cogs = self.model.income_statement.loc['Cost of Goods Sold (Net of D&A)', year]
                gross_profit = self.model.income_statement.loc['Gross Profit', year]
                sganda = self.model.income_statement.loc['SG&A (Net of D&A)', year]
                ebitda = self.model.income_statement.loc['EBITDA', year]
                
                dep = 0
                amort = 0
                if 'Depreciation' in self.model.income_statement.index:
                    dep = self.model.income_statement.loc['Depreciation', year] if pd.notna(self.model.income_statement.loc['Depreciation', year]) else 0
                if 'Amortization' in self.model.income_statement.index:
                    amort = self.model.income_statement.loc['Amortization', year] if pd.notna(self.model.income_statement.loc['Amortization', year]) else 0
                
                ebit = self.model.income_statement.loc['EBIT', year]
                interest = self.model.income_statement.loc['Interest Expense', year]
                pretax = self.model.income_statement.loc['Pretax Income', year]
                tax = self.model.income_statement.loc['Income Tax Expense', year]
                net_income = self.model.income_statement.loc['Net Income', year]
                
                if abs((revenue - cogs) - gross_profit) > 0.01:
                    validation_issues.append((year, f"Gross Profit: Revenue (${revenue:,.0f}) - COGS (${cogs:,.0f}) ≠ Gross Profit (${gross_profit:,.0f})"))
                
                if abs((gross_profit - sganda) - ebitda) > 100:
                    validation_issues.append((year, f"EBITDA: Gross Profit (${gross_profit:,.0f}) - SG&A (${sganda:,.0f}) ≠ EBITDA (${ebitda:,.0f})"))
                
                if abs((ebitda - dep - amort) - ebit) > 0.01:
                    validation_issues.append((year, f"EBIT: EBITDA (${ebitda:,.0f}) - D&A (${dep+amort:,.0f}) ≠ EBIT (${ebit:,.0f})"))
                
                if abs((ebit - interest) - pretax) > 0.01:
                    validation_issues.append((year, f"Pretax: EBIT (${ebit:,.0f}) - Interest (${interest:,.0f}) ≠ Pretax (${pretax:,.0f})"))
                
                if abs((pretax - tax) - net_income) > 0.01:
                    validation_issues.append((year, f"Net Income: Pretax (${pretax:,.0f}) - Tax (${tax:,.0f}) ≠ Net Income (${net_income:,.0f})"))
            except (KeyError, IndexError, AttributeError):
                pass
        return validation_issues
    
    def _create_income_statement_sheet(self):
        """Create Income Statement sheet following industry standards."""
        ws = self.wb.create_sheet(IndustryStandardTemplate.SHEET_INCOME_STATEMENT, 4)
        
        row = 1
        merge_range = f'A{row}:{get_column_letter(self.cols["projected_end"])}{row}'
        ws.merge_cells(merge_range)
        IndustryStandardTemplate.format_header_cell(ws[f'A{row}'], "INCOME STATEMENT", merge_range)
        
        row = self._add_standard_column_headers(ws, 3)
        
        line_items = [
            ("Revenue", 'Revenue'),
            ("Cost of Goods Sold", 'Cost of Goods Sold (Net of D&A)'),
            ("Gross Profit", 'Gross Profit'),
            ("SG&A", 'SG&A (Net of D&A)'),
            ("EBITDA", 'EBITDA'),
            ("Depreciation", 'Depreciation'),
            ("Amortization", 'Amortization'),
            ("EBIT", 'EBIT'),
            ("Interest Expense", 'Interest Expense'),
            ("Pretax Income", 'Pretax Income'),
            ("Income Tax", 'Income Tax Expense'),
            ("Net Income", 'Net Income'),
        ]
        
        for label, key in line_items:
            cell = ws.cell(row=row, column=1)
            cell.value = label
            cell.font = IndustryStandardTemplate.FONT_BODY
            cell.border = IndustryStandardTemplate.BORDER_THIN
            
            for col in range(2, 6):
                ws.cell(row=row, column=col).border = IndustryStandardTemplate.BORDER_THIN
            
            for year_idx, year in enumerate(self.model.years, start=6):
                try:
                    value = (self.model.income_statement.loc[key, year] 
                            if key in self.model.income_statement.index 
                            else 0)
                    if pd.notna(value):
                        IndustryStandardTemplate.format_calculation_cell(
                            ws.cell(row=row, column=year_idx),
                            value=value / LBOConstants.EXCEL_THOUSANDS_DIVISOR
                        )
                    else:
                        IndustryStandardTemplate.format_calculation_cell(
                            ws.cell(row=row, column=year_idx), value=0
                        )
                except (KeyError, IndexError, AttributeError):
                    IndustryStandardTemplate.format_calculation_cell(
                        ws.cell(row=row, column=year_idx), value=0
                    )
            row += 1
        
        ws.cell(row=row - 1, column=1).font = IndustryStandardTemplate.FONT_TOTAL
        
        validation_issues = self._validate_income_statement()
        self._add_validation_section(ws, "INCOME STATEMENT VALIDATION", validation_issues, row + 2)
    
    def _find_row_in_sheet(self, sheet_name: str, search_text: str) -> Optional[int]:
        """Find row number containing search text in a sheet."""
        if sheet_name not in self.wb.sheetnames:
            return None
        ws = self.wb[sheet_name]
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == search_text:
                return r
        return None
    
    def _add_cross_sheet_consistency_check(self, ws: openpyxl.worksheet.worksheet.Worksheet, 
                                          row: int) -> int:
        """Add cross-sheet consistency validation section."""
        from openpyxl.styles import Font
        
        ws.cell(row=row, column=1).value = "CROSS-SHEET CONSISTENCY"
        ws.cell(row=row, column=1).font = IndustryStandardTemplate.FONT_SECTION
        row += 1
        
        is_net_income_row = self._find_row_in_sheet(IndustryStandardTemplate.SHEET_INCOME_STATEMENT, "Net Income")
        bs_cash_row = self._find_row_in_sheet(IndustryStandardTemplate.SHEET_BALANCE_SHEET, "Cash")
        cf_net_income_row = self._find_row_in_sheet(IndustryStandardTemplate.SHEET_CASH_FLOW, "Net Income")
        cf_ending_cash_row = self._find_row_in_sheet(IndustryStandardTemplate.SHEET_CASH_FLOW, "Ending Cash")
        
        consistency_issues = []
        
        if is_net_income_row and cf_net_income_row:
            row = self._add_consistency_formula_row(ws, "Net Income Check (IS vs CF):", 
                                                    IndustryStandardTemplate.SHEET_INCOME_STATEMENT,
                                                    IndustryStandardTemplate.SHEET_CASH_FLOW,
                                                    is_net_income_row, cf_net_income_row, row)
            consistency_issues.extend(self._check_net_income_consistency())
        
        if bs_cash_row and cf_ending_cash_row:
            row = self._add_consistency_formula_row(ws, "Cash Check (BS vs CF):",
                                                    IndustryStandardTemplate.SHEET_BALANCE_SHEET,
                                                    IndustryStandardTemplate.SHEET_CASH_FLOW,
                                                    bs_cash_row, cf_ending_cash_row, row)
            consistency_issues.extend(self._check_cash_consistency())
        
        row += 1
        if not consistency_issues:
            ws.cell(row=row, column=1).value = "✓ All cross-sheet values consistent (differences < $10)"
            ws.cell(row=row, column=1).font = Font(name=IndustryStandardTemplate.FONT_NAME, size=11, bold=True, color="006100")
        else:
            ws.cell(row=row, column=1).value = "⚠ Cross-sheet consistency issues (differences > $10):"
            ws.cell(row=row, column=1).font = Font(name=IndustryStandardTemplate.FONT_NAME, size=11, bold=True, color="C00000")
            row += 1
            for year, issue in consistency_issues:
                ws.cell(row=row, column=1).value = f"  Year {year}: {issue}"
                ws.cell(row=row, column=1).font = IndustryStandardTemplate.FONT_BODY
                row += 1
        
        return row
    
    def _add_consistency_formula_row(self, ws: openpyxl.worksheet.worksheet.Worksheet,
                                    label: str, sheet1_name: str, sheet2_name: str,
                                    row1: int, row2: int, start_row: int) -> int:
        """Add a consistency check formula row."""
        from openpyxl.styles import Font
        
        ws.cell(row=start_row, column=1).value = label
        ws.cell(row=start_row, column=1).font = IndustryStandardTemplate.FONT_BODY
        
        for year_idx, year in enumerate(self.model.years, start=6):
            sheet1_quoted = quote_sheetname(sheet1_name)
            sheet2_quoted = quote_sheetname(sheet2_name)
            col = get_column_letter(year_idx)
            
            formula = f"=ABS({sheet1_quoted}!{col}{row1}-{sheet2_quoted}!{col}{row2})"
            check_cell = ws.cell(row=start_row, column=year_idx)
            check_cell.value = formula
            check_cell.number_format = '#,##0.00'
            IndustryStandardTemplate.format_calculation_cell(check_cell)
            
            cell_ref = f"{sheet1_quoted}!{col}{row1}"
            check_cell.hyperlink = cell_ref
            check_cell.font = Font(
                name=IndustryStandardTemplate.FONT_NAME,
                size=IndustryStandardTemplate.FONT_SIZE_BODY,
                underline='single',
                color='0563C1'
            )
        
        return start_row + 1
    
    def _check_net_income_consistency(self) -> List[tuple]:
        """Check Net Income consistency between Income Statement and Cash Flow."""
        issues = []
        for year in self.model.years:
            try:
                is_ni = self.model.income_statement.loc['Net Income', year]
                cf_ni = self.model.cash_flow.loc['Net Income', year]
                diff = abs(is_ni - cf_ni)
                if diff > 10:
                    issues.append((year, f"Net Income: IS=${is_ni/1000:,.0f}k, CF=${cf_ni/1000:,.0f}k"))
            except (KeyError, IndexError):
                pass
        return issues
    
    def _check_cash_consistency(self) -> List[tuple]:
        """Check Cash consistency between Balance Sheet and Cash Flow."""
        issues = []
        for year in self.model.years:
            try:
                bs_cash = self.model.balance_sheet.loc['Cash', year]
                cf_cash = self.model.cash_flow.loc['Ending Cash Balance', year]
                diff = abs(bs_cash - cf_cash)
                if diff > 10:
                    issues.append((year, f"Cash: BS=${bs_cash/1000:,.0f}k, CF=${cf_cash/1000:,.0f}k"))
            except (KeyError, IndexError):
                pass
        return issues
    
    def _create_cash_flow_sheet(self):
        """Create Cash Flow Statement sheet."""
        ws = self.wb.create_sheet(IndustryStandardTemplate.SHEET_CASH_FLOW, 5)
        
        row = 1
        merge_range = f'A{row}:{get_column_letter(self.cols["projected_end"])}{row}'
        ws.merge_cells(merge_range)
        IndustryStandardTemplate.format_header_cell(ws[f'A{row}'], "CASH FLOW STATEMENT", merge_range)
        
        row = self._add_standard_column_headers(ws, 3)
        
        cf_items = [
            ("Net Income", 'Net Income'),
            ("Depreciation & Amortization", 'Depreciation & Amortization'),
            ("Change in Accounts Receivable", 'Change in Accounts Receivable'),
            ("Change in Inventory", 'Change in Inventory'),
            ("Change in Accounts Payable", 'Change in Accounts Payable'),
            ("Net Change in Working Capital", 'Net Change in Working Capital'),
            ("Cash Flow from Operations", 'Cash Flow from Operations'),
            ("Capital Expenditures", 'Capital Expenditures'),
            ("Cash Flow from Investing", 'Cash Flow from Investing'),
            ("Debt Repayment", 'Debt Repayment'),
            ("Debt Issuance", 'Debt Issuance'),
            ("Equity Contribution", 'Equity Contribution'),
            ("Purchase Price (Equity Value)", 'Purchase Price (Equity Value)'),
            ("Existing Debt Repayment", 'Existing Debt Repayment'),
            ("Transaction Expenses", 'Transaction Expenses'),
            ("Financing Fees", 'Financing Fees'),
            ("Cash Flow from Financing", 'Cash Flow from Financing'),
            ("Net Change in Cash", 'Net Change in Cash'),
            ("Beginning Cash", 'Beginning Cash Balance'),
            ("Ending Cash", 'Ending Cash Balance'),
        ]
        
        for label, key in cf_items:
            cell = ws.cell(row=row, column=1)
            cell.value = label
            cell.font = IndustryStandardTemplate.FONT_BODY
            cell.border = IndustryStandardTemplate.BORDER_THIN
            
            for year_idx, year in enumerate(self.model.years, start=6):
                try:
                    value = (self.model.income_statement.loc['Net Income', year] 
                            if key == 'Net Income' 
                            else self.model.cash_flow.loc[key, year])
                    if pd.notna(value):
                        IndustryStandardTemplate.format_calculation_cell(
                            ws.cell(row=row, column=year_idx),
                            value=value / LBOConstants.EXCEL_THOUSANDS_DIVISOR
                        )
                except (KeyError, IndexError):
                    IndustryStandardTemplate.format_calculation_cell(
                        ws.cell(row=row, column=year_idx), value=0
                    )
            row += 1
        
        self._add_cross_sheet_consistency_check(ws, row + 2)
    
    def _validate_balance_sheet(self) -> List[tuple]:
        """Validate balance sheet balancing."""
        validation_issues = []
        for year in self.model.years:
            try:
                total_assets = self.model.balance_sheet.loc['Total Assets', year]
                total_liab_eq = self.model.balance_sheet.loc['Total Liabilities & Equity', year]
                diff = abs(total_assets - total_liab_eq)
                if diff > 1.0:
                    validation_issues.append((year, diff))
            except (KeyError, IndexError):
                pass
        return validation_issues
    
    def _create_balance_sheet(self):
        """Create Balance Sheet."""
        ws = self.wb.create_sheet(IndustryStandardTemplate.SHEET_BALANCE_SHEET, 6)
        
        row = 1
        merge_range = f'A{row}:{get_column_letter(self.cols["projected_end"])}{row}'
        ws.merge_cells(merge_range)
        IndustryStandardTemplate.format_header_cell(ws[f'A{row}'], "BALANCE SHEET", merge_range)
        
        row = self._add_standard_column_headers(ws, 3)
        
        row = self._add_section_header(ws, "ASSETS", row)
        asset_items = [
            ("Cash", 'Cash'),
            ("Accounts Receivable", 'Accounts Receivable'),
            ("Inventory", 'Inventory'),
            ("Total Current Assets", 'Total Current Assets'),
            ("PP&E, Net", 'PP&E, Net'),
            ("Goodwill", 'Goodwill'),
            ("Intangible Assets", 'Intangible Assets (Financing Fees)'),
            ("Total Assets", 'Total Assets'),
        ]
        row = self._add_financial_statement_rows(ws, asset_items, row, self.model.balance_sheet)
        
        row = self._add_section_header(ws, "LIABILITIES & EQUITY", row + 1)
        liability_items = [
            ("Accounts Payable", 'Accounts Payable'),
            ("Total Current Liabilities", 'Total Current Liabilities'),
            ("Total Debt", 'Total Debt'),
            ("Total Liabilities", 'Total Liabilities'),
            ("Shareholders Equity", 'Shareholders Equity'),
            ("Total Liabilities & Equity", 'Total Liabilities & Equity'),
        ]
        row = self._add_financial_statement_rows(ws, liability_items, row, self.model.balance_sheet)
        
        validation_issues = self._validate_balance_sheet()
        self._add_validation_section(ws, "BALANCE SHEET VALIDATION", validation_issues, row + 2)
    
    def _add_debt_instrument_row(self, ws: openpyxl.worksheet.worksheet.Worksheet, 
                                 label: str, data_key: str, debt_name: str, row: int) -> int:
        """Add a debt instrument row with year columns."""
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = IndustryStandardTemplate.FONT_BODY
        ws.cell(row=row, column=1).border = IndustryStandardTemplate.BORDER_THIN
        
        for year_idx, year in enumerate(self.model.years, start=6):
            value = self.model.debt_schedule[debt_name][data_key][year - 1]
            cell = ws.cell(row=row, column=year_idx)
            IndustryStandardTemplate.format_calculation_cell(
                cell,
                value=value / LBOConstants.EXCEL_THOUSANDS_DIVISOR
            )
            cell.border = IndustryStandardTemplate.BORDER_THIN
        
        return row + 1
    
    def _add_debt_instrument_section(self, ws: openpyxl.worksheet.worksheet.Worksheet, 
                                    debt, row: int) -> int:
        """Add complete debt instrument section."""
        ws.cell(row=row, column=1).value = debt.name
        ws.cell(row=row, column=1).font = IndustryStandardTemplate.FONT_SECTION
        row += 1
        
        row = self._add_debt_instrument_row(ws, "Beginning Balance", 'beginning_balance', debt.name, row)
        row = self._add_debt_instrument_row(ws, "Interest Paid", 'interest_paid', debt.name, row)
        row = self._add_debt_instrument_row(ws, "Principal Paid (Scheduled + Sweep)", 'principal_paid', debt.name, row)
        row = self._add_debt_instrument_row(ws, "Ending Balance", 'ending_balance', debt.name, row)
        
        return row + 2  # Space between instruments
    
    def _add_standard_column_headers(self, ws: openpyxl.worksheet.worksheet.Worksheet, row: int) -> int:
        """Add standard column headers for financial statements."""
        headers = ["", "Year -3\n($'000)", "Year -2\n($'000)", "Year -1\n($'000)", "Transaction\n($'000)", 
                   "Year 1\n($'000)", "Year 2\n($'000)", "Year 3\n($'000)", "Year 4\n($'000)", "Year 5\n($'000)"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = IndustryStandardTemplate.FONT_SECTION
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = IndustryStandardTemplate.BORDER_THIN
        
        return row + 1
    
    def _validate_debt_schedule(self) -> List[tuple]:
        """Validate debt schedule calculations."""
        validation_issues = []
        for debt in self.model.assumptions.debt_instruments:
            for year in self.model.years:
                try:
                    beg_bal = self.model.debt_schedule[debt.name]['beginning_balance'][year - 1]
                    principal = self.model.debt_schedule[debt.name]['principal_paid'][year - 1]
                    end_bal = self.model.debt_schedule[debt.name]['ending_balance'][year - 1]
                    calc_end = beg_bal - principal
                    diff = abs(calc_end - end_bal)
                    
                    if diff > 0.01:
                        validation_issues.append((debt.name, year, diff))
                except (KeyError, IndexError):
                    pass
        return validation_issues
    
    def _create_debt_schedule_sheet(self):
        """Create Debt Schedule sheet."""
        ws = self.wb.create_sheet(IndustryStandardTemplate.SHEET_DEBT_SCHEDULE, 7)
        
        row = 1
        merge_range = f'A{row}:{get_column_letter(self.cols["projected_end"])}{row}'
        ws.merge_cells(merge_range)
        IndustryStandardTemplate.format_header_cell(ws[f'A{row}'], "DEBT SCHEDULE", merge_range)
        
        row = self._add_standard_column_headers(ws, 3)
        
        for debt in self.model.assumptions.debt_instruments:
            row = self._add_debt_instrument_section(ws, debt, row)
        
        ws.cell(row=row, column=1).value = "Total Debt"
        ws.cell(row=row, column=1).font = IndustryStandardTemplate.FONT_TOTAL
        for year_idx, year in enumerate(self.model.years, start=6):
            total_debt = sum(
                self.model.debt_schedule[d.name]['ending_balance'][year - 1]
                for d in self.model.assumptions.debt_instruments
            )
            IndustryStandardTemplate.format_total_row(
                ws.cell(row=row, column=year_idx),
                value=total_debt / LBOConstants.EXCEL_THOUSANDS_DIVISOR
            )
        
        validation_issues = self._validate_debt_schedule()
        self._add_validation_section(ws, "DEBT SCHEDULE VALIDATION", validation_issues, row + 2)
    
    def _create_returns_sheet(self) -> None:
        """Create Returns Analysis sheet."""
        ws = self.wb.create_sheet(IndustryStandardTemplate.SHEET_RETURNS, 8)
        
        row = self._create_sheet_header(ws, "RETURNS ANALYSIS", 'A1:B1', 1)
        
        returns = self.model.calculate_returns()
        row = 3
        
        metrics = [
            ("Exit Year", returns['exit_year']),
            ("Exit EBITDA ($)", f"${returns['exit_ebitda']:,.0f}"),
            ("Exit Enterprise Value ($)", f"${returns['exit_ev']:,.0f}"),
            ("Exit Equity Value ($)", f"${returns['exit_equity_value']:,.0f}"),
            ("Equity Invested ($)", f"${returns['equity_invested']:,.0f}"),
            ("MOIC (x)", f"{returns['moic']:.2f}x"),
            ("IRR (%)", f"{returns['irr']*100:.2f}%"),
        ]
        
        self._add_metrics_table(ws, metrics, row)
    
    def _format_base_metric_cell(self, ws: openpyxl.worksheet.worksheet.Worksheet, 
                                 label: str, value: Any, row: int) -> None:
        """Format a base case metric cell with appropriate formatting."""
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = IndustryStandardTemplate.FONT_BODY
        
        if isinstance(value, (int, float)):
            if "MOIC" in label:
                formatted_value = f"{value:.2f}x"
            elif "IRR" in label:
                irr_pct = value * 100 if value < 1 else value
                formatted_value = f"{irr_pct:.2f}%"
            elif "Multiple" in label:
                formatted_value = f"{value:.1f}x"
            else:
                formatted_value = value
        else:
            formatted_value = value
        
        IndustryStandardTemplate.format_output_cell(ws.cell(row=row, column=2), formatted_value)
    
    def _calculate_sensitivity_matrix(self, base_returns: Dict) -> List[List]:
        """Calculate sensitivity matrix data."""
        exit_multiples = [
            self.model.assumptions.exit_multiple * 0.8,
            self.model.assumptions.exit_multiple * 0.9,
            self.model.assumptions.exit_multiple,
            self.model.assumptions.exit_multiple * 1.1,
            self.model.assumptions.exit_multiple * 1.2,
        ]
        
        sensitivity_data_rows = []
        for exit_mult in exit_multiples:
            exit_ev_change_factor = exit_mult / self.model.assumptions.exit_multiple
            approx_moic = base_returns['moic'] * exit_ev_change_factor
            approx_irr = (approx_moic ** (1.0 / base_returns['exit_year'])) - 1
            
            sensitivity_data_rows.append([
                exit_mult,
                f"{self.model.assumptions.entry_multiple:.1f}x",
                approx_moic,
                approx_irr
            ])
        
        return sensitivity_data_rows
    
    def _add_sensitivity_table_row(self, ws: openpyxl.worksheet.worksheet.Worksheet, 
                                   data_row: List, row: int) -> None:
        """Add a sensitivity table row with proper formatting."""
        cell = ws.cell(row=row, column=1)
        cell.value = data_row[0]
        cell.font = IndustryStandardTemplate.FONT_BODY
        cell.border = IndustryStandardTemplate.BORDER_THIN
        cell.number_format = '0.0"x"'
        
        cell = ws.cell(row=row, column=2)
        cell.value = data_row[1]
        cell.font = IndustryStandardTemplate.FONT_BODY
        cell.border = IndustryStandardTemplate.BORDER_THIN
        
        cell = ws.cell(row=row, column=3)
        cell.value = data_row[2]
        cell.font = IndustryStandardTemplate.FONT_BODY
        cell.border = IndustryStandardTemplate.BORDER_THIN
        cell.number_format = '0.00"x"'
        
        cell = ws.cell(row=row, column=4)
        cell.value = data_row[3]
        cell.font = IndustryStandardTemplate.FONT_BODY
        cell.border = IndustryStandardTemplate.BORDER_THIN
        cell.number_format = '0.0%'
    
    def _create_sensitivity_sheet(self) -> None:
        """Create Sensitivity Analysis sheet showing impact of key assumption changes."""
        from openpyxl.styles import Font, Alignment
        
        ws = self.wb.create_sheet(IndustryStandardTemplate.SHEET_SENSITIVITY, 9)
        row = self._create_sheet_header(ws, "SENSITIVITY ANALYSIS", 'A1:F1', 1)
        
        base_returns = self.model.calculate_returns()
        row = self._add_section_header(ws, "BASE CASE", 3)
        
        base_metrics = [
            ("Entry Multiple (x)", self.model.assumptions.entry_multiple),
            ("Exit Multiple (x)", self.model.assumptions.exit_multiple),
            ("MOIC (x)", base_returns['moic']),
            ("IRR (%)", base_returns['irr']),
        ]
        
        for label, value in base_metrics:
            self._format_base_metric_cell(ws, label, value, row)
            row += 1
        
        row += 2
        headers = ["Exit Multiple (x)", "Entry Multiple (x)", "MOIC (x)", "IRR (%)"]
        row = self._add_data_table(ws, headers, [], row)
        
        sensitivity_data_rows = self._calculate_sensitivity_matrix(base_returns)
        for data_row in sensitivity_data_rows:
            self._add_sensitivity_table_row(ws, data_row, row)
            row += 1
        
        row += 1
        note_text = (
            "Note: Sensitivity analysis uses simplified approximations assuming exit debt and cash "
            "remain constant. For accurate sensitivity analysis, the full model should be "
            "recalculated for each scenario, accounting for changes in debt paydown, cash accumulation, "
            "and working capital that may result from different exit multiples."
        )
        ws.cell(row=row, column=1).value = note_text
        ws.cell(row=row, column=1).font = Font(name=IndustryStandardTemplate.FONT_NAME, size=9, italic=True)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{row}:D{row}')
        ws.row_dimensions[row].height = 40
    
    def _configure_chart_series(self, series, title_ref, marker_symbol="circle", 
                                  marker_size=8, show_labels=False, label_format=None):
        """Helper method to configure a chart series with common settings.
        
        Args:
            series: Chart series object to configure
            title_ref: Reference object for series title
            marker_symbol: Marker symbol (default: "circle")
            marker_size: Marker size (default: 8)
            show_labels: Whether to show data labels (default: False)
            label_format: Format string for data labels (optional)
        """
        from openpyxl.chart.series import SeriesLabel
        from openpyxl.chart.data_source import StrRef
        from openpyxl.chart.label import DataLabelList
        
        series.title = SeriesLabel(strRef=StrRef(str(title_ref)))
        series.marker.symbol = marker_symbol
        series.marker.size = marker_size
        series.dLbls = DataLabelList()
        series.dLbls.showVal = show_labels
        if label_format:
            series.dLbls.numFmt = label_format
    
    def _configure_line_chart_axes(self, chart, y_title, x_title, y_number_format=None,
                                     x_number_format=None, y_min=0):
        """Helper method to configure line chart axes.
        
        Args:
            chart: LineChart object
            y_title: Y-axis title
            x_title: X-axis title
            y_number_format: Y-axis number format (optional)
            x_number_format: X-axis number format (optional)
            y_min: Y-axis minimum value (default: 0)
        """
        chart.y_axis.title = y_title
        chart.x_axis.title = x_title
        if y_number_format:
            chart.y_axis.number_format = y_number_format
        if x_number_format:
            chart.x_axis.number_format = x_number_format
        chart.y_axis.scaling.min = y_min
    
    def _create_sheet_header(self, ws: openpyxl.worksheet.worksheet.Worksheet, title: str, 
                             merge_range: str = 'A1:D1', row: int = 1) -> int:
        """Create a standard sheet header.
        
        Args:
            ws: Worksheet to add header to
            title: Header title text
            merge_range: Cell range to merge (default: 'A1:D1')
            row: Starting row (default: 1)
            
        Returns:
            Next row number after header
        """
        ws.merge_cells(merge_range)
        IndustryStandardTemplate.format_header_cell(
            ws.cell(row=row, column=1),
            title,
            merge_range
        )
        return row + 1
    
    def _add_section_header(self, ws: openpyxl.worksheet.worksheet.Worksheet, 
                           title: str, row: int, col: int = 1) -> int:
        """Add a section header to worksheet.
        
        Args:
            ws: Worksheet to add section to
            title: Section title
            row: Row number
            col: Column number (default: 1)
            
        Returns:
            Next row number after section header
        """
        ws.cell(row=row, column=col).value = title
        ws.cell(row=row, column=col).font = IndustryStandardTemplate.FONT_SECTION
        return row + 1
    
    def _add_metrics_table(self, ws: openpyxl.worksheet.worksheet.Worksheet, 
                           metrics: List[tuple], start_row: int, label_col: int = 1, 
                           value_col: int = 2) -> int:
        """Add a metrics table (label-value pairs) to worksheet.
        
        Args:
            ws: Worksheet to add metrics to
            metrics: List of (label, value) tuples
            start_row: Starting row number
            label_col: Column for labels (default: 1)
            value_col: Column for values (default: 2)
            
        Returns:
            Next row number after metrics table
        """
        row = start_row
        for label, value in metrics:
            ws.cell(row=row, column=label_col).value = label
            ws.cell(row=row, column=label_col).font = IndustryStandardTemplate.FONT_BODY
            if isinstance(value, str):
                ws.cell(row=row, column=value_col).value = value
                ws.cell(row=row, column=value_col).font = IndustryStandardTemplate.FONT_BODY
            else:
                IndustryStandardTemplate.format_output_cell(
                    ws.cell(row=row, column=value_col), 
                    value
                )
            row += 1
        return row
    
    def _add_data_table(self, ws: openpyxl.worksheet.worksheet.Worksheet, 
                       headers: List[str], data_rows: List[List], start_row: int, 
                       start_col: int = 1) -> int:
        """Add a data table with headers and rows.
        
        Args:
            ws: Worksheet to add table to
            headers: List of header strings
            data_rows: List of data row lists
            start_row: Starting row number
            start_col: Starting column number (default: 1)
            
        Returns:
            Next row number after table
        """
        row = start_row
        
        # Add headers
        for col_idx, header in enumerate(headers, start=start_col):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = IndustryStandardTemplate.FONT_SECTION
            cell.border = IndustryStandardTemplate.BORDER_THIN
            cell.fill = IndustryStandardTemplate.FILL_HEADER
        
        row += 1
        
        # Add data rows
        for data_row in data_rows:
            for col_idx, value in enumerate(data_row, start=start_col):
                cell = ws.cell(row=row, column=col_idx)
                if isinstance(value, (int, float)):
                    IndustryStandardTemplate.format_calculation_cell(cell, value)
                else:
                    cell.value = value
                    cell.font = IndustryStandardTemplate.FONT_BODY
                cell.border = IndustryStandardTemplate.BORDER_THIN
            row += 1
        
        return row
    
    def _add_financial_statement_rows(self, ws: openpyxl.worksheet.worksheet.Worksheet,
                                      items: List[Tuple[str, str]], start_row: int,
                                      data_source, data_key: str = None) -> int:
        """Add financial statement rows with year columns.
        
        Args:
            ws: Worksheet to add rows to
            items: List of (label, key) tuples for row items
            start_row: Starting row number
            data_source: DataFrame or dict to pull data from
            data_key: Optional key if data_source is dict
            
        Returns:
            Next row number after rows
        """
        row = start_row
        for label, key in items:
            cell = ws.cell(row=row, column=1)
            cell.value = label
            cell.font = IndustryStandardTemplate.FONT_BODY
            cell.border = IndustryStandardTemplate.BORDER_THIN
            
            for year_idx, year in enumerate(self.model.years, start=6):
                try:
                    if hasattr(data_source, 'loc'):
                        value = data_source.loc[key, year]
                    else:
                        value = data_source[key][year - 1] if data_key else data_source.get(key, {}).get(year - 1, 0)
                    
                    if pd.notna(value) if hasattr(pd, 'notna') else value is not None:
                        IndustryStandardTemplate.format_calculation_cell(
                            ws.cell(row=row, column=year_idx),
                            value=value / LBOConstants.EXCEL_THOUSANDS_DIVISOR
                        )
                except (KeyError, IndexError):
                    IndustryStandardTemplate.format_calculation_cell(
                        ws.cell(row=row, column=year_idx),
                        value=0
                    )
            row += 1
        return row
    
    def _add_validation_section(self, ws: openpyxl.worksheet.worksheet.Worksheet,
                               title: str, validation_issues: List, start_row: int) -> int:
        """Add a validation section to worksheet.
        
        Args:
            ws: Worksheet to add validation to
            title: Validation section title
            validation_issues: List of validation issues
            start_row: Starting row number
            
        Returns:
            Next row number after validation section
        """
        row = self._add_section_header(ws, title, start_row)
        
        if not validation_issues:
            ws.cell(row=row, column=1).value = f"✓ {title.replace('VALIDATION', '').strip()} valid for all years"
            ws.cell(row=row, column=1).font = Font(name=IndustryStandardTemplate.FONT_NAME, size=11, bold=True, color="006100")
            row += 1
        else:
            for issue in validation_issues:
                if len(issue) == 2:
                    year, diff = issue
                    ws.cell(row=row, column=1).value = f"  Year {year}: Difference of ${diff:,.0f}"
                elif len(issue) == 3:
                    name, year, diff = issue
                    ws.cell(row=row, column=1).value = f"  {name} Year {year}: Beginning - Principal ≠ Ending (diff: ${diff:,.0f})"
                ws.cell(row=row, column=1).font = IndustryStandardTemplate.FONT_BODY
                row += 1
        
        return row
    
    def _create_revenue_ebitda_chart(self, ws_is: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Create Revenue and EBITDA Growth chart on Income Statement sheet.
        
        Args:
            ws_is: Income Statement worksheet
        """
        from openpyxl.chart import LineChart, Reference
        
        # Find revenue and EBITDA rows
        revenue_row = None
        ebitda_row = None
        for row_idx in range(4, 30):
            label = ws_is.cell(row=row_idx, column=1).value
            if label == "Revenue":
                revenue_row = row_idx
            elif label == "EBITDA":
                ebitda_row = row_idx
            if revenue_row and ebitda_row:
                break
        
        if not (revenue_row and ebitda_row):
            logger.debug(f"Revenue row: {revenue_row}, EBITDA row: {ebitda_row} - skipping chart")
            return
        
        try:
            chart = LineChart()
            chart.title = "Revenue and EBITDA Growth"
            chart.style = 10
            chart.legend.position = 'b'
            chart.height = 10
            chart.width = 16
            self._configure_line_chart_axes(
                chart, "Amount ($'000)", 'Year', 
                y_number_format='#,##0', y_min=0
            )
            
            num_years = len(self.model.years)
            if num_years == 0:
                logger.warning("Cannot create Revenue/EBITDA chart - no years in model")
                return
            
            try:
                categories = Reference(ws_is, min_col=6, min_row=3, max_col=5+num_years, max_row=3)
                revenue_data = Reference(ws_is, min_col=6, min_row=revenue_row, max_col=5+num_years, max_row=revenue_row)
                ebitda_data = Reference(ws_is, min_col=6, min_row=ebitda_row, max_col=5+num_years, max_row=ebitda_row)
                
                title_col = 11
                ws_is.cell(row=revenue_row, column=title_col).value = "Revenue"
                ws_is.cell(row=ebitda_row, column=title_col).value = "EBITDA"
                revenue_title_ref = Reference(ws_is, min_col=title_col, min_row=revenue_row, max_col=title_col, max_row=revenue_row)
                ebitda_title_ref = Reference(ws_is, min_col=title_col, min_row=ebitda_row, max_col=title_col, max_row=ebitda_row)
                
                chart.add_data(revenue_data, titles_from_data=False)
                chart.add_data(ebitda_data, titles_from_data=False)
                
                if len(chart.series) >= 1:
                    self._configure_chart_series(
                        chart.series[0], revenue_title_ref, 
                        marker_symbol="circle", show_labels=False
                    )
                
                if len(chart.series) >= 2:
                    self._configure_chart_series(
                        chart.series[1], ebitda_title_ref,
                        marker_symbol="diamond", show_labels=False
                    )
                
                chart.set_categories(categories)
                chart_start_row = max(revenue_row, ebitda_row) + 2
                ws_is.add_chart(chart, f"K{chart_start_row}")
                logger.info("✓ Created Revenue and EBITDA Growth chart")
            except (ValueError, TypeError, AttributeError, IndexError) as e:
                logger.warning(f"Failed to create Revenue/EBITDA chart - data validation error: {e}")
            except Exception as e:
                logger.warning(f"Failed to create Revenue/EBITDA chart: {e}", exc_info=True)
        except Exception as e:
            logger.warning(f"Failed to create Revenue/EBITDA chart: {e}", exc_info=True)
    
    def _create_returns_chart(self, ws_returns: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Create Investment Returns Summary chart on Returns sheet.
        
        Args:
            ws_returns: Returns Analysis worksheet
        """
        from openpyxl.chart import BarChart, Reference
        
        # Get MOIC and IRR values
        moic_row = None
        irr_row = None
        for row_idx in range(3, 15):
            label = ws_returns.cell(row=row_idx, column=1).value
            if label == "MOIC":
                moic_row = row_idx
            elif label == "IRR":
                irr_row = row_idx
            if moic_row and irr_row:
                break
        
        if not (moic_row and irr_row):
            return
        
        try:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Investment Returns Summary"
            chart.style = 10
            chart.y_axis.title = 'Value'
            chart.height = 8
            chart.width = 12
            chart.y_axis.scaling.min = 0
            
            moic_value = ws_returns.cell(row=moic_row, column=2).value
            irr_value = ws_returns.cell(row=irr_row, column=2).value
            
            try:
                if isinstance(moic_value, str):
                    moic_num = float(moic_value.replace('x', '').replace('$', '').replace(',', ''))
                else:
                    moic_num = float(moic_value) if moic_value is not None else 0
                
                if isinstance(irr_value, str):
                    irr_num = float(irr_value.replace('%', '').replace('$', '').replace(',', ''))
                else:
                    irr_num = float(irr_value) if irr_value is not None else 0
            except (ValueError, TypeError, AttributeError) as e:
                logger.warning(f"Failed to extract numeric values for Returns chart: {e}")
                moic_num = 0
                irr_num = 0
            
            if moic_num > 0 or irr_num > 0:
                ws_returns['E1'] = 'MOIC'
                ws_returns['E2'] = moic_num
                ws_returns['F1'] = 'IRR (%)'
                ws_returns['F2'] = irr_num
                
                data = Reference(ws_returns, min_col=5, min_row=1, max_col=6, max_row=2)
                categories = Reference(ws_returns, min_col=5, min_row=1, max_row=1)
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)
                ws_returns.add_chart(chart, "D3")
                logger.info("✓ Created Investment Returns Summary chart")
            else:
                logger.warning("Skipping Returns chart - no valid numeric data")
        except Exception as e:
            logger.warning(f"Failed to create Returns chart: {e}", exc_info=True)
    
    def _create_capital_structure_charts(self, ws_bs: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Create Capital Structure Trends charts on Balance Sheet.
        
        Args:
            ws_bs: Balance Sheet worksheet
        """
        from openpyxl.chart import LineChart, PieChart, Reference
        from openpyxl.chart.label import DataLabelList
        from openpyxl.utils import get_column_letter
        
        # Find key balance sheet items
        total_debt_row = None
        equity_row = None
        for row_idx in range(4, 40):
            label = ws_bs.cell(row=row_idx, column=1).value
            if label == "Total Debt":
                total_debt_row = row_idx
            elif label == "Shareholders Equity":
                equity_row = row_idx
            if total_debt_row and equity_row:
                break
        
        if not (total_debt_row and equity_row and len(self.model.years) > 0):
            return
        
        try:
            # Line chart showing debt and equity trends
            chart = LineChart()
            chart.title = "Capital Structure Trends"
            chart.style = 10
            chart.legend.position = 'b'
            chart.height = 10
            chart.width = 16
            self._configure_line_chart_axes(
                chart, "Amount ($'000)", 'Year',
                y_number_format='#,##0', y_min=0
            )
            
            num_years = len(self.model.years)
            try:
                categories = Reference(ws_bs, min_col=6, min_row=3, max_col=5+num_years, max_row=3)
                debt_data = Reference(ws_bs, min_col=6, min_row=total_debt_row, max_col=5+num_years, max_row=total_debt_row)
                equity_data = Reference(ws_bs, min_col=6, min_row=equity_row, max_col=5+num_years, max_row=equity_row)
                
                title_col = 11
                ws_bs.cell(row=total_debt_row, column=title_col).value = "Total Debt"
                ws_bs.cell(row=equity_row, column=title_col).value = "Shareholders Equity"
                debt_title_ref = Reference(ws_bs, min_col=title_col, min_row=total_debt_row, max_col=title_col, max_row=total_debt_row)
                equity_title_ref = Reference(ws_bs, min_col=title_col, min_row=equity_row, max_col=title_col, max_row=equity_row)
                
                chart.add_data(debt_data, titles_from_data=False)
                chart.add_data(equity_data, titles_from_data=False)
                
                if len(chart.series) >= 1:
                    self._configure_chart_series(
                        chart.series[0], debt_title_ref,
                        marker_symbol="square", show_labels=False
                    )
                
                if len(chart.series) >= 2:
                    self._configure_chart_series(
                        chart.series[1], equity_title_ref,
                        marker_symbol="triangle", show_labels=False
                    )
                
                chart.set_categories(categories)
                chart_start_row = max(total_debt_row, equity_row) + 2
                ws_bs.add_chart(chart, f"K{chart_start_row}")
                logger.info("✓ Created Capital Structure Trends chart")
            except (ValueError, TypeError, AttributeError, IndexError) as e:
                logger.warning(f"Failed to create Capital Structure chart - data validation error: {e}")
                chart_start_row = None
            
            # Pie chart for final year capital structure
            if chart_start_row:
                try:
                    final_year_col = 5 + len(self.model.years)
                    debt_value = ws_bs.cell(row=total_debt_row, column=final_year_col).value
                    equity_value = ws_bs.cell(row=equity_row, column=final_year_col).value
                    
                    try:
                        debt_num = float(debt_value) if isinstance(debt_value, (int, float)) else 0
                        equity_num = float(equity_value) if isinstance(equity_value, (int, float)) else 0
                    except (ValueError, TypeError) as e:
                        logger.warning(f"Failed to extract numeric values for pie chart: {e}")
                        debt_num = 0
                        equity_num = 0
                    
                    if debt_num + equity_num > 0:
                        pie_chart = PieChart()
                        pie_chart.title = "Capital Structure - Final Year"
                        pie_chart.style = 10
                        pie_chart.dataLabels = DataLabelList()
                        pie_chart.dataLabels.showPercent = True
                        pie_chart.dataLabels.showVal = True
                        
                        chart_start_col = final_year_col + 2
                        ws_bs.cell(row=1, column=chart_start_col).value = 'Debt'
                        ws_bs.cell(row=2, column=chart_start_col).value = debt_num
                        ws_bs.cell(row=1, column=chart_start_col + 1).value = 'Equity'
                        ws_bs.cell(row=2, column=chart_start_col + 1).value = equity_num
                        
                        data = Reference(ws_bs, min_col=chart_start_col, min_row=2, max_col=chart_start_col + 1, max_row=2)
                        cat_ref = Reference(ws_bs, min_col=chart_start_col, min_row=1, max_col=chart_start_col + 1, max_row=1)
                        
                        pie_chart.add_data(data, titles_from_data=True)
                        pie_chart.set_categories(cat_ref)
                        pie_chart.height = 10
                        pie_chart.width = 12
                        
                        ws_bs.add_chart(pie_chart, f"{get_column_letter(chart_start_col)}{chart_start_row}")
                        logger.info("✓ Created Capital Structure pie chart")
                    else:
                        logger.debug("Skipping pie chart - no valid data")
                except Exception as e:
                    logger.warning(f"Failed to create Capital Structure pie chart: {e}", exc_info=True)
        except Exception as e:
            logger.warning(f"Failed to create Capital Structure charts: {e}", exc_info=True)
    
    def _find_sensitivity_table_location(self, ws_sens: openpyxl.worksheet.worksheet.Worksheet) -> Optional[Tuple[int, int]]:
        """Find sensitivity table header and data start row."""
        for row_idx in range(1, 50):
            cell_val = ws_sens.cell(row=row_idx, column=1).value
            if cell_val and "Exit Multiple" in str(cell_val):
                next_row_val = ws_sens.cell(row=row_idx + 1, column=1).value
                if next_row_val is not None:
                    if isinstance(next_row_val, (int, float)):
                        return (row_idx, row_idx + 1)
                    elif isinstance(next_row_val, str):
                        try:
                            float(next_row_val.replace('x', '').strip())
                            return (row_idx, row_idx + 1)
                        except (ValueError, AttributeError):
                            continue
        return None
    
    def _count_sensitivity_data_rows(self, ws_sens: openpyxl.worksheet.worksheet.Worksheet, 
                                    data_start_row: int) -> int:
        """Count number of data rows in sensitivity table."""
        data_rows = 0
        for row_idx in range(data_start_row, data_start_row + 10):
            if ws_sens.cell(row=row_idx, column=1).value:
                data_rows += 1
            else:
                break
        return data_rows
    
    def _setup_sensitivity_chart_references(self, ws_sens: openpyxl.worksheet.worksheet.Worksheet,
                                           data_start_row: int, data_end_row: int,
                                           title_col: int) -> Dict[str, Reference]:
        """Setup chart data references for sensitivity charts."""
        from openpyxl.chart import Reference
        
        categories = Reference(ws_sens, min_col=1, min_row=data_start_row, max_col=1, max_row=data_end_row)
        moic_data = Reference(ws_sens, min_col=3, min_row=data_start_row, max_col=3, max_row=data_end_row)
        irr_data = Reference(ws_sens, min_col=4, min_row=data_start_row, max_col=4, max_row=data_end_row)
        
        ws_sens.cell(row=data_start_row, column=title_col).value = "MOIC"
        moic_title_ref = Reference(ws_sens, min_col=title_col, min_row=data_start_row, 
                                  max_col=title_col, max_row=data_start_row)
        ws_sens.cell(row=data_start_row, column=title_col + 1).value = "IRR"
        irr_title_ref = Reference(ws_sens, min_col=title_col + 1, min_row=data_start_row, 
                                 max_col=title_col + 1, max_row=data_start_row)
        ws_sens.cell(row=data_start_row, column=title_col + 2).value = "IRR (%)"
        irr_combined_title_ref = Reference(ws_sens, min_col=title_col + 2, min_row=data_start_row, 
                                          max_col=title_col + 2, max_row=data_start_row)
        
        return {
            'categories': categories,
            'moic_data': moic_data,
            'irr_data': irr_data,
            'moic_title_ref': moic_title_ref,
            'irr_title_ref': irr_title_ref,
            'irr_combined_title_ref': irr_combined_title_ref
        }
    
    def _create_moic_sensitivity_chart(self, ws_sens: openpyxl.worksheet.worksheet.Worksheet,
                                      categories: Reference, moic_data: Reference,
                                      moic_title_ref: Reference) -> None:
        """Create MOIC sensitivity chart."""
        from openpyxl.chart import LineChart
        
        try:
            chart = LineChart()
            chart.title = "MOIC Sensitivity to Exit Multiple"
            chart.style = 10
            chart.legend = None
            chart.height = 8
            chart.width = 12
            self._configure_line_chart_axes(
                chart, 'MOIC (x)', 'Exit Multiple (x)',
                y_number_format='0.0"x"', x_number_format='0.0"x"', y_min=0
            )
            
            chart.set_categories(categories)
            chart.add_data(moic_data, titles_from_data=False)
            
            if len(chart.series) >= 1:
                s1 = chart.series[0]
                self._configure_chart_series(
                    s1, moic_title_ref, marker_symbol="circle",
                    marker_size=10, show_labels=True, label_format='0.00"x"'
                )
                s1.graphicalProperties.line.width = 25000
            
            ws_sens.add_chart(chart, "F3")
            logger.info("✓ Created MOIC Sensitivity chart")
        except (ValueError, TypeError, AttributeError, IndexError) as e:
            logger.warning(f"Failed to create MOIC Sensitivity chart - data validation error: {e}")
        except Exception as e:
            logger.warning(f"Failed to create MOIC Sensitivity chart: {e}", exc_info=True)
    
    def _create_irr_sensitivity_chart(self, ws_sens: openpyxl.worksheet.worksheet.Worksheet,
                                      categories: Reference, irr_data: Reference,
                                      irr_title_ref: Reference) -> None:
        """Create IRR sensitivity chart."""
        from openpyxl.chart import LineChart
        
        try:
            chart = LineChart()
            chart.title = "IRR Sensitivity to Exit Multiple"
            chart.style = 10
            chart.legend = None
            chart.height = 8
            chart.width = 12
            self._configure_line_chart_axes(
                chart, 'IRR (%)', 'Exit Multiple (x)',
                y_number_format='0.0%', x_number_format='0.0"x"', y_min=0
            )
            
            chart.set_categories(categories)
            chart.add_data(irr_data, titles_from_data=False)
            
            if len(chart.series) >= 1:
                s2 = chart.series[0]
                self._configure_chart_series(
                    s2, irr_title_ref, marker_symbol="diamond",
                    marker_size=10, show_labels=True, label_format='0.0"%'
                )
                s2.graphicalProperties.line.width = 25000
                s2.graphicalProperties.line.solidFill = "FF0000"
            
            ws_sens.add_chart(chart, "F18")
            logger.info("✓ Created IRR Sensitivity chart")
        except (ValueError, TypeError, AttributeError, IndexError) as e:
            logger.warning(f"Failed to create IRR Sensitivity chart - data validation error: {e}")
        except Exception as e:
            logger.warning(f"Failed to create IRR Sensitivity chart: {e}", exc_info=True)
    
    def _create_combined_sensitivity_chart(self, ws_sens: openpyxl.worksheet.worksheet.Worksheet,
                                           categories: Reference, moic_data: Reference,
                                           irr_data: Reference, moic_title_ref: Reference,
                                           irr_combined_title_ref: Reference) -> None:
        """Create combined MOIC and IRR sensitivity chart."""
        from openpyxl.chart import LineChart
        
        try:
            chart = LineChart()
            chart.title = "MOIC and IRR Sensitivity to Exit Multiple"
            chart.style = 10
            chart.legend.position = 'b'
            chart.height = 10
            chart.width = 16
            self._configure_line_chart_axes(
                chart, 'MOIC (x) / IRR (%)', 'Exit Multiple (x)',
                y_number_format='0.0', x_number_format='0.0"x"', y_min=0
            )
            
            chart.set_categories(categories)
            chart.add_data(moic_data, titles_from_data=False)
            chart.add_data(irr_data, titles_from_data=False)
            
            if len(chart.series) >= 1:
                self._configure_chart_series(
                    chart.series[0], moic_title_ref,
                    marker_symbol="circle", show_labels=False
                )
            
            if len(chart.series) >= 2:
                s4 = chart.series[1]
                self._configure_chart_series(
                    s4, irr_combined_title_ref,
                    marker_symbol="diamond", show_labels=False
                )
                s4.graphicalProperties.line.solidFill = "FF0000"
            
            ws_sens.add_chart(chart, "Q3")
            logger.info("✓ Created Combined MOIC/IRR Sensitivity chart")
        except (ValueError, TypeError, AttributeError, IndexError) as e:
            logger.warning(f"Failed to create Combined Sensitivity chart - data validation error: {e}")
        except Exception as e:
            logger.warning(f"Failed to create Combined Sensitivity chart: {e}", exc_info=True)
    
    def _create_sensitivity_charts(self, ws_sens: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Create Sensitivity Analysis charts on Sensitivity sheet.
        
        Args:
            ws_sens: Sensitivity Analysis worksheet
        """
        location = self._find_sensitivity_table_location(ws_sens)
        if not location:
            logger.debug("Skipping Sensitivity charts - header or data start row not found")
            return
        
        header_row, data_start_row = location
        data_rows = self._count_sensitivity_data_rows(ws_sens, data_start_row)
        
        if data_rows == 0:
            logger.debug("Skipping Sensitivity charts - no data rows found")
            return
        
        data_end_row = data_start_row + data_rows - 1
        title_col = 6
        
        try:
            refs = self._setup_sensitivity_chart_references(ws_sens, data_start_row, 
                                                           data_end_row, title_col)
            
            self._create_moic_sensitivity_chart(ws_sens, refs['categories'], 
                                                refs['moic_data'], refs['moic_title_ref'])
            self._create_irr_sensitivity_chart(ws_sens, refs['categories'], 
                                              refs['irr_data'], refs['irr_title_ref'])
            self._create_combined_sensitivity_chart(ws_sens, refs['categories'], 
                                                   refs['moic_data'], refs['irr_data'],
                                                   refs['moic_title_ref'], 
                                                   refs['irr_combined_title_ref'])
        except Exception as e:
            logger.warning(f"Failed to create Sensitivity Analysis charts: {e}", exc_info=True)
    
    def _add_charts(self) -> None:
        """Add professional charts to relevant sheets.
        
        Charts are created with error handling - if one chart fails, others will still be created.
        Errors are logged but do not stop the export process.
        """
        logger.info("Starting chart creation...")
        
        # Import chart improvement utilities
        try:
            from .lbo_chart_improvements import ChartStructureImprover
            chart_improver = ChartStructureImprover()
        except ImportError:
            chart_improver = None
            logger.debug("Chart improvement utilities not available")
        
        # Chart 1: Revenue and EBITDA Growth (Line Chart on Income Statement)
        if IndustryStandardTemplate.SHEET_INCOME_STATEMENT in self.wb.sheetnames:
            self._create_revenue_ebitda_chart(self.wb[IndustryStandardTemplate.SHEET_INCOME_STATEMENT])
        
        # Chart 2: Returns Analysis (Bar Chart on Returns sheet)
        if IndustryStandardTemplate.SHEET_RETURNS in self.wb.sheetnames:
            self._create_returns_chart(self.wb[IndustryStandardTemplate.SHEET_RETURNS])
        
        # Chart 3: Balance Sheet - Capital Structure Trends (Line Chart + Pie Chart on Balance Sheet)
        if IndustryStandardTemplate.SHEET_BALANCE_SHEET in self.wb.sheetnames:
            self._create_capital_structure_charts(self.wb[IndustryStandardTemplate.SHEET_BALANCE_SHEET])
        
        # Chart 4: Sensitivity Analysis - MOIC and IRR by Exit Multiple
        if IndustryStandardTemplate.SHEET_SENSITIVITY in self.wb.sheetnames:
            self._create_sensitivity_charts(self.wb[IndustryStandardTemplate.SHEET_SENSITIVITY])
        
        logger.info("Chart creation completed")
    
    def _add_navigation_hyperlinks(self):
        """Add hyperlinks for navigation from Cover sheet using 'Place in This Document' format.
        
        Maps text in cells A7-A15 to their corresponding sheets:
        - A7: "Summary" → Summary sheet
        - A8: "Sources & Uses" → Sources & Uses sheet
        - A9: "Assumptions" → Assumptions sheet
        - A10: "Income Statement" → Income Statement sheet
        - A11: "Cash Flow" → Cash Flow sheet
        - A12: "Balance Sheet" → Balance Sheet sheet
        - A13: "Debt Schedule" → Debt Schedule sheet
        - A14: "Returns Analysis" → Returns Analysis sheet
        - A15: "Sensitivity Analysis" → Sensitivity Analysis sheet
        """
        if IndustryStandardTemplate.SHEET_COVER not in self.wb.sheetnames:
            return
        
        ws_cover = self.wb[IndustryStandardTemplate.SHEET_COVER]
        row = 7
        
        # Use shared navigation items mapping to ensure consistency
        nav_items = self._get_navigation_items()
        
        for label, sheet_name in nav_items:
            if sheet_name in self.wb.sheetnames:
                cell = ws_cover.cell(row=row, column=1)
                # Verify cell text matches expected label before adding hyperlink
                if cell.value == label:
                    # Use "Place in This Document" format: #SheetName!A1
                    # Sheet names with spaces or special characters must be quoted: #'Sheet Name'!A1
                    # The # prefix tells Excel this is a "Place in This Document" hyperlink
                    if ' ' in sheet_name or '&' in sheet_name:
                        # Quote sheet names with spaces or special characters
                        hyperlink_target = f"#'{sheet_name}'!A1"
                    else:
                        # No quotes needed for sheet names without spaces
                        hyperlink_target = f"#{sheet_name}!A1"
                    cell.hyperlink = hyperlink_target
                    cell.font = Font(name=IndustryStandardTemplate.FONT_NAME, 
                                   size=IndustryStandardTemplate.FONT_SIZE_BODY,
                                   underline='single', color='0563C1')
                else:
                    # Log warning if text doesn't match expected label
                    logger.warning(
                        f"Cover sheet row {row} has text '{cell.value}' but expected '{label}'. "
                        f"Hyperlink not added."
                    )
                row += 1
    
    def _set_column_widths(self):
        """Set appropriate column widths."""
        for ws in self.wb.worksheets:
            ws.column_dimensions['A'].width = 30  # Label column
            for col in range(2, 12):
                ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _protect_formulas(self):
        """Protect formula cells from accidental changes."""
        for ws in self.wb.worksheets:
            # Unlock all cells first
            for row in ws.iter_rows():
                for cell in row:
                    cell.protection = openpyxl.styles.Protection(locked=False)
            
            # Lock cells with formulas (calculation cells)
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':  # Formula
                        cell.protection = openpyxl.styles.Protection(locked=True)

