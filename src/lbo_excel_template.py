"""
Excel Template Utilities for LBO Models
Provides formatting utilities for alternative Excel export format.
Note: Industry-standard format (default) uses IndustryStandardExcelExporter.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from typing import Dict, Optional, Any, List
import openpyxl


class LBOExcelTemplate:
    """Excel template utilities for alternative export format.

    Note: For industry-standard formatting, use IndustryStandardExcelExporter.
    This class provides utilities for the alternative export format with additional sheets.
    """

    # Color codes from AI recommendations
    COLOR_HEADER = "4F81BD"  # Blue
    COLOR_INPUT = "D9EAD3"  # Light green
    COLOR_OUTPUT = "F4CCCC"  # Light red/pink
    COLOR_TOTAL = "FFC000"  # Orange/amber for totals
    COLOR_BORDER = "A6A6A6"  # Gray

    # Fonts
    FONT_HEADER = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    FONT_SUBHEADER = Font(name="Arial", size=11, bold=True)
    FONT_BODY = Font(name="Arial", size=10)
    FONT_INPUT = Font(name="Arial", size=10, italic=True)
    FONT_TOTAL = Font(name="Arial", size=10, bold=True)

    # Fills
    FILL_HEADER = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    FILL_INPUT = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT, fill_type="solid")
    FILL_OUTPUT = PatternFill(start_color=COLOR_OUTPUT, end_color=COLOR_OUTPUT, fill_type="solid")
    FILL_TOTAL = PatternFill(start_color=COLOR_TOTAL, end_color=COLOR_TOTAL, fill_type="solid")

    # Borders
    BORDER_THICK = Border(
        left=Side(style="thick", color=COLOR_BORDER),
        right=Side(style="thick", color=COLOR_BORDER),
        top=Side(style="thick", color=COLOR_BORDER),
        bottom=Side(style="thick", color=COLOR_BORDER),
    )
    BORDER_THIN = Border(
        left=Side(style="thin", color=COLOR_BORDER),
        right=Side(style="thin", color=COLOR_BORDER),
        top=Side(style="thin", color=COLOR_BORDER),
        bottom=Side(style="thin", color=COLOR_BORDER),
    )

    @staticmethod
    def _get_or_create_sheet(
        wb: openpyxl.Workbook, sheet_name: str, index: Optional[int] = None
    ) -> openpyxl.worksheet.worksheet.Worksheet:
        """Get existing sheet or create new one."""
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name, index)
        ws.delete_rows(1, ws.max_row)
        return ws

    @staticmethod
    def _create_title_header(
        ws: openpyxl.worksheet.worksheet.Worksheet,
        title: str,
        merge_range: str,
        row: int = 1,
        font_size: int = 16,
    ) -> None:
        """Create title header with formatting."""
        ws.merge_cells(merge_range)
        cell = ws.cell(row=row, column=1)
        cell.value = title
        cell.font = Font(name="Arial", size=font_size, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = LBOExcelTemplate.FILL_HEADER
        cell.border = LBOExcelTemplate.BORDER_THICK

    @staticmethod
    def _create_section_header(
        ws: openpyxl.worksheet.worksheet.Worksheet, text: str, row: int
    ) -> None:
        """Create section header with formatting."""
        cell = ws.cell(row=row, column=1)
        cell.value = text
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell.fill = LBOExcelTemplate.FILL_HEADER

    @staticmethod
    def _add_label_value_row(
        ws: openpyxl.worksheet.worksheet.Worksheet, label: str, value: Any, row: int
    ) -> None:
        """Add a label-value row."""
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = LBOExcelTemplate.FONT_BODY
        ws.cell(row=row, column=2).value = value
        ws.cell(row=row, column=2).font = LBOExcelTemplate.FONT_BODY

    @staticmethod
    def _set_column_widths(
        ws: openpyxl.worksheet.worksheet.Worksheet, widths: Dict[str, float]
    ) -> None:
        """Set column widths."""
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    @staticmethod
    def create_cover_page(
        wb: openpyxl.Workbook, company_name: str, assumptions: Dict
    ) -> openpyxl.worksheet.worksheet.Worksheet:
        """Create professional cover page."""
        ws = LBOExcelTemplate._get_or_create_sheet(wb, "Cover Page", 0)

        LBOExcelTemplate._create_title_header(ws, "LEVERAGED BUYOUT MODEL", "A1:D1", 1, 16)

        ws.merge_cells("A2:D2")
        cell = ws["A2"]
        cell.value = company_name
        cell.font = Font(name="Arial", size=14, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        row = 4
        LBOExcelTemplate._create_section_header(ws, "KEY METRICS", row)
        row += 1

        metrics = [
            ("Entry EBITDA", f"${assumptions.get('entry_ebitda', 0):,.0f}"),
            ("Entry Multiple", f"{assumptions.get('entry_multiple', 0):.1f}x"),
            (
                "Enterprise Value",
                f"${assumptions.get('entry_ebitda', 0) * assumptions.get('entry_multiple', 0):,.0f}",
            ),
            ("Exit Year", assumptions.get("exit_year", 5)),
            ("Exit Multiple", f"{assumptions.get('exit_multiple', 0):.1f}x"),
        ]

        for label, value in metrics:
            LBOExcelTemplate._add_label_value_row(ws, label, value, row)
            row += 1

        LBOExcelTemplate._set_column_widths(ws, {"A": 20, "B": 20, "C": 20, "D": 20})
        return ws

    @staticmethod
    def create_executive_summary(
        wb: openpyxl.Workbook, model_data: Dict
    ) -> openpyxl.worksheet.worksheet.Worksheet:
        """Create executive summary sheet with key highlights."""
        ws = LBOExcelTemplate._get_or_create_sheet(wb, "Executive Summary", 1)

        LBOExcelTemplate._create_title_header(ws, "EXECUTIVE SUMMARY", "A1:D1", 1, 12)

        row = 3
        returns = model_data.get("returns", {})
        summary_items = [
            ("Investment Highlights", ""),
            ("", ""),
            ("Equity Invested", f"${returns.get('equity_invested', 0):,.0f}"),
            ("Exit Equity Value", f"${returns.get('exit_equity_value', 0):,.0f}"),
            ("MOIC", f"{returns.get('moic', 0):.2f}x"),
            ("IRR", f"{returns.get('irr', 0):.2f}%"),
            ("", ""),
            ("Transaction Details", ""),
            ("Entry EBITDA", f"${model_data.get('entry_ebitda', 0):,.0f}"),
            ("Entry Multiple", f"{model_data.get('entry_multiple', 0):.1f}x"),
            ("Exit EBITDA", f"${returns.get('exit_ebitda', 0):,.0f}"),
            ("Exit Multiple", f"{model_data.get('exit_multiple', 0):.1f}x"),
        ]

        for label, value in summary_items:
            if label:
                if label in ["Investment Highlights", "Transaction Details"]:
                    LBOExcelTemplate._create_section_header(ws, label, row)
                else:
                    ws.cell(row=row, column=1).value = label
                    ws.cell(row=row, column=1).font = LBOExcelTemplate.FONT_BODY
            if value:
                ws.cell(row=row, column=2).value = value
                ws.cell(row=row, column=2).font = LBOExcelTemplate.FONT_BODY
            row += 1

        LBOExcelTemplate._set_column_widths(ws, {"A": 25, "B": 20})
        return ws

    @staticmethod
    def _add_category_section(
        ws: openpyxl.worksheet.worksheet.Worksheet, category: str, items: List[tuple], row: int
    ) -> int:
        """Add a category section with header and items."""
        LBOExcelTemplate._create_section_header(ws, category, row)
        row += 1

        for label, value in items:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font = LBOExcelTemplate.FONT_BODY
            ws.cell(row=row, column=2).value = value
            ws.cell(row=row, column=2).font = LBOExcelTemplate.FONT_INPUT
            ws.cell(row=row, column=2).fill = LBOExcelTemplate.FILL_INPUT
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
            row += 1

        return row + 1  # Return next row with spacing

    @staticmethod
    def create_assumptions_sheet(
        wb: openpyxl.Workbook, assumptions: Dict
    ) -> openpyxl.worksheet.worksheet.Worksheet:
        """Create assumptions sheet with all model inputs."""
        ws = LBOExcelTemplate._get_or_create_sheet(wb, "Assumptions")

        LBOExcelTemplate._create_title_header(ws, "MODEL ASSUMPTIONS", "A1:B1", 1, 12)

        row = 3
        categories = {
            "Transaction Assumptions": [
                ("Entry EBITDA", assumptions.get("entry_ebitda", 0)),
                ("Entry Multiple", assumptions.get("entry_multiple", 0)),
                ("Existing Debt", assumptions.get("existing_debt", 0)),
                ("Existing Cash", assumptions.get("existing_cash", 0)),
            ],
            "Operating Assumptions": [
                (
                    "Revenue Growth Rate (Year 1)",
                    f"{assumptions.get('revenue_growth_rate', [0])[0]*100:.1f}%",
                ),
                ("COGS % of Revenue", f"{assumptions.get('cogs_pct_of_revenue', 0)*100:.1f}%"),
                ("SG&A % of Revenue", f"{assumptions.get('sganda_pct_of_revenue', 0)*100:.1f}%"),
                ("CapEx % of Revenue", f"{assumptions.get('capex_pct_of_revenue', 0)*100:.1f}%"),
                ("Tax Rate", f"{assumptions.get('tax_rate', 0)*100:.1f}%"),
            ],
            "Working Capital Assumptions": [
                ("Days Sales Outstanding", assumptions.get("days_sales_outstanding", 0)),
                ("Days Inventory Outstanding", assumptions.get("days_inventory_outstanding", 0)),
                ("Days Payable Outstanding", assumptions.get("days_payable_outstanding", 0)),
            ],
            "Exit Assumptions": [
                ("Exit Year", assumptions.get("exit_year", 5)),
                ("Exit Multiple", assumptions.get("exit_multiple", 0)),
            ],
        }

        for category, items in categories.items():
            row = LBOExcelTemplate._add_category_section(ws, category, items, row)

        LBOExcelTemplate._set_column_widths(ws, {"A": 35, "B": 20})
        return ws

    @staticmethod
    def format_header_cell(
        cell: openpyxl.cell.cell.Cell, text: str, merge_range: Optional[str] = None
    ) -> openpyxl.cell.cell.Cell:
        """Format a header cell with professional styling."""
        if merge_range:
            cell.parent.merge_cells(merge_range)

        cell.value = text
        cell.font = LBOExcelTemplate.FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = LBOExcelTemplate.FILL_HEADER
        cell.border = LBOExcelTemplate.BORDER_THICK
        return cell

    @staticmethod
    def format_input_cell(cell: openpyxl.cell.cell.Cell, value: Any) -> openpyxl.cell.cell.Cell:
        """Format an input cell with green background."""
        cell.value = value
        cell.font = LBOExcelTemplate.FONT_INPUT
        cell.fill = LBOExcelTemplate.FILL_INPUT
        cell.border = LBOExcelTemplate.BORDER_THIN
        cell.alignment = Alignment(horizontal="right", vertical="center")

    @staticmethod
    def format_output_cell(cell: openpyxl.cell.cell.Cell, value: Any) -> openpyxl.cell.cell.Cell:
        """Format an output cell with light red background."""
        cell.value = value
        cell.font = LBOExcelTemplate.FONT_BODY
        cell.fill = LBOExcelTemplate.FILL_OUTPUT
        cell.border = LBOExcelTemplate.BORDER_THIN
        cell.alignment = Alignment(horizontal="right", vertical="center")
        return cell

    @staticmethod
    def format_total_cell(cell: openpyxl.cell.cell.Cell, value: Any) -> openpyxl.cell.cell.Cell:
        """Format a total cell with orange background and bold font."""
        cell.value = value
        cell.font = LBOExcelTemplate.FONT_TOTAL
        cell.fill = LBOExcelTemplate.FILL_TOTAL
        cell.border = LBOExcelTemplate.BORDER_THICK
        cell.alignment = Alignment(horizontal="right", vertical="center")
        return cell

    @staticmethod
    def format_section_header(cell: openpyxl.cell.cell.Cell, text: str) -> openpyxl.cell.cell.Cell:
        """Format a section header."""
        cell.value = text
        cell.font = LBOExcelTemplate.FONT_SUBHEADER
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        cell.border = LBOExcelTemplate.BORDER_THIN
        return cell
