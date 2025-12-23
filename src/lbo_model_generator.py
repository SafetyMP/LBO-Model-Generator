"""
LBO Model Generator

A comprehensive tool to generate Leveraged Buyout (LBO) models from user inputs.
Generates Income Statement, Balance Sheet, Cash Flow Statement, Debt Schedule, and Returns Analysis.
"""

import pandas as pd
import numpy as np
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import json
import logging

# Handle both package and direct imports
try:
    from .lbo_constants import LBOConstants
    from .lbo_exceptions import LBOValidationError, LBOCalculationError, LBOAIServiceError, LBOConfigurationError
    from .lbo_industry_standards import IndustryStandardTemplate
    from .lbo_industry_excel import IndustryStandardExcelExporter
except ImportError:
    from lbo_constants import LBOConstants
    from lbo_exceptions import LBOValidationError, LBOCalculationError, LBOAIServiceError, LBOConfigurationError
    from lbo_industry_standards import IndustryStandardTemplate
    from lbo_industry_excel import IndustryStandardExcelExporter

# Configure logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)


# Excel layout constants
class ExcelConstants:
    """Constants for Excel sheet layout matching reference model."""

    # Row indices
    SOURCES_SECTION_START_ROW = 5
    INCOME_STATEMENT_START_ROW = 22
    BALANCE_SHEET_START_ROW = 57
    CASH_FLOW_START_ROW = 110
    DEBT_SCHEDULE_START_ROW = 147
    MAX_ROW_SCAN = 300

    # Column indices
    HISTORICAL_COL = 4  # Column D
    TRANSACTION_COL = 5  # Column E
    PROJECTED_START_COL = 6  # Column F
    SOURCES_PCT_COL = 7  # Column G
    USES_LABEL_COL = 8  # Column H
    USES_VALUE_COL = 9  # Column I
    USES_PCT_COL = 10  # Column J

    # P&L Sheet references
    PL_REVENUE_ROW = 8
    PL_COGS_ROW = 24
    PL_EBITDA_ROW = 84
    PL_HISTORICAL_YEARS = {"2017": 2, "2018": 3, "2019": 4, "2020": 5}  # Columns B-E
    PL_LTM_COL = 11  # Column K


@dataclass
class LBODebtStructure:
    """Structure for debt financing in LBO."""

    name: str
    amount: float
    interest_rate: float
    ebitda_multiple: Optional[float] = None
    amortization_schedule: str = "bullet"  # "bullet", "amortizing", "cash_flow_sweep"
    amortization_periods: int = 5  # For amortizing debt
    priority: int = 1  # 1 = senior, 2 = subordinated, etc.


@dataclass
class LBOAssumptions:
    """Core assumptions for LBO model."""

    # Transaction details
    entry_ebitda: float
    entry_multiple: float
    existing_debt: float = 0.0
    existing_cash: float = 0.0
    transaction_expenses_pct: float = 0.03  # % of EV
    financing_fees_pct: float = 0.02  # % of total debt

    # Debt structure
    debt_instruments: List[LBODebtStructure] = field(default_factory=list)
    equity_amount: float = 0.0

    # Operating assumptions
    revenue_growth_rate: List[float] = field(default_factory=lambda: [0.05] * 5)  # 5 years default
    cogs_pct_of_revenue: float = 0.70
    sganda_pct_of_revenue: float = 0.15
    depreciation_pct_of_ppe: float = 0.10
    capex_pct_of_revenue: float = 0.03
    tax_rate: float = 0.25

    # Working capital assumptions
    days_sales_outstanding: float = 45.0
    days_inventory_outstanding: float = 30.0
    days_payable_outstanding: float = 30.0

    # Balance sheet assumptions
    initial_ppe: float = 0.0
    initial_ar: float = 0.0
    initial_inventory: float = 0.0
    initial_ap: float = 0.0
    min_cash_balance: float = 0.0

    # Exit assumptions
    exit_year: int = 5
    exit_multiple: float = 7.5
    target_exit_debt: float = 0.0  # Target debt level at exit (0 = pay down all debt, >0 = maintain strategic debt)
    max_debt_paydown_per_year: float = 0.0  # Maximum annual debt paydown (0 = unlimited)
    fcf_conversion_rate: float = 0.0  # FCF conversion rate (0 = calculate from CFO - CapEx, >0 = EBITDA * rate)

    # Starting revenue (for projections)
    starting_revenue: float = 0.0

    def __post_init__(self):
        """Validate assumptions after initialization."""
        errors = []

        if self.entry_ebitda <= 0:
            errors.append("entry_ebitda must be positive")

        if self.entry_multiple <= 0:
            errors.append("entry_multiple must be positive")

        if len(self.revenue_growth_rate) == 0:
            errors.append("revenue_growth_rate cannot be empty")

        # Validate revenue growth rate length matches expected projection period
        # Industry standard: growth rates should be provided for each projection year
        expected_years = len(self.revenue_growth_rate)
        if expected_years < 1:
            errors.append("revenue_growth_rate must have at least one year")
        elif expected_years > 10:
            errors.append(
                f"revenue_growth_rate has {expected_years} years, which exceeds typical LBO projection period (typically 5-7 years)"
            )

        if not (0 <= self.cogs_pct_of_revenue <= 1):
            errors.append("cogs_pct_of_revenue must be between 0 and 1")

        if not (0 <= self.sganda_pct_of_revenue <= 1):
            errors.append("sganda_pct_of_revenue must be between 0 and 1")

        if not (0 <= self.tax_rate <= 1):
            errors.append("tax_rate must be between 0 and 1")

        if not (0 <= self.transaction_expenses_pct <= 1):
            errors.append("transaction_expenses_pct must be between 0 and 1")

        if not (0 <= self.financing_fees_pct <= 1):
            errors.append("financing_fees_pct must be between 0 and 1")

        if self.exit_year <= 0:
            errors.append("exit_year must be positive")

        if self.exit_multiple <= 0:
            errors.append("exit_multiple must be positive")

        if self.days_sales_outstanding < 0 or self.days_sales_outstanding > 365:
            errors.append("days_sales_outstanding must be between 0 and 365")

        if self.days_inventory_outstanding < 0 or self.days_inventory_outstanding > 365:
            errors.append("days_inventory_outstanding must be between 0 and 365")

        if self.days_payable_outstanding < 0 or self.days_payable_outstanding > 365:
            errors.append("days_payable_outstanding must be between 0 and 365")

        if errors:
            raise ValueError("LBOAssumptions validation failed:\n" + "\n".join(f"  - {e}" for e in errors))


class ExcelRowFinder:
    """Helper class for finding rows in Excel worksheets."""

    @staticmethod
    def find_row_by_label(
        ws: openpyxl.worksheet.worksheet.Worksheet,
        label: str,
        start_row: int = 1,
        end_row: int = None,
        column: int = 1,
        partial_match: bool = False,
    ) -> Optional[int]:
        """Find row index containing label in specified column.

        Args:
            ws: Excel worksheet to search
            label: Label text to find
            start_row: Starting row index (1-based)
            end_row: Ending row index (1-based, None = use MAX_ROW_SCAN)
            column: Column number to search (default: 1 = column A)
            partial_match: If True, match if label is contained in cell value

        Returns:
            Row index (1-based) if found, None otherwise
        """
        if end_row is None:
            end_row = ExcelConstants.MAX_ROW_SCAN

        for row_idx in range(start_row, min(end_row, ws.max_row + 1)):
            cell_val = ws.cell(row=row_idx, column=column).value
            if cell_val:
                if partial_match:
                    if label in str(cell_val):
                        return row_idx
                else:
                    if str(cell_val).strip() == label.strip():
                        return row_idx
        return None

    @staticmethod
    def build_row_map(
        ws: openpyxl.worksheet.worksheet.Worksheet, start_row: int = 1, end_row: int = None, column: int = 1
    ) -> Dict[str, int]:
        """Build a map of all row labels to row indices.

        Args:
            ws: Excel worksheet to scan
            start_row: Starting row index
            end_row: Ending row index
            column: Column to scan

        Returns:
            Dictionary mapping label strings to row indices
        """
        if end_row is None:
            end_row = ExcelConstants.MAX_ROW_SCAN

        row_map = {}
        for row_idx in range(start_row, min(end_row, ws.max_row + 1)):
            cell_val = ws.cell(row=row_idx, column=column).value
            if cell_val and isinstance(cell_val, str):
                # Store both exact match and common variations
                row_map[cell_val.strip()] = row_idx
                # Also store common variations
                if cell_val.strip().lower() in ["revenue", "cogs", "ebitda", "ebit"]:
                    row_map[cell_val.strip().lower()] = row_idx
        return row_map

    @staticmethod
    def find_rows_by_patterns(
        ws: openpyxl.worksheet.worksheet.Worksheet,
        patterns: Dict[str, List[str]],
        start_row: int = 1,
        end_row: int = None,
        column: int = 1,
    ) -> Dict[str, Optional[int]]:
        """Find multiple rows using pattern matching.

        Args:
            ws: Excel worksheet to search
            patterns: Dictionary mapping keys to lists of possible labels
            start_row: Starting row index
            end_row: Ending row index
            column: Column to search

        Returns:
            Dictionary mapping keys to row indices (or None if not found)
        """
        if end_row is None:
            end_row = ExcelConstants.MAX_ROW_SCAN

        result = {}
        for key, labels in patterns.items():
            result[key] = None
            for row_idx in range(start_row, min(end_row, ws.max_row + 1)):
                cell_val = ws.cell(row=row_idx, column=column).value
                if cell_val:
                    cell_str = str(cell_val).strip()
                    for label in labels:
                        if label in cell_str or cell_str == label:
                            result[key] = row_idx
                            break
                if result[key]:
                    break
        return result


class LBOModel:
    """Main LBO model class that generates all financial statements.

    This class generates comprehensive LBO financial models including:
    - Income Statement with multi-year projections
    - Balance Sheet with working capital calculations
    - Cash Flow Statement
    - Debt Schedule with multiple debt instruments
    - Returns Analysis (IRR and MOIC)

    Example:
        >>> assumptions = LBOAssumptions(
        ...     entry_ebitda=10000,
        ...     entry_multiple=6.5,
        ...     revenue_growth_rate=[0.05, 0.05, 0.05, 0.05, 0.05]
        ... )
        >>> model = LBOModel(assumptions)
        >>> model.export_to_excel("output.xlsx")
    """

    @staticmethod
    def _round_value(value: float) -> float:
        """Round financial value to 2 decimal places."""
        try:
            return round(float(value), LBOConstants.DECIMAL_PLACES)
        except (ValueError, TypeError):
            # Fallback if value is not numeric
            try:
                return round(float(value), 2)
            except (ValueError, TypeError):
                return value

    def __init__(self, assumptions: LBOAssumptions):
        """Initialize LBO model with assumptions.

        Args:
            assumptions: LBOAssumptions object with all model parameters

        Raises:
            ValueError: If assumptions fail validation
        """
        logger.info("Initializing LBO model")
        self.assumptions = assumptions
        self.num_years = len(assumptions.revenue_growth_rate)
        self.years = list(range(1, self.num_years + 1))

        try:
            # Calculate transaction values
            self._calculate_transaction_values()

            # Initialize financial statements
            self.income_statement = pd.DataFrame(index=self._get_is_line_items(), columns=self.years)
            self.balance_sheet = pd.DataFrame(index=self._get_bs_line_items(), columns=self.years)
            self.cash_flow = pd.DataFrame(index=self._get_cf_line_items(), columns=self.years)
            self.debt_schedule = {}

            # Build model
            self._build_model()
            logger.info("LBO model initialized successfully")
        except (LBOValidationError, LBOCalculationError) as e:
            logger.error(f"Error initializing LBO model: {e}")
            raise
        except Exception as e:
            logger.error(f"Unexpected error initializing LBO model: {e}", exc_info=True)
            raise LBOCalculationError(f"Failed to initialize LBO model: {e}") from e

    def _calculate_transaction_values(self) -> None:
        """Calculate transaction-related values following industry standards.

        Industry Standard Calculations:
        - Enterprise Value (EV) = Entry EBITDA × Entry Multiple
        - Equity Value (Purchase Price) = EV - Existing Debt + Existing Cash
          (Cash reduces net debt, thus increases equity value)
        - Transaction Expenses = EV × Transaction Expenses %
        - Financing Fees = Total Debt × Financing Fees %
        - Total Uses = Equity Value + Existing Debt Repayment + Transaction Expenses + Financing Fees
        - Total Sources = Equity Contribution + Total Debt
        - Sources must equal Uses (industry requirement)

        Calculates:
        - Enterprise value (EV = EBITDA × Multiple)
        - Equity value (EV - existing debt + existing cash) - industry standard formula
        - Total sources and uses (must balance)
        - Individual debt instrument amounts (if using EBITDA multiples)
        """
        logger.debug("Calculating transaction values")

        # Industry standard: EV = Entry EBITDA × Entry Multiple
        self.enterprise_value = self.assumptions.entry_ebitda * self.assumptions.entry_multiple

        # Industry standard: Equity Value = EV - Existing Debt + Existing Cash
        # Cash reduces net debt, effectively increasing the equity purchase price
        self.equity_value = self.enterprise_value - self.assumptions.existing_debt + self.assumptions.existing_cash

        # Calculate debt amounts if using EBITDA multiples
        total_debt_calc = 0.0
        for debt in self.assumptions.debt_instruments:
            if debt.ebitda_multiple:
                debt.amount = debt.ebitda_multiple * self.assumptions.entry_ebitda
            total_debt_calc += debt.amount

        # Calculate transaction expenses and financing fees first
        self.transaction_expenses = self._round_value(self.enterprise_value * self.assumptions.transaction_expenses_pct)
        total_debt = sum(d.amount for d in self.assumptions.debt_instruments)
        self.financing_fees = self._round_value(total_debt * self.assumptions.financing_fees_pct)

        # Calculate total uses
        total_uses_calc = self._round_value(
            self.equity_value + self.assumptions.existing_debt + self.transaction_expenses + self.financing_fees
        )

        # If equity not specified, calculate it to balance sources and uses
        if self.assumptions.equity_amount == 0.0 or self.assumptions.equity_amount is None:
            # Equity needed = Total Uses - Total Debt
            calculated_equity = self._round_value(max(0, total_uses_calc - total_debt))

            # Validate calculated equity
            if calculated_equity < 0:
                logger.warning(
                    f"Calculated equity is negative: ${calculated_equity:,.0f}. "
                    f"This may indicate insufficient debt capacity or incorrect assumptions."
                )
            elif calculated_equity > 0 and self.equity_value > 0:
                equity_ratio = calculated_equity / self.equity_value
                if equity_ratio < LBOConstants.MIN_EQUITY_TO_EV_RATIO:
                    logger.warning(
                        f"Equity is very small relative to EV: {equity_ratio:.1%}. "
                        f"Calculated equity: ${calculated_equity:,.0f}, EV: ${self.equity_value:,.0f}. "
                        f"This may indicate an aggressive debt structure."
                    )

            self.assumptions.equity_amount = calculated_equity

        # Calculate sources and uses (should now balance)
        self.total_sources = self._round_value(self.assumptions.equity_amount + total_debt)
        self.total_uses = self._round_value(
            self.equity_value + self.assumptions.existing_debt + self.transaction_expenses + self.financing_fees
        )

        # Validate balance
        balance_diff = abs(self.total_sources - self.total_uses)
        if balance_diff > 0.01:
            logger.warning(
                f"Sources and Uses do not balance: Sources=${self.total_sources:,.0f}, "
                f"Uses=${self.total_uses:,.0f}, Difference=${balance_diff:,.0f}"
            )

        # Calculate goodwill (industry standard: Purchase Price - Fair Value of Net Identifiable Assets)
        # Goodwill represents the premium paid over the fair value of net assets
        net_book_value = (
            self.assumptions.initial_ppe
            + self.assumptions.initial_ar
            + self.assumptions.initial_inventory
            - self.assumptions.initial_ap
        )
        # Purchase price is the equity value paid
        self.goodwill = max(0, self.equity_value - net_book_value)

        if self.goodwill < 0:
            logger.debug(
                f"Negative goodwill calculated: ${self.goodwill:,.0f}. This may indicate net book value exceeds purchase price."
            )

    def _get_is_line_items(self) -> List[str]:
        """Income statement line items (detailed format matching reference model)."""
        return [
            "Revenue",
            "Growth (%)",
            "Cost of Goods Sold (Net of D&A)",
            "% of Sales",
            "Gross Profit",
            "% of Sales",
            "SG&A (Net of D&A)",
            "% of Sales",
            "EBITDA",
            "% of Sales",
            "Depreciation",
            "Amortization",
            "EBIT",
            "Interest Expense",
            "Pretax Income",
            "Income Tax Expense",
            "Tax Rate",
            "Net Income",
        ]

    def _get_bs_line_items(self) -> List[str]:
        """Balance sheet line items."""
        return [
            "Cash",
            "Accounts Receivable",
            "Inventory",
            "Total Current Assets",
            "PP&E, Net",
            "Goodwill",
            "Intangible Assets (Financing Fees)",
            "Total Assets",
            "Accounts Payable",
            "Total Current Liabilities",
            "Total Debt",
            "Total Liabilities",
            "Shareholders Equity",
            "Total Liabilities & Equity",
        ]

    def _get_cf_line_items(self) -> List[str]:
        """Cash flow statement line items."""
        return [
            "Net Income",
            "Depreciation & Amortization",
            "Change in Accounts Receivable",
            "Change in Inventory",
            "Change in Accounts Payable",
            "Net Change in Working Capital",
            "Cash Flow from Operations",
            "Capital Expenditures",
            "Cash Flow from Investing",
            "Debt Repayment",
            "Debt Issuance",
            "Equity Contribution",
            "Purchase Price (Equity Value)",
            "Existing Debt Repayment",
            "Transaction Expenses",
            "Financing Fees",
            "Cash Flow from Financing",
            "Net Change in Cash",
            "Beginning Cash Balance",
            "Ending Cash Balance",
        ]

    def _build_model(self) -> None:
        """Build complete LBO model."""
        self._build_income_statement()
        self._build_debt_schedule()  # Must be before balance sheet
        self._build_balance_sheet()
        self._build_cash_flow_statement()
        self._apply_cash_flow_sweep()  # Apply cash flow sweep to use excess cash for debt paydown
        self._reconcile_model()

    def _build_income_statement(self) -> None:
        """Build income statement projections following industry-standard order.

        Calculation order (industry standard):
        1. Revenue (with growth rates)
        2. COGS = Revenue × COGS%
        3. Gross Profit = Revenue - COGS
        4. SG&A = Revenue × SG&A%
        5. EBITDA = Gross Profit - SG&A (before D&A)
        6. Depreciation & Amortization
        7. EBIT = EBITDA - D&A
        8. Interest Expense (from debt schedule)
        9. Pretax Income = EBIT - Interest
        10. Income Tax = Pretax × Tax Rate
        11. Net Income = Pretax - Tax
        """
        # Step 1: Revenue (with growth rates)
        if self.assumptions.starting_revenue == 0.0:
            # Estimate from EBITDA margin
            ebitda_margin = 1 - self.assumptions.cogs_pct_of_revenue - self.assumptions.sganda_pct_of_revenue
            self.assumptions.starting_revenue = self.assumptions.entry_ebitda / ebitda_margin

        prev_revenue = self.assumptions.starting_revenue
        for year in self.years:
            if year == 1:
                # Base case revenue (or use starting revenue)
                revenue = prev_revenue
                growth_rate = 0.0
            else:
                # Use growth rate for year (index year - 1 for 0-based indexing)
                # Handle case where growth rate list may be shorter than years
                growth_rate_idx = min(year - 2, len(self.assumptions.revenue_growth_rate) - 1)
                growth_rate = self.assumptions.revenue_growth_rate[growth_rate_idx]
                revenue = prev_revenue * (1 + growth_rate)

            self.income_statement.loc["Revenue", year] = self._round_value(revenue)
            self.income_statement.loc["Growth (%)", year] = self._round_value(growth_rate)
            prev_revenue = revenue

        # Step 2 & 3: COGS and Gross Profit
        for year in self.years:
            rev = self.income_statement.loc["Revenue", year]
            cogs = self._round_value(rev * self.assumptions.cogs_pct_of_revenue)
            self.income_statement.loc["Cost of Goods Sold (Net of D&A)", year] = cogs

            gp = self._round_value(rev - cogs)
            self.income_statement.loc["Gross Profit", year] = gp

        # Step 4: SG&A
        for year in self.years:
            rev = self.income_statement.loc["Revenue", year]
            sganda = self._round_value(rev * self.assumptions.sganda_pct_of_revenue)
            self.income_statement.loc["SG&A (Net of D&A)", year] = sganda

        # Step 5: EBITDA (before D&A)
        for year in self.years:
            gp = self.income_statement.loc["Gross Profit", year]
            sganda = self.income_statement.loc["SG&A (Net of D&A)", year]
            ebitda = self._round_value(gp - sganda)
            self.income_statement.loc["EBITDA", year] = ebitda

        # Calculate % of Sales rows (second pass after all base items)
        for year in self.years:
            rev = self.income_statement.loc["Revenue", year]
            if rev > 0:
                # Map each item to its corresponding % of Sales row index
                items_map = [
                    ("Cost of Goods Sold (Net of D&A)", 0),
                    ("Gross Profit", 1),
                    ("SG&A (Net of D&A)", 2),
                    ("EBITDA", 3),
                ]

                pct_rows = [idx for idx, item in enumerate(self.income_statement.index) if item == "% of Sales"]

                for item_name, pct_idx in items_map:
                    try:
                        item_idx = list(self.income_statement.index).index(item_name)
                        if pct_idx < len(pct_rows) and pct_rows[pct_idx] == item_idx + 1:
                            value = self.income_statement.loc[item_name, year]
                            self.income_statement.iloc[pct_rows[pct_idx], year - 1] = value / rev
                    except (ValueError, IndexError):
                        continue

        # D&A (separate Depreciation and Amortization)
        if self.assumptions.initial_ppe > 0:
            annual_depreciation = self.assumptions.initial_ppe * self.assumptions.depreciation_pct_of_ppe
            annual_amortization = self.financing_fees / self.num_years
        else:
            # Estimate from revenue using constants
            annual_depreciation = self.assumptions.starting_revenue * LBOConstants.DEFAULT_DEPRECIATION_TO_REVENUE_RATIO
            annual_amortization = (
                self.financing_fees / self.num_years
                if self.financing_fees > 0
                else self.assumptions.starting_revenue * LBOConstants.DEFAULT_AMORTIZATION_TO_REVENUE_RATIO
            )

        for year in self.years:
            self.income_statement.loc["Depreciation", year] = self._round_value(annual_depreciation)
            self.income_statement.loc["Amortization", year] = self._round_value(annual_amortization)

        # EBIT
        for year in self.years:
            ebitda = self.income_statement.loc["EBITDA", year]
            depreciation = self.income_statement.loc["Depreciation", year]
            amortization = self.income_statement.loc["Amortization", year]
            self.income_statement.loc["EBIT", year] = self._round_value(ebitda - depreciation - amortization)

        # Interest Expense (calculated from debt schedule later)
        for year in self.years:
            self.income_statement.loc["Interest Expense", year] = 0.0  # Will be updated

    def _build_debt_schedule(self) -> None:
        """Build debt repayment schedule."""
        # Initialize debt balances
        for debt in self.assumptions.debt_instruments:
            self.debt_schedule[debt.name] = {
                "beginning_balance": [],
                "interest_paid": [],
                "principal_paid": [],
                "ending_balance": [],
            }

            balance = debt.amount
            for year in self.years:
                self.debt_schedule[debt.name]["beginning_balance"].append(balance)

                # Interest
                interest = balance * debt.interest_rate
                self.debt_schedule[debt.name]["interest_paid"].append(interest)

                # Principal payment
                principal_paid = 0.0
                if debt.amortization_schedule == "amortizing":
                    # Annual amortization
                    scheduled_amount = self._round_value(debt.amount / debt.amortization_periods)

                    # If target_exit_debt is specified, adjust scheduled amortization
                    # to maintain target debt level
                    if self.assumptions.target_exit_debt > 0.01:
                        # Calculate total debt paydown needed to reach target
                        # Total debt = sum of all debt instruments' beginning balances for this year
                        # We need to calculate this by summing beginning balances from all debt instruments
                        current_total_debt = 0.0
                        for d in self.assumptions.debt_instruments:
                            if d.name in self.debt_schedule:
                                # Use beginning balance for this year if available
                                if year - 1 < len(self.debt_schedule[d.name]["beginning_balance"]):
                                    current_total_debt += self.debt_schedule[d.name]["beginning_balance"][year - 1]
                                else:
                                    # If not yet calculated, use the debt amount
                                    current_total_debt += d.amount
                            else:
                                # If debt schedule not yet initialized, use debt amount
                                current_total_debt += d.amount

                        remaining_paydown_needed = current_total_debt - self.assumptions.target_exit_debt
                        # Limit scheduled principal to not exceed remaining paydown needed
                        if remaining_paydown_needed > 0.01:
                            # Distribute remaining paydown across remaining years
                            years_remaining = debt.amortization_periods - (year - 1)
                            if years_remaining > 0:
                                max_principal_per_year = remaining_paydown_needed / years_remaining
                                scheduled_amount = min(scheduled_amount, max_principal_per_year)

                    # CRITICAL: Principal paid cannot exceed beginning balance
                    principal_paid = min(scheduled_amount, balance)
                    principal_paid = self._round_value(principal_paid)
                elif debt.amortization_schedule == "bullet":
                    # No payment until final year
                    if year == self.num_years:
                        principal_paid = self._round_value(balance)
                # cash_flow_sweep handled separately

                balance = self._round_value(balance - principal_paid)
                self.debt_schedule[debt.name]["principal_paid"].append(principal_paid)
                self.debt_schedule[debt.name]["ending_balance"].append(balance)

        # Update interest expense in income statement
        for year in self.years:
            total_interest = self._calculate_debt_interest(year)
            self.income_statement.loc["Interest Expense", year] = self._round_value(total_interest)

        # Calculate remaining IS items
        for year in self.years:
            self._update_income_statement_from_ebit(year)

    def _build_balance_sheet(self) -> None:
        """Build balance sheet projections."""
        # Initial balances (transaction adjustments)
        initial_cash = self.assumptions.existing_cash
        initial_ar = (
            self.assumptions.initial_ar
            if self.assumptions.initial_ar > 0
            else (self.assumptions.starting_revenue * self.assumptions.days_sales_outstanding / 365)
        )
        initial_inv = (
            self.assumptions.initial_inventory
            if self.assumptions.initial_inventory > 0
            else (
                self.assumptions.starting_revenue
                * self.assumptions.cogs_pct_of_revenue
                * self.assumptions.days_inventory_outstanding
                / 365
            )
        )
        # Calculate initial_ap if needed (used in net_book_value calculation)
        _ = (
            self.assumptions.initial_ap
            if self.assumptions.initial_ap > 0
            else (
                self.assumptions.starting_revenue
                * self.assumptions.cogs_pct_of_revenue
                * self.assumptions.days_payable_outstanding
                / 365
            )
        )
        initial_ppe = (
            self.assumptions.initial_ppe
            if self.assumptions.initial_ppe > 0
            else (self.assumptions.starting_revenue * LBOConstants.DEFAULT_PPE_TO_REVENUE_RATIO)
        )

        # Year 1 (post-transaction)
        self.balance_sheet.loc["Cash", 1] = self._round_value(initial_cash)
        self.balance_sheet.loc["Accounts Receivable", 1] = self._round_value(initial_ar)
        self.balance_sheet.loc["Inventory", 1] = self._round_value(initial_inv)
        self.balance_sheet.loc["Total Current Assets", 1] = self._round_value(initial_cash + initial_ar + initial_inv)

        # PP&E (reduced by depreciation each year)
        cumulative_depreciation = 0.0
        for year in self.years:
            depreciation = self.income_statement.loc["Depreciation", year]
            ppe_net = self._round_value(initial_ppe - cumulative_depreciation)
            self.balance_sheet.loc["PP&E, Net", year] = ppe_net
            cumulative_depreciation += depreciation

        # Goodwill
        for year in self.years:
            self.balance_sheet.loc["Goodwill", year] = self._round_value(self.goodwill)

        # Financing fees (amortized)
        remaining_fees = self.financing_fees
        for year in self.years:
            amortization = self.financing_fees / self.num_years
            if year == 1:
                self.balance_sheet.loc["Intangible Assets (Financing Fees)", year] = self._round_value(remaining_fees)
            else:
                remaining_fees -= amortization
                self.balance_sheet.loc["Intangible Assets (Financing Fees)", year] = self._round_value(remaining_fees)

        # Accounts Receivable and Inventory (update each year)
        for year in self.years:
            revenue = self.income_statement.loc["Revenue", year]
            self.balance_sheet.loc["Accounts Receivable", year] = self._round_value(
                revenue * self.assumptions.days_sales_outstanding / 365
            )
            self.balance_sheet.loc["Inventory", year] = self._round_value(
                revenue * self.assumptions.cogs_pct_of_revenue * self.assumptions.days_inventory_outstanding / 365
            )

        # Accounts Payable
        for year in self.years:
            revenue = self.income_statement.loc["Revenue", year]
            self.balance_sheet.loc["Accounts Payable", year] = self._round_value(
                revenue * self.assumptions.cogs_pct_of_revenue * self.assumptions.days_payable_outstanding / 365
            )
            self.balance_sheet.loc["Total Current Liabilities", year] = self._round_value(
                self.balance_sheet.loc["Accounts Payable", year]
            )

        # Total Debt
        for year in self.years:
            total_debt = self._calculate_total_debt(year)
            self.balance_sheet.loc["Total Debt", year] = self._round_value(total_debt)

        # Liabilities
        for year in self.years:
            self.balance_sheet.loc["Total Liabilities", year] = self._round_value(
                self.balance_sheet.loc["Total Current Liabilities", year] + self.balance_sheet.loc["Total Debt", year]
            )

        # Assets (will be updated with cash from CF)
        for year in self.years:
            self.balance_sheet.loc["Total Assets", year] = (
                self.balance_sheet.loc["Total Current Assets", year]
                + self.balance_sheet.loc["PP&E, Net", year]
                + self.balance_sheet.loc["Goodwill", year]
                + self.balance_sheet.loc["Intangible Assets (Financing Fees)", year]
            )

    def _calculate_operating_activities(self, year: int) -> float:
        """Calculate operating activities cash flow."""
        net_income = self.income_statement.loc["Net Income", year]
        self.cash_flow.loc["Net Income", year] = net_income

        depreciation = self.income_statement.loc["Depreciation", year]
        amortization = self.income_statement.loc["Amortization", year]
        da = depreciation + amortization
        self.cash_flow.loc["Depreciation & Amortization", year] = self._round_value(da)

        net_wc_change = self._calculate_working_capital_changes(year)
        cfo = self._round_value(net_income + da + net_wc_change)

        return cfo

    def _calculate_working_capital_changes(self, year: int) -> float:
        """Calculate working capital changes."""
        if year == 1:
            prev_ar = (
                self.balance_sheet.loc["Accounts Receivable", 1]
                if pd.notna(self.balance_sheet.loc["Accounts Receivable", 1])
                else 0
            )
            prev_inv = self.balance_sheet.loc["Inventory", 1] if pd.notna(self.balance_sheet.loc["Inventory", 1]) else 0
            prev_ap = (
                self.balance_sheet.loc["Accounts Payable", 1]
                if pd.notna(self.balance_sheet.loc["Accounts Payable", 1])
                else 0
            )
        else:
            prev_ar = (
                self.balance_sheet.loc["Accounts Receivable", year - 1]
                if pd.notna(self.balance_sheet.loc["Accounts Receivable", year - 1])
                else 0
            )
            prev_inv = (
                self.balance_sheet.loc["Inventory", year - 1]
                if pd.notna(self.balance_sheet.loc["Inventory", year - 1])
                else 0
            )
            prev_ap = (
                self.balance_sheet.loc["Accounts Payable", year - 1]
                if pd.notna(self.balance_sheet.loc["Accounts Payable", year - 1])
                else 0
            )

        curr_ar = (
            self.balance_sheet.loc["Accounts Receivable", year]
            if pd.notna(self.balance_sheet.loc["Accounts Receivable", year])
            else 0
        )
        curr_inv = (
            self.balance_sheet.loc["Inventory", year] if pd.notna(self.balance_sheet.loc["Inventory", year]) else 0
        )
        curr_ap = (
            self.balance_sheet.loc["Accounts Payable", year]
            if pd.notna(self.balance_sheet.loc["Accounts Payable", year])
            else 0
        )

        self.cash_flow.loc["Change in Accounts Receivable", year] = self._round_value(-(curr_ar - prev_ar))
        self.cash_flow.loc["Change in Inventory", year] = self._round_value(-(curr_inv - prev_inv))
        self.cash_flow.loc["Change in Accounts Payable", year] = self._round_value(curr_ap - prev_ap)

        net_wc_change = self._round_value(
            self.cash_flow.loc["Change in Accounts Receivable", year]
            + self.cash_flow.loc["Change in Inventory", year]
            + self.cash_flow.loc["Change in Accounts Payable", year]
        )
        self.cash_flow.loc["Net Change in Working Capital", year] = net_wc_change

        return net_wc_change

    def _calculate_investing_activities(self, year: int) -> float:
        """Calculate investing activities cash flow."""
        revenue = self.income_statement.loc["Revenue", year]
        capex = self._round_value(-revenue * self.assumptions.capex_pct_of_revenue)
        self.cash_flow.loc["Capital Expenditures", year] = capex
        self.cash_flow.loc["Cash Flow from Investing", year] = capex
        return capex

    def _adjust_cfo_for_fcf_target(self, year: int, cfo: float, capex: float) -> float:
        """Adjust CFO to match target FCF conversion rate if specified."""
        if self.assumptions.fcf_conversion_rate > 0.01:
            ebitda = self.income_statement.loc["EBITDA", year]
            target_fcf = ebitda * self.assumptions.fcf_conversion_rate
            cfo = self._round_value(target_fcf - capex)
        return cfo

    def _calculate_financing_activities_year1(self, total_debt_repayment: float) -> float:
        """Calculate financing activities for Year 1 (transaction year)."""
        debt_issuance = self._round_value(sum(d.amount for d in self.assumptions.debt_instruments))
        self.cash_flow.loc["Debt Issuance", 1] = debt_issuance

        equity_contribution = self._round_value(self.assumptions.equity_amount)
        self.cash_flow.loc["Equity Contribution", 1] = equity_contribution

        purchase_price = self._round_value(-self.equity_value)
        self.cash_flow.loc["Purchase Price (Equity Value)", 1] = purchase_price

        existing_debt_repayment = self._round_value(-self.assumptions.existing_debt)
        self.cash_flow.loc["Existing Debt Repayment", 1] = existing_debt_repayment

        transaction_expenses_cf = self._round_value(-self.transaction_expenses)
        self.cash_flow.loc["Transaction Expenses", 1] = transaction_expenses_cf

        financing_fees_cf = self._round_value(-self.financing_fees)
        self.cash_flow.loc["Financing Fees", 1] = financing_fees_cf

        cff = self._round_value(
            debt_issuance
            + equity_contribution
            + purchase_price
            + existing_debt_repayment
            + transaction_expenses_cf
            + financing_fees_cf
            - total_debt_repayment
        )
        return cff

    def _calculate_financing_activities_future_years(self, year: int, total_debt_repayment: float) -> float:
        """Calculate financing activities for Years 2+ (no transaction items)."""
        self.cash_flow.loc["Debt Issuance", year] = 0.0
        self.cash_flow.loc["Equity Contribution", year] = 0.0
        self.cash_flow.loc["Purchase Price (Equity Value)", year] = 0.0
        self.cash_flow.loc["Existing Debt Repayment", year] = 0.0
        self.cash_flow.loc["Transaction Expenses", year] = 0.0
        self.cash_flow.loc["Financing Fees", year] = 0.0

        return self._round_value(-total_debt_repayment)

    def _reconcile_cash_balance(self, year: int, beginning_cash: float, cfo: float, capex: float, cff: float) -> float:
        """Reconcile cash balance for a year."""
        net_change = self._round_value(cfo + capex + cff)
        self.cash_flow.loc["Net Change in Cash", year] = net_change
        self.cash_flow.loc["Beginning Cash Balance", year] = self._round_value(beginning_cash)

        ending_cash = self._round_value(beginning_cash + net_change)
        if self.assumptions.min_cash_balance > 0:
            ending_cash = self._round_value(max(ending_cash, self.assumptions.min_cash_balance))

        self.cash_flow.loc["Ending Cash Balance", year] = ending_cash
        self.balance_sheet.loc["Cash", year] = ending_cash

        return ending_cash

    def _build_cash_flow_statement(self) -> None:
        """Build cash flow statement."""
        beginning_cash = self.assumptions.existing_cash

        for year in self.years:
            cfo = self._calculate_operating_activities(year)
            capex = self._calculate_investing_activities(year)
            cfo = self._adjust_cfo_for_fcf_target(year, cfo, capex)
            self.cash_flow.loc["Cash Flow from Operations", year] = cfo

            year_idx = year - 1
            total_debt_repayment = sum(
                self.debt_schedule[debt.name]["principal_paid"][year_idx] for debt in self.assumptions.debt_instruments
            )
            self.cash_flow.loc["Debt Repayment", year] = self._round_value(-total_debt_repayment)

            if year == 1:
                cff = self._calculate_financing_activities_year1(total_debt_repayment)
            else:
                cff = self._calculate_financing_activities_future_years(year, total_debt_repayment)

            self.cash_flow.loc["Cash Flow from Financing", year] = cff
            beginning_cash = self._reconcile_cash_balance(year, beginning_cash, cfo, capex, cff)

        for year in self.years:
            self._update_balance_sheet_totals(year)

    def _apply_cash_flow_sweep(self) -> None:
        """Apply cash flow sweep: Use excess cash to pay down debt (industry standard).

        Industry Standard Cash Flow Sweep:
        - Debt can only be paid with cash that the firm generates (CFO - CapEx - required payments)
        - Excess cash above minimum balance can be used for voluntary debt paydown
        - Sweep applies in priority order (senior debt first)
        - Debt paydown carries over between years (excess cash from Year N can pay Year N+1 debt)

        Implementation uses iterative approach to handle cascading effects:
        1. Calculate available cash for sweep (CFO - CapEx - required debt service - min cash)
        2. Apply sweep to use excess cash for debt paydown
        3. Recalculate interest expense for future years
        4. Recalculate affected income statement and cash flow items
        5. Repeat until no more excess cash can be swept

        Key refinement: The sweep only uses cash generated by operations, not just ending cash balance.
        This follows industry standards where debt can only be paid with cash the firm generates.
        """
        # Determine minimum cash balance
        if self.assumptions.min_cash_balance > 0:
            min_cash = self.assumptions.min_cash_balance
        else:
            # Default: 1% of starting revenue or $10M, whichever is higher
            min_cash = max(self.assumptions.starting_revenue * 0.01, 10000)

        # Store original debt schedule BEFORE any modifications
        # This is the debt schedule after scheduled payments but BEFORE sweep
        original_debt_schedule = {}
        for debt in self.assumptions.debt_instruments:
            original_debt_schedule[debt.name] = {
                "beginning_balance": [b for b in self.debt_schedule[debt.name]["beginning_balance"]],
                "principal_paid": [p for p in self.debt_schedule[debt.name]["principal_paid"]],
                "ending_balance": [e for e in self.debt_schedule[debt.name]["ending_balance"]],
                "interest_paid": [i for i in self.debt_schedule[debt.name]["interest_paid"]],
            }

        # Iterative sweep: up to 5 passes to handle cascading effects
        max_iterations = 5
        for iteration in range(max_iterations):
            total_sweep_this_iteration = 0.0

            # Apply sweep for each year sequentially
            for year_idx, year in enumerate(self.years):
                # Calculate available cash for sweep using helper method
                total_available_for_sweep = self._calculate_available_cash_for_sweep(
                    year, year_idx, min_cash, original_debt_schedule
                )

                # Only apply sweep if we have positive available cash
                if total_available_for_sweep > 0.01:  # Small tolerance for rounding
                    remaining_sweep = total_available_for_sweep

                    # Sort debt by priority (1 = senior, 2 = subordinated, etc.)
                    sorted_debt = sorted(self.assumptions.debt_instruments, key=lambda d: d.priority)

                    # Determine if this is the final/exit year
                    is_exit_year = year_idx == len(self.years) - 1

                    for debt in sorted_debt:
                        if remaining_sweep <= 0.01:
                            break

                        # Apply sweep to this debt instrument using helper method
                        actual_sweep = self._apply_sweep_to_single_debt(
                            debt, year, year_idx, remaining_sweep, min_cash, original_debt_schedule, is_exit_year
                        )

                        remaining_sweep -= actual_sweep
                        total_sweep_this_iteration += actual_sweep

                    # Update total debt on balance sheet after sweep for this year
                    total_debt_after_sweep = sum(
                        self.debt_schedule[debt.name]["ending_balance"][year_idx]
                        for debt in self.assumptions.debt_instruments
                    )
                    self.balance_sheet.loc["Total Debt", year] = self._round_value(total_debt_after_sweep)

                    # CRITICAL: Recalculate Total Liabilities and Equity after debt changes
                    total_current_liab = self.balance_sheet.loc["Total Current Liabilities", year]
                    total_debt = self.balance_sheet.loc["Total Debt", year]
                    self.balance_sheet.loc["Total Liabilities", year] = self._round_value(
                        total_current_liab + total_debt
                    )

                    # Recalculate equity to keep balance sheet balanced
                    total_assets = self.balance_sheet.loc["Total Assets", year]
                    total_liab = self.balance_sheet.loc["Total Liabilities", year]
                    equity = self._round_value(total_assets - total_liab)
                    self.balance_sheet.loc["Shareholders Equity", year] = equity
                    self.balance_sheet.loc["Total Liabilities & Equity", year] = self._round_value(total_liab + equity)

            # If no sweep was applied in this iteration, we're done
            if total_sweep_this_iteration < 0.01:
                break

            # Recalculate debt schedule, interest expense, and affected items for all future years
            # This handles cascading effects of debt paydown
            for future_year_idx in range(len(self.years)):
                future_year = self.years[future_year_idx]
                self._recalculate_financial_statements_after_sweep(
                    future_year, future_year_idx, original_debt_schedule, iteration
                )

        # FINAL VALIDATION PASS: Ensure all debt schedules are valid after sweep
        # This corrects any remaining issues where principal might exceed beginning balance
        for debt in self.assumptions.debt_instruments:
            for year_idx, year in enumerate(self.years):
                beg_balance = self.debt_schedule[debt.name]["beginning_balance"][year_idx]
                principal = self.debt_schedule[debt.name]["principal_paid"][year_idx]

                # CRITICAL: Ensure principal never exceeds beginning balance
                if principal > beg_balance:
                    logger.warning(
                        f"Final correction: {debt.name} Year {year}: "
                        f"Principal ${principal:,.2f} exceeds beginning balance ${beg_balance:,.2f}. "
                        f"Limiting principal to beginning balance."
                    )
                    principal = beg_balance
                    self.debt_schedule[debt.name]["principal_paid"][year_idx] = self._round_value(principal)

                # Recalculate ending balance to ensure Beginning - Principal = Ending
                ending = self._round_value(beg_balance - principal)
                ending = max(0, ending)  # Ensure non-negative
                self.debt_schedule[debt.name]["ending_balance"][year_idx] = ending

                # Recalculate interest based on beginning balance
                interest = beg_balance * debt.interest_rate
                self.debt_schedule[debt.name]["interest_paid"][year_idx] = self._round_value(interest)

                # Update next year's beginning balance if not last year
                if year_idx < len(self.years) - 1:
                    self.debt_schedule[debt.name]["beginning_balance"][year_idx + 1] = ending

        logger.debug(f"Cash flow sweep completed after {iteration + 1} iteration(s)")

    def _calculate_available_cash_for_sweep(
        self, year: int, year_idx: int, min_cash: float, original_debt_schedule: Dict
    ) -> float:
        """Calculate available cash for debt sweep in a given year.

        Args:
            year: Year number
            year_idx: Year index (0-based)
            min_cash: Minimum cash balance required
            original_debt_schedule: Original debt schedule before sweep

        Returns:
            Total available cash for sweep
        """
        # Industry standard: Calculate available cash for sweep
        # Available cash = Cash Generated by Operations - Required Payments - Minimum Cash
        cfo = self.cash_flow.loc["Cash Flow from Operations", year]
        capex = abs(self.cash_flow.loc["Cash Flow from Investing", year])  # CapEx is negative, make positive

        # Required debt service = scheduled principal + interest payments
        required_principal = sum(
            original_debt_schedule[debt.name]["principal_paid"][year_idx] for debt in self.assumptions.debt_instruments
        )
        required_interest = sum(
            original_debt_schedule[debt.name]["interest_paid"][year_idx] for debt in self.assumptions.debt_instruments
        )
        required_debt_service = required_principal + required_interest

        # Available cash from operations for sweep
        available_from_operations = cfo - capex - required_debt_service - min_cash

        # Also consider excess beginning cash (carryover from previous year)
        if year_idx == 0:
            beginning_cash = self.assumptions.existing_cash
        else:
            beginning_cash = self.cash_flow.loc["Ending Cash Balance", self.years[year_idx - 1]]

        excess_beginning_cash = max(0, beginning_cash - min_cash)

        # Total available for sweep
        total_available = available_from_operations + excess_beginning_cash

        # Apply target exit debt limit if specified
        current_total_debt = sum(
            self.debt_schedule[debt.name]["ending_balance"][year_idx] for debt in self.assumptions.debt_instruments
        )

        if self.assumptions.target_exit_debt > 0.01:
            remaining_paydown_needed = current_total_debt - self.assumptions.target_exit_debt
            if remaining_paydown_needed <= 0.01:
                total_available = 0.0

        # Apply max annual paydown limit if specified
        if self.assumptions.max_debt_paydown_per_year > 0.01:
            total_available = min(total_available, self.assumptions.max_debt_paydown_per_year)

        return max(0, total_available)

    def _calculate_available_debt_to_sweep(self, debt: LBODebtStructure, year_idx: int) -> float:
        """Calculate available debt that can be swept."""
        current_beginning_balance = self.debt_schedule[debt.name]["beginning_balance"][year_idx]
        current_total_principal = self.debt_schedule[debt.name]["principal_paid"][year_idx]
        remaining_debt = current_beginning_balance - current_total_principal
        return max(0, remaining_debt)

    def _apply_target_exit_debt_limit(self, actual_sweep: float, year_idx: int) -> float:
        """Apply target exit debt limit to sweep amount."""
        if self.assumptions.target_exit_debt <= 0.01:
            return actual_sweep

        current_total_debt = sum(
            self.debt_schedule[d.name]["ending_balance"][year_idx] for d in self.assumptions.debt_instruments
        )
        current_debt_after_sweep = current_total_debt - actual_sweep

        if current_debt_after_sweep < self.assumptions.target_exit_debt:
            max_sweep = current_total_debt - self.assumptions.target_exit_debt
            actual_sweep = min(actual_sweep, max_sweep)
            actual_sweep = max(0, actual_sweep)

        return actual_sweep

    def _update_debt_schedule_after_sweep(self, debt: LBODebtStructure, year_idx: int, actual_sweep: float) -> None:
        """Update debt schedule after applying sweep."""
        current_beginning_balance = self.debt_schedule[debt.name]["beginning_balance"][year_idx]
        current_total_principal = self.debt_schedule[debt.name]["principal_paid"][year_idx]

        new_total_principal = self._round_value(current_total_principal + actual_sweep)
        new_total_principal = min(new_total_principal, current_beginning_balance)
        self.debt_schedule[debt.name]["principal_paid"][year_idx] = new_total_principal

        new_ending = self._round_value(current_beginning_balance - new_total_principal)
        new_ending = max(0, new_ending)
        self.debt_schedule[debt.name]["ending_balance"][year_idx] = new_ending

        for future_idx in range(year_idx + 1, len(self.years)):
            if future_idx == year_idx + 1:
                self.debt_schedule[debt.name]["beginning_balance"][future_idx] = new_ending
            else:
                prev_ending = self.debt_schedule[debt.name]["ending_balance"][future_idx - 1]
                self.debt_schedule[debt.name]["beginning_balance"][future_idx] = prev_ending

        recalc_interest = current_beginning_balance * debt.interest_rate
        self.debt_schedule[debt.name]["interest_paid"][year_idx] = self._round_value(recalc_interest)

    def _update_cash_flow_after_sweep(
        self, debt: LBODebtStructure, year: int, year_idx: int, actual_sweep: float, min_cash: float
    ) -> float:
        """Update cash flow after sweep and return adjusted sweep amount."""
        current_repayment = self.cash_flow.loc["Debt Repayment", year]
        self.cash_flow.loc["Debt Repayment", year] = self._round_value(current_repayment - actual_sweep)

        current_cff = self.cash_flow.loc["Cash Flow from Financing", year]
        self.cash_flow.loc["Cash Flow from Financing", year] = self._round_value(current_cff - actual_sweep)

        current_net_change = self.cash_flow.loc["Net Change in Cash", year]
        self.cash_flow.loc["Net Change in Cash", year] = self._round_value(current_net_change - actual_sweep)

        current_ending_cash = self.cash_flow.loc["Ending Cash Balance", year]
        new_ending_cash = self._round_value(current_ending_cash - actual_sweep)

        if new_ending_cash < min_cash:
            excess_sweep = min_cash - new_ending_cash
            actual_sweep = self._round_value(actual_sweep - excess_sweep)
            new_ending_cash = min_cash

        self.cash_flow.loc["Ending Cash Balance", year] = new_ending_cash
        self.balance_sheet.loc["Cash", year] = new_ending_cash

        if year_idx < len(self.years) - 1:
            next_year = year + 1
            self.cash_flow.loc["Beginning Cash Balance", next_year] = new_ending_cash

        return actual_sweep

    def _apply_sweep_to_single_debt(
        self,
        debt: LBODebtStructure,
        year: int,
        year_idx: int,
        sweep_amount: float,
        min_cash: float,
        original_debt_schedule: Dict,
        is_exit_year: bool,
    ) -> float:
        """Apply sweep to a single debt instrument.

        Args:
            debt: Debt instrument to sweep
            year: Year number
            year_idx: Year index
            sweep_amount: Amount available to sweep
            min_cash: Minimum cash balance
            original_debt_schedule: Original debt schedule
            is_exit_year: Whether this is the exit year

        Returns:
            Actual sweep amount applied
        """
        if debt.amortization_schedule == "bullet" and not is_exit_year:
            return 0.0

        available_debt_to_sweep = self._calculate_available_debt_to_sweep(debt, year_idx)
        if available_debt_to_sweep <= 0.01 or sweep_amount <= 0.01:
            return 0.0

        actual_sweep = min(sweep_amount, available_debt_to_sweep)
        actual_sweep = self._apply_target_exit_debt_limit(actual_sweep, year_idx)
        actual_sweep = self._round_value(actual_sweep)

        current_total_principal = self.debt_schedule[debt.name]["principal_paid"][year_idx]
        current_beginning_balance = self.debt_schedule[debt.name]["beginning_balance"][year_idx]
        proposed_total_principal = current_total_principal + actual_sweep

        if proposed_total_principal > current_beginning_balance:
            actual_sweep = max(0, current_beginning_balance - current_total_principal)
            actual_sweep = self._round_value(actual_sweep)

        if actual_sweep <= 0.01:
            return 0.0

        self._update_debt_schedule_after_sweep(debt, year_idx, actual_sweep)
        actual_sweep = self._update_cash_flow_after_sweep(debt, year, year_idx, actual_sweep, min_cash)
        self._update_balance_sheet_totals(year)

        return actual_sweep

    def _recalculate_financial_statements_after_sweep(
        self, future_year: int, future_year_idx: int, original_debt_schedule: Dict, iteration: int
    ) -> None:
        """Recalculate financial statements after debt sweep.

        Args:
            future_year: Year to recalculate
            future_year_idx: Year index
            original_debt_schedule: Original debt schedule
            iteration: Current iteration number
        """
        # Recalculate debt schedule for this year
        for debt in self.assumptions.debt_instruments:
            beg_balance = self.debt_schedule[debt.name]["beginning_balance"][future_year_idx]

            if beg_balance <= 0.01:
                self.debt_schedule[debt.name]["principal_paid"][future_year_idx] = 0.0
                self.debt_schedule[debt.name]["ending_balance"][future_year_idx] = 0.0
                self.debt_schedule[debt.name]["interest_paid"][future_year_idx] = 0.0
                if future_year_idx < len(self.years) - 1:
                    self.debt_schedule[debt.name]["beginning_balance"][future_year_idx + 1] = 0.0
            else:
                interest = beg_balance * debt.interest_rate
                self.debt_schedule[debt.name]["interest_paid"][future_year_idx] = self._round_value(interest)

                original_scheduled = original_debt_schedule[debt.name]["principal_paid"][future_year_idx]
                current_total = self.debt_schedule[debt.name]["principal_paid"][future_year_idx]
                already_swept = current_total - original_scheduled

                if already_swept <= 0.01:
                    scheduled_principal = 0.0
                    if debt.amortization_schedule == "amortizing":
                        years_amortized = sum(
                            1
                            for i in range(future_year_idx)
                            if self.debt_schedule[debt.name]["principal_paid"][i] > 0.01
                        )
                        if years_amortized < debt.amortization_periods:
                            original_scheduled = self._round_value(debt.amount / debt.amortization_periods)
                            scheduled_principal = min(original_scheduled, beg_balance)
                            scheduled_principal = self._round_value(scheduled_principal)
                    elif debt.amortization_schedule == "bullet":
                        if future_year == self.num_years:
                            scheduled_principal = self._round_value(beg_balance)

                    scheduled_principal = min(scheduled_principal, beg_balance)
                    self.debt_schedule[debt.name]["principal_paid"][future_year_idx] = self._round_value(
                        scheduled_principal
                    )

                current_principal = self.debt_schedule[debt.name]["principal_paid"][future_year_idx]
                if current_principal > beg_balance:
                    logger.warning(
                        f"Correcting principal paid for {debt.name} Year {future_year}: "
                        f"${current_principal:,.2f} exceeds beginning balance ${beg_balance:,.2f}. "
                        f"Limiting to beginning balance."
                    )
                    current_principal = beg_balance
                current_principal = self._round_value(current_principal)
                self.debt_schedule[debt.name]["principal_paid"][future_year_idx] = current_principal

                new_ending = self._round_value(beg_balance - current_principal)
                new_ending = max(0, new_ending)
                self.debt_schedule[debt.name]["ending_balance"][future_year_idx] = new_ending

                if future_year_idx < len(self.years) - 1:
                    self.debt_schedule[debt.name]["beginning_balance"][future_year_idx + 1] = new_ending

        # Update total debt on balance sheet
        total_debt = self._calculate_total_debt(future_year)
        self.balance_sheet.loc["Total Debt", future_year] = self._round_value(total_debt)

        # Recalculate income statement
        self._update_income_statement_from_ebit(future_year)

        # Update cash flow
        net_income = self.income_statement.loc["Net Income", future_year]
        self.cash_flow.loc["Net Income", future_year] = net_income

        da = self.cash_flow.loc["Depreciation & Amortization", future_year]
        net_wc_change = self.cash_flow.loc["Net Change in Working Capital", future_year]
        cfo = self._round_value(net_income + da + net_wc_change)

        if self.assumptions.fcf_conversion_rate > 0.01:
            ebitda = self.income_statement.loc["EBITDA", future_year]
            capex = self.cash_flow.loc["Cash Flow from Investing", future_year]
            target_fcf = ebitda * self.assumptions.fcf_conversion_rate
            cfo = self._round_value(target_fcf - capex)

        self.cash_flow.loc["Cash Flow from Operations", future_year] = cfo

        capex = self.cash_flow.loc["Cash Flow from Investing", future_year]
        cff = self.cash_flow.loc["Cash Flow from Financing", future_year]
        net_change = self._round_value(cfo + capex + cff)
        self.cash_flow.loc["Net Change in Cash", future_year] = net_change

        if future_year_idx == 0:
            beginning_cash = self.assumptions.existing_cash
        else:
            beginning_cash = self.cash_flow.loc["Ending Cash Balance", self.years[future_year_idx - 1]]

        ending_cash = self._round_value(beginning_cash + net_change)
        if self.assumptions.min_cash_balance > 0 and iteration == 0:
            ending_cash = self._round_value(max(ending_cash, self.assumptions.min_cash_balance))

        self.cash_flow.loc["Beginning Cash Balance", future_year] = self._round_value(beginning_cash)
        self.cash_flow.loc["Ending Cash Balance", future_year] = ending_cash
        self.balance_sheet.loc["Cash", future_year] = ending_cash

        # Update balance sheet
        self._update_balance_sheet_totals(future_year)

    def _validate_debt_basic_checks(
        self,
        debt_name: str,
        year: int,
        year_idx: int,
        beg_bal: float,
        principal: float,
        interest: float,
        end_bal: float,
        debt: LBODebtStructure,
        tolerance: float,
    ) -> List[str]:
        """Validate basic debt schedule checks for a year."""
        errors = []

        # Check principal doesn't exceed beginning balance
        if principal > beg_bal + tolerance:
            errors.append(
                f"{debt_name} Year {year}: Principal paid (${principal:,.2f}) "
                f"exceeds beginning balance (${beg_bal:,.2f})"
            )

        # Check ending balance is non-negative
        if end_bal < -tolerance:
            errors.append(f"{debt_name} Year {year}: Ending balance (${end_bal:,.2f}) is negative")

        # Check interest calculation
        calc_interest = beg_bal * debt.interest_rate
        if abs(calc_interest - interest) > tolerance:
            errors.append(
                f"{debt_name} Year {year}: Interest calculation error. "
                f"Expected (${calc_interest:,.2f}) ≠ Actual (${interest:,.2f})"
            )

        return errors

    def _validate_amortizing_schedule(
        self,
        debt_name: str,
        year: int,
        year_idx: int,
        principal: float,
        end_bal: float,
        debt: LBODebtStructure,
        tolerance: float,
        scenarios: Dict,
    ) -> List[str]:
        """Validate amortizing debt schedule."""
        warnings = []
        expected_principal = debt.amount / debt.amortization_periods

        if year_idx < debt.amortization_periods:
            if abs(principal - expected_principal) > tolerance * 10:
                if principal < expected_principal:
                    warnings.append(
                        f"{debt_name} Year {year}: Principal (${principal:,.2f}) "
                        f"less than expected amortization (${expected_principal:,.2f})"
                    )
                else:
                    scenarios["cash_flow_sweep"].append(
                        f"{debt_name} Year {year}: Principal includes sweep "
                        f"(${principal:,.2f} vs expected ${expected_principal:,.2f})"
                    )
        else:
            if end_bal > tolerance:
                warnings.append(
                    f"{debt_name} Year {year}: Should be fully paid after "
                    f"{debt.amortization_periods} periods, but ending balance is ${end_bal:,.2f}"
                )

        return warnings

    def _validate_bullet_schedule(
        self,
        debt_name: str,
        year: int,
        year_idx: int,
        principal: float,
        beg_bal: float,
        has_sweep: bool,
        tolerance: float,
        scenarios: Dict,
    ) -> List[str]:
        """Validate bullet debt schedule."""
        errors = []

        if year == self.num_years:
            if abs(principal - beg_bal) > tolerance:
                errors.append(
                    f"{debt_name} Year {year}: Bullet debt should pay full balance "
                    f"(${beg_bal:,.2f}) but paid ${principal:,.2f}"
                )
            scenarios["bullet"].append(f"{debt_name} Year {year}: Bullet payment of ${principal:,.2f}")
        else:
            if principal > tolerance:
                if has_sweep:
                    scenarios["cash_flow_sweep"].append(
                        f"{debt_name} Year {year}: Bullet debt paid early via sweep " f"(${principal:,.2f})"
                    )
                else:
                    errors.append(
                        f"{debt_name} Year {year}: Bullet debt should have no payment " f"but paid ${principal:,.2f}"
                    )

        return errors

    def _validate_final_year_debt(self, debt_name: str, final_end_bal: float, tolerance: float) -> List[str]:
        """Validate final year debt balance."""
        warnings = []
        if final_end_bal > tolerance * 100:
            warnings.append(
                f"{debt_name}: Final year ending balance (${final_end_bal:,.2f}) is not zero. "
                f"Debt may not be fully repaid."
            )
        return warnings

    def _track_debt_scenarios(
        self,
        debt: LBODebtStructure,
        debt_name: str,
        schedule: Dict,
        is_amortizing: bool,
        is_bullet: bool,
        has_sweep: bool,
        tolerance: float,
        scenarios: Dict,
    ) -> None:
        """Track debt payment scenarios."""
        if is_amortizing:
            scenarios["amortizing"].append(debt_name)
        elif is_bullet:
            scenarios["bullet"].append(debt_name)

        if has_sweep:
            total_sweep = sum(
                max(
                    0, schedule["principal_paid"][i] - (debt.amount / debt.amortization_periods if is_amortizing else 0)
                )
                for i in range(len(self.years))
            )
            if total_sweep > tolerance:
                scenarios["cash_flow_sweep"].append(f"{debt_name}: Total sweep payments of ${total_sweep:,.2f}")

    def _validate_total_debt_consistency(self, tolerance: float) -> List[str]:
        """Validate total debt matches sum of instruments."""
        errors = []
        for year_idx, year in enumerate(self.years):
            total_debt_calc = sum(
                self.debt_schedule[debt.name]["ending_balance"][year_idx] for debt in self.assumptions.debt_instruments
            )
            total_debt_bs = self.balance_sheet.loc["Total Debt", year]

            if abs(total_debt_calc - total_debt_bs) > tolerance:
                errors.append(
                    f"Year {year}: Total debt mismatch. "
                    f"Sum of instruments (${total_debt_calc:,.2f}) ≠ Balance Sheet (${total_debt_bs:,.2f})"
                )
        return errors

    def _identify_mixed_structure(self, scenarios: Dict) -> None:
        """Identify mixed debt structure scenario."""
        amortizing_count = len(
            [d for d in self.assumptions.debt_instruments if d.amortization_schedule == "amortizing"]
        )
        bullet_count = len([d for d in self.assumptions.debt_instruments if d.amortization_schedule == "bullet"])

        if amortizing_count > 0 and bullet_count > 0:
            scenarios["mixed_structure"] = [
                f"Model has {len(scenarios['amortizing'])} amortizing and "
                f"{len(scenarios['bullet'])} bullet instruments"
            ]

    def _validate_debt_schedule(self) -> Dict[str, List[str]]:
        """Validate debt schedule calculations and payment scenarios.

        Validates:
        1. Beginning balance continuity (Year N ending = Year N+1 beginning)
        2. Balance equation (Beginning - Principal = Ending)
        3. Principal paid doesn't exceed beginning balance
        4. Ending balance is non-negative
        5. Interest calculation (Interest = Beginning Balance × Rate)
        6. Payment schedule compliance (amortizing, bullet, sweep)
        7. Total debt matches sum of individual instruments

        Returns:
            Dictionary with validation results:
            {
                'errors': List of error messages,
                'warnings': List of warning messages,
                'scenarios': Dictionary of payment scenario validations
            }
        """
        errors = []
        warnings = []
        scenarios = {"amortizing": [], "bullet": [], "cash_flow_sweep": [], "mixed_structure": []}

        tolerance = 0.01

        for debt in self.assumptions.debt_instruments:
            debt_name = debt.name
            schedule = self.debt_schedule[debt_name]

            is_amortizing = debt.amortization_schedule == "amortizing"
            is_bullet = debt.amortization_schedule == "bullet"
            has_sweep = self.assumptions.min_cash_balance > 0

            for year_idx, year in enumerate(self.years):
                beg_bal = schedule["beginning_balance"][year_idx]
                principal = schedule["principal_paid"][year_idx]
                interest = schedule["interest_paid"][year_idx]
                end_bal = schedule["ending_balance"][year_idx]

                error = self._validate_debt_balance_equation(debt_name, year, year_idx, tolerance)
                if error:
                    errors.append(error)

                errors.extend(
                    self._validate_debt_basic_checks(
                        debt_name, year, year_idx, beg_bal, principal, interest, end_bal, debt, tolerance
                    )
                )

                error = self._validate_debt_continuity(debt_name, year, year_idx, tolerance)
                if error:
                    errors.append(error)

                if is_amortizing:
                    warnings.extend(
                        self._validate_amortizing_schedule(
                            debt_name, year, year_idx, principal, end_bal, debt, tolerance, scenarios
                        )
                    )
                elif is_bullet:
                    errors.extend(
                        self._validate_bullet_schedule(
                            debt_name, year, year_idx, principal, beg_bal, has_sweep, tolerance, scenarios
                        )
                    )

            final_year_idx = len(self.years) - 1
            final_end_bal = schedule["ending_balance"][final_year_idx]
            warnings.extend(self._validate_final_year_debt(debt_name, final_end_bal, tolerance))

            self._track_debt_scenarios(
                debt, debt_name, schedule, is_amortizing, is_bullet, has_sweep, tolerance, scenarios
            )

        errors.extend(self._validate_total_debt_consistency(tolerance))
        self._identify_mixed_structure(scenarios)

        return {"errors": errors, "warnings": warnings, "scenarios": scenarios}

    def _reconcile_balance_sheet(self, year: int) -> None:
        """Reconcile balance sheet for a specific year.

        Args:
            year: Year to reconcile
        """
        # Ensure Total Liabilities is correctly calculated
        total_current_liab = self.balance_sheet.loc["Total Current Liabilities", year]
        total_debt = self.balance_sheet.loc["Total Debt", year]
        calculated_total_liab = self._round_value(total_current_liab + total_debt)
        if abs(calculated_total_liab - self.balance_sheet.loc["Total Liabilities", year]) > 0.01:
            self.balance_sheet.loc["Total Liabilities", year] = calculated_total_liab

        # Ensure Total Assets is correctly calculated
        cash = self.balance_sheet.loc["Cash", year]
        ar = self.balance_sheet.loc["Accounts Receivable", year]
        inv = self.balance_sheet.loc["Inventory", year]
        calculated_current_assets = self._round_value(cash + ar + inv)
        if abs(calculated_current_assets - self.balance_sheet.loc["Total Current Assets", year]) > 0.01:
            self.balance_sheet.loc["Total Current Assets", year] = calculated_current_assets

        calculated_total_assets = self._round_value(
            self.balance_sheet.loc["Total Current Assets", year]
            + self.balance_sheet.loc["PP&E, Net", year]
            + self.balance_sheet.loc["Goodwill", year]
            + self.balance_sheet.loc["Intangible Assets (Financing Fees)", year]
        )
        if abs(calculated_total_assets - self.balance_sheet.loc["Total Assets", year]) > 0.01:
            self.balance_sheet.loc["Total Assets", year] = calculated_total_assets

        # Check if balance sheet balances
        assets = self.balance_sheet.loc["Total Assets", year]
        liab_eq = self.balance_sheet.loc["Total Liabilities & Equity", year]
        diff = abs(assets - liab_eq)
        if diff > LBOConstants.BALANCE_SHEET_TOLERANCE:
            # Adjust equity to balance
            total_liab = self.balance_sheet.loc["Total Liabilities", year]
            equity = self._round_value(assets - total_liab)
            self.balance_sheet.loc["Shareholders Equity", year] = equity
            self.balance_sheet.loc["Total Liabilities & Equity", year] = self._round_value(total_liab + equity)
            logger.debug(
                f"Adjusted equity in year {year} to ${equity:,.2f} to balance sheet (Assets=${assets:,.2f}, Liab=${total_liab:,.2f})"
            )

    def _reconcile_cash_flow(self, year: int, year_idx: int) -> float:
        """Reconcile cash flow for a specific year.

        Args:
            year: Year to reconcile
            year_idx: Index of year in years list

        Returns:
            Ending cash balance after reconciliation
        """
        # Check: Beginning + Net Change = Ending
        beginning = self.cash_flow.loc["Beginning Cash Balance", year]
        net_change = self.cash_flow.loc["Net Change in Cash", year]
        ending = self.cash_flow.loc["Ending Cash Balance", year]
        calc_ending = beginning + net_change

        if abs(calc_ending - ending) > LBOConstants.CASH_FLOW_TOLERANCE:
            logger.warning(
                f"Cash flow reconciliation issue in year {year}: "
                f"Beginning (${beginning:,.2f}) + Net Change (${net_change:,.2f}) = "
                f"${calc_ending:,.2f} but Ending = ${ending:,.2f}. Recalculating..."
            )
            ending_cash = beginning + net_change
            if self.assumptions.min_cash_balance > 0:
                ending_cash = max(ending_cash, self.assumptions.min_cash_balance)
            self.cash_flow.loc["Ending Cash Balance", year] = ending_cash
            self.balance_sheet.loc["Cash", year] = ending_cash
            ending = ending_cash

        # Check: Ending Cash Year N = Beginning Cash Year N+1
        if year_idx < len(self.years) - 1:
            next_year = self.years[year_idx + 1]
            next_beginning = self.cash_flow.loc["Beginning Cash Balance", next_year]
            if abs(ending - next_beginning) > LBOConstants.CASH_FLOW_TOLERANCE:
                logger.warning(
                    f"Cash flow continuity issue: Year {year} Ending (${ending:,.2f}) != "
                    f"Year {next_year} Beginning (${next_beginning:,.2f}). Fixing..."
                )
                self.cash_flow.loc["Beginning Cash Balance", next_year] = ending
                next_net_change = self.cash_flow.loc["Net Change in Cash", next_year]
                next_ending = ending + next_net_change
                if self.assumptions.min_cash_balance > 0:
                    next_ending = max(next_ending, self.assumptions.min_cash_balance)
                self.cash_flow.loc["Ending Cash Balance", next_year] = next_ending
                self.balance_sheet.loc["Cash", next_year] = next_ending
                self._update_balance_sheet_totals(next_year)

        return ending

    def _update_balance_sheet_totals(self, year: int) -> None:
        """Update balance sheet totals after cash changes.

        Args:
            year: Year to update
        """
        cash = self.balance_sheet.loc["Cash", year]
        ar = self.balance_sheet.loc["Accounts Receivable", year]
        inv = self.balance_sheet.loc["Inventory", year]
        self.balance_sheet.loc["Total Current Assets", year] = self._round_value(cash + ar + inv)
        self.balance_sheet.loc["Total Assets", year] = self._round_value(
            self.balance_sheet.loc["Total Current Assets", year]
            + self.balance_sheet.loc["PP&E, Net", year]
            + self.balance_sheet.loc["Goodwill", year]
            + self.balance_sheet.loc["Intangible Assets (Financing Fees)", year]
        )
        total_assets = self.balance_sheet.loc["Total Assets", year]
        total_liab = self.balance_sheet.loc["Total Liabilities", year]
        equity = self._round_value(total_assets - total_liab)
        self.balance_sheet.loc["Shareholders Equity", year] = equity
        self.balance_sheet.loc["Total Liabilities & Equity", year] = self._round_value(total_liab + equity)

    def _reconcile_cross_sheet_cash(self, year: int, ending_cash: float) -> None:
        """Reconcile cash between balance sheet and cash flow statement.

        Args:
            year: Year to reconcile
            ending_cash: Ending cash from cash flow statement
        """
        bs_cash = self.balance_sheet.loc["Cash", year]
        if abs(bs_cash - ending_cash) > LBOConstants.CASH_FLOW_TOLERANCE:
            logger.warning(
                f"Cross-sheet cash mismatch in year {year}: "
                f"Balance Sheet Cash (${bs_cash:,.2f}) != Cash Flow Ending (${ending_cash:,.2f}). "
                f"Updating balance sheet..."
            )
            self.balance_sheet.loc["Cash", year] = ending_cash
            self._update_balance_sheet_totals(year)

    def _validate_year1_sources_uses(self, year: int, ending_cash: float) -> None:
        """Validate Year 1 ending cash reconciles with Sources & Uses.

        Args:
            year: Year to validate (should be 1)
            ending_cash: Ending cash balance
        """
        if year != 1:
            return

        expected_sources = self.assumptions.equity_amount + sum(d.amount for d in self.assumptions.debt_instruments)
        expected_uses = (
            self.equity_value + self.assumptions.existing_debt + self.transaction_expenses + self.financing_fees
        )
        first_year_debt_repayment = sum(
            self.debt_schedule[debt.name]["principal_paid"][0] for debt in self.assumptions.debt_instruments
        )

        net_transaction = expected_sources - expected_uses - first_year_debt_repayment
        cfo_year1 = self.cash_flow.loc["Cash Flow from Operations", year]
        capex_year1 = self.cash_flow.loc["Cash Flow from Investing", year]

        expected_ending_cash = self.assumptions.existing_cash + net_transaction + cfo_year1 + capex_year1

        if abs(expected_ending_cash - ending_cash) > LBOConstants.CASH_FLOW_TOLERANCE:
            logger.warning(
                f"Year 1 ending cash doesn't reconcile with Sources & Uses: "
                f"Expected (${expected_ending_cash:,.2f}) != Actual (${ending_cash:,.2f}). "
                f"Sources=${expected_sources:,.2f}, Uses=${expected_uses:,.2f}, "
                f"Debt Repayment=${first_year_debt_repayment:,.2f}, CFO=${cfo_year1:,.2f}, CapEx=${capex_year1:,.2f}"
            )

    def _validate_income_statement_assumptions(self, year: int) -> None:
        """Validate Income Statement aligns with assumptions.

        Args:
            year: Year to validate
        """
        revenue = self.income_statement.loc["Revenue", year]

        if revenue <= 0:
            return

        # Check COGS aligns with assumption
        cogs = self.income_statement.loc["Cost of Goods Sold (Net of D&A)", year]
        calc_cogs_pct = cogs / revenue
        if abs(calc_cogs_pct - self.assumptions.cogs_pct_of_revenue) > 0.01:
            logger.warning(
                f"Year {year}: COGS % ({calc_cogs_pct:.1%}) doesn't match assumption "
                f"({self.assumptions.cogs_pct_of_revenue:.1%})"
            )

        # Check SG&A aligns with assumption
        sganda = self.income_statement.loc["SG&A (Net of D&A)", year]
        calc_sganda_pct = sganda / revenue
        if abs(calc_sganda_pct - self.assumptions.sganda_pct_of_revenue) > 0.01:
            logger.warning(
                f"Year {year}: SG&A % ({calc_sganda_pct:.1%}) doesn't match assumption "
                f"({self.assumptions.sganda_pct_of_revenue:.1%})"
            )

        # Check tax rate aligns with assumption (for years with positive pretax income)
        pretax = self.income_statement.loc["Pretax Income", year]
        if pretax > 0:
            tax = self.income_statement.loc["Income Tax Expense", year]
            calc_tax_rate = tax / pretax if pretax > 0 else 0
            if abs(calc_tax_rate - self.assumptions.tax_rate) > 0.01:
                logger.warning(
                    f"Year {year}: Tax Rate ({calc_tax_rate:.1%}) doesn't match assumption "
                    f"({self.assumptions.tax_rate:.1%})"
                )

    def _validate_returns_analysis(self) -> None:
        """Validate Returns Analysis aligns with assumptions."""
        returns = self.calculate_returns()
        exit_year = returns["exit_year"]

        # Check exit year aligns with assumption
        if exit_year != min(self.assumptions.exit_year, self.num_years):
            logger.warning(
                f"Returns exit year ({exit_year}) doesn't match assumption "
                f"({self.assumptions.exit_year}) or model years ({self.num_years})"
            )

        # Check exit EBITDA growth aligns with revenue growth assumptions
        entry_ebitda = self.assumptions.entry_ebitda
        exit_ebitda = returns["exit_ebitda"]
        ebitda_growth = (exit_ebitda / entry_ebitda) ** (1.0 / (exit_year - 1)) - 1 if exit_year > 1 else 0

        # Estimate expected EBITDA growth from revenue growth (simplified check)
        if exit_year > 1:
            avg_revenue_growth = sum(self.assumptions.revenue_growth_rate[: exit_year - 1]) / (exit_year - 1)
            # EBITDA should grow roughly in line with revenue (simplified check)
            if abs(ebitda_growth - avg_revenue_growth) > 0.10:  # Allow 10% difference for margin changes
                logger.debug(
                    f"Exit EBITDA growth ({ebitda_growth:.1%}) differs from average revenue growth "
                    f"({avg_revenue_growth:.1%}). This may be due to margin expansion/compression."
                )

    def _reconcile_model(self) -> None:
        """Reconcile and balance the model following industry standards.

        Industry Standard Reconciliation Checks:
        1. Balance Sheet: Assets = Liabilities + Equity
        2. Cash Flow: Beginning + Net Change = Ending
        3. Cash Flow: Ending Cash Year N = Beginning Cash Year N+1
        4. Cross-sheet: Cash from Balance Sheet = Ending Cash from Cash Flow
        5. Debt Schedule: All debt schedule validations
        6. Income Statement: Values align with assumptions (COGS%, SG&A%, Tax Rate)
        7. Returns Analysis: Exit values align with assumptions and model outputs
        """
        # Validate debt schedule first
        debt_validation = self._validate_debt_schedule()
        if debt_validation["errors"]:
            for error in debt_validation["errors"]:
                logger.error(f"Debt schedule validation error: {error}")
        if debt_validation["warnings"]:
            for warning in debt_validation["warnings"]:
                logger.warning(f"Debt schedule validation warning: {warning}")

        # Log payment scenarios
        for scenario_type, details in debt_validation["scenarios"].items():
            if details:
                logger.debug(f"Debt payment scenario '{scenario_type}': {details}")

        # Reconcile balance sheets and cash flows for each year
        for i, year in enumerate(self.years):
            # Reconcile balance sheet
            self._reconcile_balance_sheet(year)

            # Reconcile cash flow
            ending_cash = self._reconcile_cash_flow(year, i)

            # Reconcile cross-sheet cash
            self._reconcile_cross_sheet_cash(year, ending_cash)

            # Validate Year 1 Sources & Uses
            self._validate_year1_sources_uses(year, ending_cash)

            # Validate Income Statement assumptions
            self._validate_income_statement_assumptions(year)

        # Validate Returns Analysis
        self._validate_returns_analysis()

    def _calculate_irr(self, cash_flows: List[float], guess: float = 0.1, max_iter: int = None) -> float:
        """Calculate IRR using Newton-Raphson method."""
        if len(cash_flows) < 2:
            return 0.0

        if max_iter is None:
            max_iter = LBOConstants.MAX_IRR_ITERATIONS

        def npv(rate: float) -> float:
            return sum(cf / ((1 + rate) ** i) for i, cf in enumerate(cash_flows))

        def npv_derivative(rate: float) -> float:
            return sum(-i * cf / ((1 + rate) ** (i + 1)) for i, cf in enumerate(cash_flows))

        rate = guess
        for _ in range(max_iter):
            npv_val = npv(rate)
            if abs(npv_val) < LBOConstants.IRR_EPSILON:
                return rate
            derivative = npv_derivative(rate)
            if abs(derivative) < 1e-10:
                break
            rate = rate - npv_val / derivative
            rate = max(rate, -0.99)  # Prevent negative rates below -100%
            rate = min(rate, LBOConstants.MAX_IRR_RATE)  # Prevent unrealistic high rates

        return rate

    # ==================== AI INTEGRATION METHODS ====================

    def validate_with_ai(self, industry: Optional[str] = None, api_key: Optional[str] = None) -> Dict:
        """
        Validate model assumptions using AI.

        Args:
            industry: Industry sector for context
            api_key: OpenAI API key (optional, uses env var if not provided)

        Returns:
            Dictionary with validation results
        """
        try:
            from .lbo_ai_validator import LBOModelAIValidator
        except ImportError:
            from lbo_ai_validator import LBOModelAIValidator

        try:
            validator = LBOModelAIValidator(api_key=api_key)
            assumptions_dict = self._assumptions_to_dict()
            result = validator.validate_model_quality(assumptions_dict, industry)

            logger.info(
                f"AI validation completed: Valid={result.is_valid}, "
                f"Warnings={len(result.warnings)}, Errors={len(result.errors)}"
            )

            return {
                "is_valid": result.is_valid,
                "warnings": result.warnings,
                "errors": result.errors,
                "suggestions": result.suggestions,
                "confidence_score": result.confidence_score,
                "details": result.details,
            }
        except ImportError:
            logger.warning("lbo_ai_validator not available. Install openai package and set API key.")
            return {"error": "AI validator not available"}

    def review_generated_model_ai(self, excel_file_path: str, api_key: Optional[str] = None) -> Dict:
        """
        Review generated Excel model using AI.

        Args:
            excel_file_path: Path to Excel file
            api_key: OpenAI API key (optional)

        Returns:
            Dictionary with review results
        """
        try:
            from .lbo_ai_validator import LBOModelAIValidator
        except ImportError:
            from lbo_ai_validator import LBOModelAIValidator

            from lbo_ai_validator import LBOModelAIValidator
        try:
            validator = LBOModelAIValidator(api_key=api_key)
            returns = self.calculate_returns()
            model_summary = {
                "assumptions": self._assumptions_to_dict(),
                "returns": returns,
                "key_metrics": {
                    "irr": returns.get("irr", 0),
                    "moic": returns.get("moic", 0),
                    "exit_ebitda": returns.get("exit_ebitda", 0),
                    "exit_ev": returns.get("exit_ev", 0),
                },
            }

            result = validator.review_generated_model(excel_file_path, model_summary)

            logger.info(
                f"AI review completed: Valid={result.is_valid}, "
                f"Warnings={len(result.warnings)}, Errors={len(result.errors)}"
            )

            return {
                "warnings": result.warnings,
                "errors": result.errors,
                "suggestions": result.suggestions,
                "confidence_score": result.confidence_score,
                "details": result.details,
            }
        except ImportError:
            logger.warning("lbo_ai_validator not available.")
            return {"error": "AI validator not available"}
        except Exception as e:
            logger.error(f"AI review error: {e}")
            return {"error": str(e)}

    def generate_scenarios_with_ai(self, industry: Optional[str] = None, api_key: Optional[str] = None) -> Dict:
        """
        Generate sensitivity scenarios using AI.

        Args:
            industry: Industry sector
            api_key: OpenAI API key (optional)

        Returns:
            Dictionary with scenario analysis
        """
        try:
            from .lbo_ai_validator import LBOModelAIValidator
        except ImportError:
            from lbo_ai_validator import LBOModelAIValidator

        try:
            validator = LBOModelAIValidator(api_key=api_key)
            assumptions_dict = self._assumptions_to_dict()
            scenarios = validator.generate_sensitivity_scenarios(assumptions_dict, industry)

            logger.info("AI scenario generation completed")

            return {
                "base_case": scenarios.base_case,
                "high_case": scenarios.high_case,
                "low_case": scenarios.low_case,
                "key_assumptions": scenarios.key_assumptions,
                "sensitivity_matrix": scenarios.sensitivity_matrix,
            }
        except ImportError:
            logger.warning("lbo_ai_validator not available.")
            return {"error": "AI validator not available"}

    def answer_question_ai(self, question: str, api_key: Optional[str] = None) -> str:
        """
        Answer questions about the model using AI.

        Args:
            question: Natural language question
            api_key: OpenAI API key (optional)

        Returns:
            Answer text
        """
        try:
            from .lbo_ai_validator import LBOModelAIValidator
        except ImportError:
            from lbo_ai_validator import LBOModelAIValidator

        try:
            validator = LBOModelAIValidator(api_key=api_key)
            returns = self.calculate_returns()
            model_data = {
                "assumptions": self._assumptions_to_dict(),
                "returns": returns,
                "key_metrics": {
                    "irr": returns.get("irr", 0),
                    "moic": returns.get("moic", 0),
                    "exit_ebitda": returns.get("exit_ebitda", 0),
                    "exit_ev": returns.get("exit_ev", 0),
                    "equity_invested": returns.get("equity_invested", 0),
                    "exit_equity_value": returns.get("exit_equity_value", 0),
                },
            }

            answer = validator.query_model(question, model_data, self._assumptions_to_dict())
            logger.info(f"AI query answered: {question[:50]}...")

            return answer
        except ImportError:
            return "AI query feature not available. Install openai package and set API key."
            return "AI query feature not available. Install openai package and set API key."
        """
        Benchmark model against market using AI.
        
        Args:
            industry: Industry sector
            api_key: OpenAI API key (optional)
            
        Returns:
            Dictionary with benchmarking results
        """
        try:
            from .lbo_ai_validator import LBOModelAIValidator
        except ImportError:
            from lbo_ai_validator import LBOModelAIValidator

        try:
            validator = LBOModelAIValidator(api_key=api_key)
            assumptions_dict = self._assumptions_to_dict()
            industry = self.assumptions.industry if hasattr(self.assumptions, 'industry') else None
            benchmark = validator.benchmark_against_market(assumptions_dict, industry)

            logger.info("AI benchmarking completed")

            return {
                "industry_averages": benchmark.industry_averages,
                "quartiles": benchmark.quartiles,
                "deviations": benchmark.deviations,
                "recommendations": benchmark.recommendations,
            }
        except ImportError:
            logger.warning("lbo_ai_validator not available.")
            return {"error": "AI validator not available"}

    def generate_documentation_ai(self, excel_file_path: str, api_key: Optional[str] = None) -> str:
        """
        Generate model documentation using AI.

        Args:
            excel_file_path: Path to Excel file
            api_key: OpenAI API key (optional)

        Returns:
            Documentation string (markdown)
        """
        try:
            from .lbo_ai_validator import LBOModelAIValidator
        except ImportError:
            from lbo_ai_validator import LBOModelAIValidator

        try:
            validator = LBOModelAIValidator(api_key=api_key)
            assumptions_dict = self._assumptions_to_dict()
            documentation = validator.generate_model_documentation(excel_file_path, assumptions_dict)

            logger.info("AI documentation generation completed")

            return documentation
        except ImportError:
            return "# Model Documentation\n\nAI documentation feature not available. Install openai package and set API key."
        except Exception as e:
            logger.error(f"AI documentation error: {e}")
            return f"# Model Documentation\n\nError: {str(e)}"

    def diagnose_error_ai(
        self, error_message: str, stack_trace: Optional[str] = None, api_key: Optional[str] = None
    ) -> Dict:
        """
        Diagnose errors using AI.

        Args:
            error_message: Error message
            stack_trace: Optional stack trace
            api_key: OpenAI API key (optional)

        Returns:
            Dictionary with diagnosis and fixes
        """
        try:
            from .lbo_ai_validator import LBOModelAIValidator
        except ImportError:
            from lbo_ai_validator import LBOModelAIValidator

        try:
            validator = LBOModelAIValidator(api_key=api_key)
            assumptions_dict = self._assumptions_to_dict()
            diagnosis = validator.diagnose_model_errors(error_message, assumptions_dict, stack_trace)

            logger.info("AI error diagnosis completed")

            return diagnosis
        except ImportError:
            return {"error": "AI validator not available"}

    def _prepare_model_data_for_ai(self) -> Dict:
        """Prepare model data dictionary for AI processing."""
        returns = self.calculate_returns()
        return {
            "irr": returns.get("irr", 0),
            "moic": returns.get("moic", 0),
            "exit_ebitda": returns.get("exit_ebitda", 0),
            "exit_ev": returns.get("exit_ev", 0),
            "equity_invested": returns.get("equity_invested", 0),
        }

    def optimize_debt_structure_with_ai(self, api_key: Optional[str] = None) -> Dict:
        """
        Get optimization suggestions using AI.

        Args:
            api_key: OpenAI API key (optional)

        Returns:
            Dictionary with optimization recommendations
        """
        try:
            from .lbo_ai_validator import LBOModelAIValidator
        except ImportError:
            from lbo_ai_validator import LBOModelAIValidator

        try:
            validator = LBOModelAIValidator(api_key=api_key)
            model_data = self._prepare_model_data_for_ai()
            assumptions_dict = self._assumptions_to_dict()
            optimization = validator.optimize_debt_structure(model_data, assumptions_dict)

            logger.info("AI optimization analysis completed")
            return optimization
        except ImportError:
            logger.warning("lbo_ai_validator not available.")
            return {"error": "AI validator not available"}

    def _assumptions_to_dict(self) -> Dict:
        """Convert LBOAssumptions to dictionary for AI processing."""
        assumptions_dict = {
            "entry_ebitda": self.assumptions.entry_ebitda,
            "entry_multiple": self.assumptions.entry_multiple,
            "existing_debt": self.assumptions.existing_debt,
            "existing_cash": self.assumptions.existing_cash,
            "transaction_expenses_pct": self.assumptions.transaction_expenses_pct,
            "financing_fees_pct": self.assumptions.financing_fees_pct,
            "revenue_growth_rate": self.assumptions.revenue_growth_rate,
            "cogs_pct_of_revenue": self.assumptions.cogs_pct_of_revenue,
            "sganda_pct_of_revenue": self.assumptions.sganda_pct_of_revenue,
            "depreciation_pct_of_ppe": self.assumptions.depreciation_pct_of_ppe,
            "capex_pct_of_revenue": self.assumptions.capex_pct_of_revenue,
            "tax_rate": self.assumptions.tax_rate,
            "days_sales_outstanding": self.assumptions.days_sales_outstanding,
            "days_inventory_outstanding": self.assumptions.days_inventory_outstanding,
            "days_payable_outstanding": self.assumptions.days_payable_outstanding,
            "exit_year": self.assumptions.exit_year,
            "exit_multiple": self.assumptions.exit_multiple,
            "starting_revenue": self.assumptions.starting_revenue,
            "debt_instruments": [
                {
                    "name": debt.name,
                    "amount": debt.amount,
                    "interest_rate": debt.interest_rate,
                    "ebitda_multiple": debt.ebitda_multiple,
                    "amortization_schedule": debt.amortization_schedule,
                    "amortization_periods": debt.amortization_periods,
                }
                for debt in self.assumptions.debt_instruments
            ],
            "equity_amount": self.assumptions.equity_amount,
        }
        return assumptions_dict

    def calculate_returns(self) -> Dict:
        """Calculate returns analysis following industry standards.

        Industry Standard Returns Calculation:
        - Exit EV = Exit EBITDA × Exit Multiple (from assumptions)
        - Exit Equity Value = Exit EV - Exit Debt + Exit Cash
        - MOIC = Exit Equity Value / Equity Invested
        - IRR = Internal Rate of Return on equity investment

        The IRR calculation uses a simplified approach assuming:
        - Equity investment at time 0
        - No intermediate cash flows (no dividends)
        - Exit proceeds at exit year

        Note: This is a simplified IRR. For more sophisticated analysis including
        intermediate cash flows, the full equity cash flow waterfall should be used.
        """
        exit_year = min(self.assumptions.exit_year, self.num_years)

        # Validate exit year doesn't exceed model years
        if exit_year > self.num_years:
            logger.warning(
                f"Exit year ({exit_year}) exceeds model years ({self.num_years}). "
                f"Using final year ({self.num_years}) for calculations."
            )
            exit_year = self.num_years

        # Get exit EBITDA from income statement (validates alignment)
        if "EBITDA" not in self.income_statement.index:
            raise LBOCalculationError("EBITDA not found in income statement. Model may not be fully built.")
        exit_ebitda = self.income_statement.loc["EBITDA", exit_year]

        # Calculate exit EV using exit multiple from assumptions
        exit_ev = exit_ebitda * self.assumptions.exit_multiple

        # Get exit debt and cash from balance sheet (validates alignment)
        if "Total Debt" not in self.balance_sheet.index:
            raise LBOCalculationError("Total Debt not found in balance sheet. Model may not be fully built.")
        if "Cash" not in self.balance_sheet.index:
            raise LBOCalculationError("Cash not found in balance sheet. Model may not be fully built.")

        exit_debt = self.balance_sheet.loc["Total Debt", exit_year]
        exit_cash = self.balance_sheet.loc["Cash", exit_year]
        exit_equity_value = exit_ev - exit_debt + exit_cash

        # Equity invested: Use specified amount, or calculate from sources & uses
        equity_invested = (
            self.assumptions.equity_amount
            if self.assumptions.equity_amount > 0
            else (self.equity_value - sum(d.amount for d in self.assumptions.debt_instruments))
        )

        if equity_invested <= 0:
            logger.warning(
                f"Equity invested is ${equity_invested:,.0f}. "
                f"This may indicate incorrect assumptions or a calculation error."
            )

        moic = exit_equity_value / equity_invested if equity_invested > 0 else 0

        # IRR calculation: Equity investment at time 0, exit proceeds at exit year
        # Simplified approach: assumes no intermediate cash flows (no dividends)
        cash_flows = [-equity_invested] + [0] * (exit_year - 1) + [exit_equity_value]
        irr = self._calculate_irr(cash_flows) if len(cash_flows) > 1 and equity_invested > 0 else 0.0

        # Validate exit EBITDA aligns with assumptions (should match entry EBITDA growth)
        if exit_year == 1:
            entry_ebitda_check = self.income_statement.loc["EBITDA", 1]
            if abs(entry_ebitda_check - self.assumptions.entry_ebitda) > 100:
                logger.warning(
                    f"Year 1 EBITDA (${entry_ebitda_check:,.0f}) doesn't match entry EBITDA "
                    f"(${self.assumptions.entry_ebitda:,.0f}). This may indicate calculation issues."
                )

        return {
            "exit_year": exit_year,
            "exit_ebitda": exit_ebitda,
            "exit_ev": exit_ev,
            "exit_debt": exit_debt,
            "exit_cash": exit_cash,
            "exit_equity_value": exit_equity_value,
            "equity_invested": equity_invested,
            "moic": moic,
            "irr": irr,  # IRR is already a decimal (e.g., 0.3846 for 38.46%)
        }

    def get_debt_schedule_validation(self) -> Dict[str, List]:
        """Get debt schedule validation results.

        Returns:
            Dictionary with validation results including errors, warnings, and scenarios
        """
        return self._validate_debt_schedule()

    def _run_ai_validation_before_export(self, industry: Optional[str], api_key: Optional[str]) -> None:
        """Run AI validation before export if requested."""
        logger.info("Running AI validation before export...")
        validation_result = self.validate_with_ai(industry=industry, api_key=api_key)
        if not validation_result.get("is_valid", True):
            logger.warning(f"AI validation found errors: {validation_result.get('errors', [])}")
            if validation_result.get("warnings"):
                logger.warning(f"AI validation warnings: {validation_result.get('warnings', [])}")
        else:
            logger.info("AI validation passed")

    def export_to_excel(
        self,
        filename: str,
        company_name: str = "LBO Model",
        interactive: bool = True,
        validate_with_ai: bool = False,
        industry: Optional[str] = None,
        api_key: Optional[str] = None,
    ) -> None:
        """Export model to Excel with professional formatting matching industry standards.

        Args:
            filename: Output Excel filename
            company_name: Company name for title
            interactive: Create interactive Excel with formulas and multiple sheets
            validate_with_ai: Run AI validation before export
            industry: Industry sector for AI validation (if validate_with_ai=True)
            api_key: OpenAI API key for AI validation (optional, uses env var if not provided)
        """
        if validate_with_ai:
            self._run_ai_validation_before_export(industry, api_key)

        logger.info("Exporting using industry-standard format...")
        exporter = IndustryStandardExcelExporter(self)
        exporter.export(filename, company_name)
        logger.info(f"Industry-standard Excel file created: {filename}")


def create_lbo_from_inputs(input_config: Dict, validate: bool = True) -> LBOModel:
    """
    Create LBO model from input dictionary.

    Expected input_config structure:
    {
        "entry_ebitda": 10000,
        "entry_multiple": 6.5,
        "revenue_growth_rate": [0.05, 0.05, 0.05, 0.05, 0.05],
        "debt_instruments": [
            {"name": "Senior Debt", "interest_rate": 0.08, "ebitda_multiple": 1.0, "amortization_schedule": "cash_flow_sweep"},
            {"name": "Sub Debt", "interest_rate": 0.12, "ebitda_multiple": 2.0, "amortization_schedule": "bullet"}
        ],
        "exit_year": 5,
        "exit_multiple": 7.5,
        ...
    }
    """
    # Enhanced validation if requested
    if validate:
        try:
            from .lbo_validation_enhanced import EnhancedLBOValidator

            validation_result = EnhancedLBOValidator.validate_comprehensive(input_config)
            if not validation_result.is_valid:
                for error in validation_result.errors:
                    logger.warning(f"Validation error: {error}")
            if validation_result.warnings:
                for warning in validation_result.warnings:
                    logger.warning(f"Validation warning: {warning}")
        except ImportError:
            logger.debug("Enhanced validation not available, skipping")

    # Standardize configuration
    try:
        from .lbo_consistency_helpers import LBOConsistencyHelper

        input_config = LBOConsistencyHelper.standardize_config(input_config)
    except ImportError:
        logger.debug("Consistency helpers not available, skipping standardization")

    # Parse debt instruments
    debt_instruments = []
    for debt_dict in input_config.get("debt_instruments", []):
        # Ensure priority is set
        if "priority" not in debt_dict:
            debt_dict["priority"] = len(debt_instruments) + 1

        debt = LBODebtStructure(
            name=debt_dict["name"],
            amount=debt_dict.get("amount", 0),
            interest_rate=debt_dict["interest_rate"],
            ebitda_multiple=debt_dict.get("ebitda_multiple"),
            amortization_schedule=debt_dict.get("amortization_schedule", "bullet"),
            amortization_periods=debt_dict.get("amortization_periods", 5),
            priority=debt_dict.get("priority", len(debt_instruments) + 1),
        )
        debt_instruments.append(debt)

    assumptions = LBOAssumptions(
        entry_ebitda=input_config["entry_ebitda"],
        entry_multiple=input_config["entry_multiple"],
        existing_debt=input_config.get("existing_debt", 0),
        existing_cash=input_config.get("existing_cash", 0),
        transaction_expenses_pct=input_config.get("transaction_expenses_pct", 0.03),
        financing_fees_pct=input_config.get("financing_fees_pct", 0.02),
        debt_instruments=debt_instruments,
        equity_amount=input_config.get("equity_amount") or 0.0,
        revenue_growth_rate=input_config.get("revenue_growth_rate", [0.05] * 5),
        cogs_pct_of_revenue=input_config.get("cogs_pct_of_revenue", 0.70),
        sganda_pct_of_revenue=input_config.get("sganda_pct_of_revenue", 0.15),
        depreciation_pct_of_ppe=input_config.get("depreciation_pct_of_ppe", 0.10),
        capex_pct_of_revenue=input_config.get("capex_pct_of_revenue", 0.03),
        tax_rate=input_config.get("tax_rate", 0.25),
        days_sales_outstanding=input_config.get("days_sales_outstanding", 45),
        days_inventory_outstanding=input_config.get("days_inventory_outstanding", 30),
        days_payable_outstanding=input_config.get("days_payable_outstanding", 30),
        initial_ppe=input_config.get("initial_ppe", 0),
        initial_ar=input_config.get("initial_ar", 0),
        initial_inventory=input_config.get("initial_inventory", 0),
        initial_ap=input_config.get("initial_ap", 0),
        min_cash_balance=input_config.get("min_cash_balance", 0),
        exit_year=input_config.get("exit_year", 5),
        exit_multiple=input_config.get("exit_multiple", 7.5),
        target_exit_debt=input_config.get("target_exit_debt", 0.0),
        max_debt_paydown_per_year=input_config.get("max_debt_paydown_per_year", 0.0),
        fcf_conversion_rate=input_config.get("fcf_conversion_rate", 0.0),
        starting_revenue=input_config.get("starting_revenue", 0),
    )

    return LBOModel(assumptions)


if __name__ == "__main__":
    # Example usage
    example_inputs = {
        "entry_ebitda": 10000,
        "entry_multiple": 6.5,
        "revenue_growth_rate": [0.05, 0.05, 0.05, 0.05, 0.05],
        "debt_instruments": [
            {
                "name": "Senior Debt",
                "interest_rate": 0.08,
                "ebitda_multiple": 1.0,
                "amortization_schedule": "amortizing",
                "amortization_periods": 5,
            },
            {
                "name": "Subordinated Debt",
                "interest_rate": 0.12,
                "ebitda_multiple": 2.0,
                "amortization_schedule": "bullet",
            },
        ],
        "exit_year": 5,
        "exit_multiple": 7.5,
        "tax_rate": 0.25,
        "starting_revenue": 50000,
    }

    model = create_lbo_from_inputs(example_inputs)
    model.export_to_excel("lbo_model_output.xlsx")
    print("LBO model generated successfully!")
    print("\nReturns Analysis:")
    returns = model.calculate_returns()
    for key, value in returns.items():
        print(f"  {key}: {value:,.2f}")
