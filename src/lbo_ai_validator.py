"""
AI-Powered LBO Model Validator and Enhancement Suite

Comprehensive AI integration for LBO model validation, quality assurance,
scenario analysis, documentation, and user assistance.
"""

import json
import logging
import openpyxl
from typing import Dict, List, Optional, Any
from dataclasses import dataclass
import openai

# Handle both package and direct imports
try:
    from .lbo_exceptions import LBOAIServiceError, LBOConfigurationError
    from .lbo_validation import validate_api_key
except ImportError:
    from lbo_exceptions import LBOAIServiceError, LBOConfigurationError
    from lbo_validation import validate_api_key

logger = logging.getLogger(__name__)


@dataclass
class ValidationResult:
    """Result of AI validation."""

    is_valid: bool
    warnings: List[str]
    errors: List[str]
    suggestions: List[str]
    confidence_score: float
    details: Dict[str, Any]


@dataclass
class ScenarioAnalysis:
    """Scenario analysis result."""

    base_case: Dict[str, Any]
    high_case: Dict[str, Any]
    low_case: Dict[str, Any]
    key_assumptions: List[str]
    sensitivity_matrix: Dict[str, Dict[str, float]]


@dataclass
class BenchmarkResult:
    """Market benchmarking result."""

    industry_averages: Dict[str, float]
    quartiles: Dict[str, Dict[str, float]]
    deviations: Dict[str, float]
    recommendations: List[str]


class LBOModelAIValidator:
    """Comprehensive AI-powered validator and enhancer for LBO models."""

    def __init__(self, api_key: Optional[str] = None, model: str = "gpt-4o-mini"):
        """
        Initialize AI validator.

        Args:
            api_key: OpenAI API key (or set OPENAI_API_KEY env variable)
            model: Model to use (default: gpt-4o-mini for cost efficiency)
        """
        # Validate API key
        try:
            self.api_key = validate_api_key(api_key)
        except LBOConfigurationError as e:
            raise LBOConfigurationError(str(e)) from e

        self.client = openai.OpenAI(api_key=self.api_key)
        self.model = model

    # ==================== HELPER METHODS FOR AI OPERATIONS ====================

    def _call_openai_api(
        self,
        system_message: str,
        user_prompt: str,
        temperature: float = 0.3,
        response_format: Optional[Dict] = None,
    ) -> str:
        """Make OpenAI API call with error handling.

        Args:
            system_message: System role message
            user_prompt: User prompt content
            temperature: Temperature setting (default: 0.3)
            response_format: Optional response format (e.g., {"type": "json_object"})

        Returns:
            Response content string

        Raises:
            openai.OpenAIError: For API errors
        """
        messages = [
            {"role": "system", "content": system_message},
            {"role": "user", "content": user_prompt},
        ]

        kwargs = {"model": self.model, "messages": messages, "temperature": temperature}

        if response_format:
            kwargs["response_format"] = response_format

        response = self.client.chat.completions.create(**kwargs)
        return response.choices[0].message.content

    def _parse_json_response(self, content: str, default: Optional[Dict] = None) -> Dict:
        """Parse JSON response with error handling.

        Args:
            content: JSON string to parse
            default: Default dict to return on error (default: {})

        Returns:
            Parsed dictionary
        """
        if default is None:
            default = {}

        try:
            return json.loads(content)
        except json.JSONDecodeError as e:
            logger.error(f"JSON decode error: {e}", exc_info=True)
            return default

    def _create_error_validation_result(
        self, error: Exception, error_type: str
    ) -> ValidationResult:
        """Create error ValidationResult.

        Args:
            error: Exception that occurred
            error_type: Type of error

        Returns:
            ValidationResult with error information
        """
        return ValidationResult(
            is_valid=False,
            errors=[f"AI validation error: {str(error)}"],
            warnings=[],
            suggestions=[],
            confidence_score=0.0,
            details={"error": str(error), "error_type": error_type},
        )

    def _create_error_dict_response(
        self, error: Exception, default_keys: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Create error dictionary response.

        Args:
            error: Exception that occurred
            default_keys: Dictionary with default values for all keys

        Returns:
            Dictionary with error information
        """
        result = default_keys.copy()
        if "risk_analysis" in result:
            result["risk_analysis"] = f"Error: {str(error)}"
        elif "root_cause" in result:
            result["root_cause"] = f"Error: {str(error)}"
        return result

    def _handle_ai_error(
        self, error: Exception, error_type: str, result_type: str = "validation"
    ) -> Any:
        """Handle AI errors and return appropriate error response.

        Args:
            error: Exception that occurred
            error_type: Type of error
            result_type: Type of result ("validation", "dict", "string")

        Returns:
            Appropriate error response based on result_type
        """
        logger.error(f"AI {error_type} error: {error}", exc_info=True)

        if result_type == "validation":
            return self._create_error_validation_result(error, error_type)
        elif result_type == "dict":
            return {"error": str(error), "error_type": error_type}
        else:  # string
            return f"Error: {str(error)}"

    # ==================== 1. MODEL VALIDATION AND QUALITY ASSURANCE ====================

    def validate_model_quality(
        self, assumptions: Dict, industry: Optional[str] = None
    ) -> ValidationResult:
        """
        Validate LBO model assumptions for realism and common errors.

        Args:
            assumptions: LBO assumptions dictionary
            industry: Industry sector for context

        Returns:
            ValidationResult with warnings, errors, and suggestions
        """
        prompt = self._create_validation_prompt(assumptions, industry)
        system_message = (
            "You are a financial modeling expert specializing in LBO validation. "
            "Analyze assumptions for realism, consistency, and common errors. "
            "Provide specific, actionable feedback."
        )

        try:
            content = self._call_openai_api(
                system_message, prompt, temperature=0.3, response_format={"type": "json_object"}
            )
            result_json = self._parse_json_response(content)

            return ValidationResult(
                is_valid=result_json.get("is_valid", True),
                warnings=result_json.get("warnings", []),
                errors=result_json.get("errors", []),
                suggestions=result_json.get("suggestions", []),
                confidence_score=result_json.get("confidence_score", 0.8),
                details=result_json.get("details", {}),
            )
        except openai.OpenAIError as e:
            return self._handle_ai_error(e, "openai_api", "validation")
        except json.JSONDecodeError as e:
            return self._create_error_validation_result(e, "json_decode_error")
        except (ValueError, TypeError, KeyError) as e:
            return self._create_error_validation_result(e, "data_processing_error")
        except Exception as e:
            return self._handle_ai_error(e, "unexpected", "validation")

    def _create_validation_prompt(self, assumptions: Dict, industry: Optional[str]) -> str:
        """Create validation prompt."""
        prompt = f"""Analyze the following LBO model assumptions for quality and realism.

ASSUMPTIONS:
{json.dumps(assumptions, indent=2)}
"""
        if industry:
            prompt += f"\nINDUSTRY: {industry}"

        prompt += """
Validate for:
1. Realistic growth rates (industry-appropriate)
2. Reasonable EBITDA margins
3. Appropriate debt/equity ratios
4. Consistent working capital assumptions
5. Market-appropriate entry/exit multiples
6. Sustainable cash flow projections
7. Appropriate tax rates
8. Reasonable transaction and financing costs

Return JSON with:
{
    "is_valid": true/false,
    "warnings": ["list of warnings"],
    "errors": ["list of critical errors"],
    "suggestions": ["list of improvement suggestions"],
    "confidence_score": 0.0-1.0,
    "details": {
        "growth_rate_assessment": "analysis",
        "margin_assessment": "analysis",
        "debt_structure_assessment": "analysis",
        "multiple_assessment": "analysis"
    }
}
"""
        return prompt

    # ==================== 2. OUTPUT REVIEW AND CONSISTENCY CHECKS ====================

    def review_generated_model(self, excel_file_path: str, model_summary: Dict) -> ValidationResult:
        """
        Review generated Excel model for consistency and errors.

        Args:
            excel_file_path: Path to generated Excel file
            model_summary: Summary of model assumptions and key outputs

        Returns:
            ValidationResult with review findings
        """
        try:
            # Extract key information from Excel
            wb = openpyxl.load_workbook(excel_file_path, data_only=False)
            model_data = self._extract_model_data(wb)

            prompt = f"""Review the following LBO model Excel file for consistency and errors.

MODEL SUMMARY:
{json.dumps(model_summary, indent=2)}

MODEL DATA EXTRACT:
{json.dumps(model_data, indent=2)}

Check for:
1. Balance sheet balancing (Assets = Liabilities + Equity)
2. Cash flow reconciliation issues
3. Formula inconsistencies
4. Missing required sections
5. Circular references
6. Incorrect linkages between sheets
7. Debt schedule accuracy
8. Returns calculation correctness

Return JSON:
{
    "is_valid": true/false,
    "warnings": ["warnings"],
    "errors": ["critical errors"],
    "suggestions": ["fixes"],
    "confidence_score": 0.0-1.0,
    "details": {{
        "balance_sheet_check": "status",
        "cash_flow_check": "status",
        "formula_check": "status",
        "debt_schedule_check": "status"
    }}
}
"""

            system_message = (
                "You are an Excel financial model auditor. Review LBO models for errors, "
                "inconsistencies, and best practices."
            )
            content = self._call_openai_api(
                system_message, prompt, temperature=0.2, response_format={"type": "json_object"}
            )
            result_json = self._parse_json_response(content)

            return ValidationResult(
                is_valid=result_json.get("is_valid", True),
                warnings=result_json.get("warnings", []),
                errors=result_json.get("errors", []),
                suggestions=result_json.get("suggestions", []),
                confidence_score=result_json.get("confidence_score", 0.8),
                details=result_json.get("details", {}),
            )
        except openai.OpenAIError as e:
            return self._handle_ai_error(e, "openai_api", "validation")
        except Exception as e:
            return self._handle_ai_error(e, "unexpected", "validation")

    def _extract_model_data(self, wb: openpyxl.Workbook) -> Dict:
        """Extract key data from Excel workbook."""
        data = {}

        try:
            if "Final" in wb.sheetnames:
                ws = wb["Final"]
                # Extract key values (simplified extraction)
                data["sheets"] = wb.sheetnames
                data["has_sources_uses"] = any(
                    "SOURCES" in str(cell.value) for row in ws.iter_rows() for cell in row
                )
                data["has_income_statement"] = any(
                    "INCOME STATEMENT" in str(cell.value) for row in ws.iter_rows() for cell in row
                )
                data["has_balance_sheet"] = any(
                    "BALANCE SHEET" in str(cell.value) for row in ws.iter_rows() for cell in row
                )
                data["has_cash_flow"] = any(
                    "CASH FLOW" in str(cell.value) for row in ws.iter_rows() for cell in row
                )
                data["has_debt_schedule"] = any(
                    "DEBT SCHEDULE" in str(cell.value) for row in ws.iter_rows() for cell in row
                )
        except (KeyError, ValueError, TypeError):
            # Ignore missing or invalid keys in response
            pass

        return data

    # ==================== 3. SCENARIO ANALYSIS AND SENSITIVITY TESTING ====================

    def generate_sensitivity_scenarios(
        self, base_assumptions: Dict, industry: Optional[str] = None
    ) -> ScenarioAnalysis:
        """
        Generate High/Base/Low scenarios and sensitivity analysis.

        Args:
            base_assumptions: Base case assumptions
            industry: Industry sector

        Returns:
            ScenarioAnalysis with scenarios and sensitivity matrix
        """
        prompt = f"""Generate sensitivity scenarios for the following LBO model.

BASE ASSUMPTIONS:
{json.dumps(base_assumptions, indent=2)}
"""
        if industry:
            prompt += f"\nINDUSTRY: {industry}"

        prompt += """
Create:
1. High Case scenario (optimistic but realistic)
2. Low Case scenario (conservative/pessimistic)
3. Identify key assumptions to vary (revenue growth, EBITDA margin, exit multiple, etc.)
4. Sensitivity matrix showing impact of each assumption on IRR and MOIC

Return JSON:
{
    "base_case": {{"assumptions": {...}, "expected_irr": X, "expected_moic": Y}},
    "high_case": {{"assumptions": {...}, "expected_irr": X, "expected_moic": Y}},
    "low_case": {{"assumptions": {...}, "expected_irr": X, "expected_moic": Y}},
    "key_assumptions": ["assumption1", "assumption2"],
    "sensitivity_matrix": {{
        "revenue_growth": {{"-20%": {{"irr": X, "moic": Y}}, "+20%": {{"irr": X, "moic": Y}}}},
        "ebitda_margin": {{"-5pp": {{"irr": X, "moic": Y}}, "+5pp": {{"irr": X, "moic": Y}}}}
    }}
}
"""

        try:
            system_message = (
                "You are a financial analyst expert in scenario analysis and "
                "sensitivity testing for LBO models."
            )
            content = self._call_openai_api(
                system_message, prompt, temperature=0.4, response_format={"type": "json_object"}
            )
            result_json = self._parse_json_response(content)

            return ScenarioAnalysis(
                base_case=result_json.get("base_case", {}),
                high_case=result_json.get("high_case", {}),
                low_case=result_json.get("low_case", {}),
                key_assumptions=result_json.get("key_assumptions", []),
                sensitivity_matrix=result_json.get("sensitivity_matrix", {}),
            )
        except (openai.OpenAIError, json.JSONDecodeError, LBOAIServiceError) as e:
            logger.error(f"Error generating scenarios: {e}", exc_info=True)
            return ScenarioAnalysis(
                base_case=base_assumptions,
                high_case={},
                low_case={},
                key_assumptions=[],
                sensitivity_matrix={},
            )

    # ==================== 4. NATURAL LANGUAGE QUERY INTERFACE ====================

    def query_model(
        self, question: str, model_data: Dict, assumptions: Optional[Dict] = None
    ) -> str:
        """
        Answer natural language questions about the model.

        Args:
            question: User's question
            model_data: Model outputs and key metrics
            assumptions: Model assumptions (optional)

        Returns:
            Natural language answer
        """
        prompt = f"""Answer the following question about an LBO financial model.

QUESTION: {question}

MODEL DATA:
{json.dumps(model_data, indent=2)}
"""
        if assumptions:
            prompt += f"\nASSUMPTIONS:\n{json.dumps(assumptions, indent=2)}"

        prompt += "\n\nProvide a clear, concise answer. If you need more information, say so."

        try:
            system_message = (
                "You are a financial modeling expert helping users understand LBO models. "
                "Answer questions clearly and accurately."
            )
            content = self._call_openai_api(system_message, prompt, temperature=0.5)
            return content.strip()
        except Exception as e:
            return self._handle_ai_error(e, "query", "string")

    # ==================== 5. MARKET BENCHMARKING AND COMPARISON ====================

    def benchmark_against_market(self, assumptions: Dict, industry: str) -> BenchmarkResult:
        """
        Compare model assumptions against market benchmarks.

        Args:
            assumptions: Model assumptions
            industry: Industry sector

        Returns:
            BenchmarkResult with comparisons and recommendations
        """
        prompt = f"""Compare the following LBO model assumptions against industry benchmarks.

ASSUMPTIONS:
{json.dumps(assumptions, indent=2)}

INDUSTRY: {industry}

Compare:
1. Entry multiple vs. industry average
2. Revenue growth rates vs. industry norms
3. EBITDA margins vs. typical margins
4. Debt structure vs. market standards
5. Exit multiples vs. historical transactions
6. Expected returns (IRR/MOIC) vs. LBO fund targets

Return JSON:
{{
    "industry_averages": {{
        "entry_multiple": X,
        "revenue_growth": X,
        "ebitda_margin": X,
        "debt_to_ebitda": X,
        "exit_multiple": X
    }},
    "quartiles": {{
        "entry_multiple": {{"q25": X, "q50": X, "q75": X}},
        "revenue_growth": {{"q25": X, "q50": X, "q75": X}},
        "ebitda_margin": {{"q25": X, "q50": X, "q75": X}}
    }},
    "deviations": {{
        "entry_multiple_deviation": X,
        "revenue_growth_deviation": X,
        "ebitda_margin_deviation": X
    }},
    "recommendations": ["recommendation1", "recommendation2"]
}}
"""

        try:
            system_message = (
                "You are a market research analyst with expertise in LBO transaction data "
                "and industry benchmarks."
            )
            content = self._call_openai_api(
                system_message, prompt, temperature=0.3, response_format={"type": "json_object"}
            )
            result_json = self._parse_json_response(content)

            return BenchmarkResult(
                industry_averages=result_json.get("industry_averages", {}),
                quartiles=result_json.get("quartiles", {}),
                deviations=result_json.get("deviations", {}),
                recommendations=result_json.get("recommendations", []),
            )
        except Exception as e:
            logger.error(f"Error during benchmarking: {e}", exc_info=True)
            return BenchmarkResult(
                industry_averages={},
                quartiles={},
                deviations={},
                recommendations=[f"Benchmarking error: {str(e)}"],
            )

    # ==================== 6. FORMULA EXPLANATION AND DOCUMENTATION ====================

    def generate_model_documentation(self, excel_file_path: str, assumptions: Dict) -> str:
        """
        Generate comprehensive documentation for the model.

        Args:
            excel_file_path: Path to Excel file
            assumptions: Model assumptions

        Returns:
            Markdown documentation string
        """
        prompt = f"""Generate comprehensive documentation for an LBO financial model.

ASSUMPTIONS:
{json.dumps(assumptions, indent=2)}

Create documentation including:
1. Executive summary
2. Model overview and structure
3. Key assumptions explanation
4. Financial statement methodology
5. Debt schedule mechanics
6. Returns calculation methodology
7. How to use the model
8. Key formulas and their purposes
9. Scenario analysis guide
10. Troubleshooting common issues

Return markdown-formatted documentation.
"""

        try:
            system_message = (
                "You are a technical writer specializing in financial model documentation. "
                "Write clear, comprehensive documentation."
            )
            content = self._call_openai_api(system_message, prompt, temperature=0.4)
            return content.strip()
        except Exception as e:
            return self._handle_ai_error(e, "documentation", "string")

    # ==================== 7. ERROR DIAGNOSIS AND TROUBLESHOOTING ====================

    def diagnose_model_errors(
        self, error_message: str, assumptions: Dict, stack_trace: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Diagnose errors and provide fix suggestions.

        Args:
            error_message: Error message
            assumptions: Model assumptions
            stack_trace: Optional stack trace

        Returns:
            Dictionary with diagnosis and fixes
        """
        prompt = f"""Diagnose the following error in an LBO model generation process.

ERROR MESSAGE:
{error_message}
"""
        if stack_trace:
            prompt += f"\nSTACK TRACE:\n{stack_trace}\n"

        prompt += f"\nASSUMPTIONS:\n{json.dumps(assumptions, indent=2)}\n"
        prompt += """
Provide:
1. Root cause analysis
2. Specific fix suggestions
3. Common solutions for this error type
4. Prevention strategies

Return JSON:
{
    "root_cause": "analysis",
    "fix_suggestions": ["fix1", "fix2"],
    "common_solutions": ["solution1", "solution2"],
    "prevention": ["prevention1", "prevention2"],
    "severity": "high/medium/low"
}
"""

        default_response = {
            "root_cause": "",
            "fix_suggestions": [],
            "common_solutions": [],
            "prevention": [],
            "severity": "unknown",
        }

        try:
            system_message = (
                "You are a debugging expert specializing in financial modeling errors. "
                "Diagnose errors accurately and provide actionable fixes."
            )
            content = self._call_openai_api(
                system_message, prompt, temperature=0.3, response_format={"type": "json_object"}
            )
            return self._parse_json_response(content, default_response)
        except Exception as e:
            return self._create_error_dict_response(e, default_response)

    # ==================== 8. OPTIMIZATION SUGGESTIONS ====================

    def optimize_debt_structure(self, model_data: Dict, assumptions: Dict) -> Dict[str, Any]:
        """
        Suggest optimal debt structure and deal terms.

        Args:
            model_data: Model outputs (IRR, MOIC, cash flows)
            assumptions: Current assumptions

        Returns:
            Optimization recommendations
        """
        prompt = f"""Suggest optimizations for the following LBO structure.

CURRENT ASSUMPTIONS:
{json.dumps(assumptions, indent=2)}

MODEL OUTPUTS:
{json.dumps(model_data, indent=2)}

Suggest:
1. Optimal debt/equity mix
2. Best amortization schedules
3. Interest rate optimization
4. Timing improvements
5. Deal structure enhancements

Return JSON:
{
    "recommended_debt_structure": {...},
    "expected_irr_improvement": X,
    "expected_moic_improvement": X,
    "risk_analysis": "analysis",
    "implementation_steps": ["step1", "step2"]
}
"""

        default_response = {
            "recommended_debt_structure": {},
            "expected_irr_improvement": 0,
            "expected_moic_improvement": 0,
            "risk_analysis": "",
            "implementation_steps": [],
        }

        try:
            system_message = (
                "You are an LBO structuring expert. Optimize deal structures for maximum "
                "returns while managing risk."
            )
            content = self._call_openai_api(
                system_message, prompt, temperature=0.4, response_format={"type": "json_object"}
            )
            return self._parse_json_response(content, default_response)
        except Exception as e:
            return self._create_error_dict_response(e, default_response)

    # ==================== 9. DYNAMIC PROMPT ENHANCEMENT ====================

    def enhance_business_description(self, user_description: str) -> Dict[str, Any]:
        """
        Enhance and enrich business descriptions.

        Args:
            user_description: User's business description

        Returns:
            Enhanced description and suggestions
        """
        prompt = f"""Analyze and enhance the following business description for LBO modeling.

USER DESCRIPTION:
{user_description}

Provide:
1. Enhanced business description
2. Missing information to gather
3. Suggested metrics to include
4. Clarifying questions to ask

Return JSON:
{
    "enhanced_description": "enhanced version",
    "missing_information": ["info1", "info2"],
    "suggested_metrics": ["metric1", "metric2"],
    "clarifying_questions": ["question1", "question2"]
}
"""

        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {
                        "role": "system",
                        "content": "You are a business analyst expert at extracting and organizing business information for financial modeling.",
                    },
                    {"role": "user", "content": prompt},
                ],
                temperature=0.5,
                response_format={"type": "json_object"},
            )

            return json.loads(response.choices[0].message.content)

        except openai.OpenAIError as e:
            logger.error(f"OpenAI API error enhancing description: {e}", exc_info=True)
            return {
                "enhanced_description": user_description,
                "missing_information": [],
                "suggested_metrics": [],
                "clarifying_questions": [],
            }
        except Exception as e:
            logger.error(f"Unexpected error enhancing description: {e}", exc_info=True)
            return {
                "enhanced_description": user_description,
                "missing_information": [],
                "suggested_metrics": [],
                "clarifying_questions": [],
            }

    # ==================== 10. REAL-TIME GUIDANCE DURING MODEL BUILDING ====================

    def get_contextual_help(
        self, current_step: str, user_inputs: Dict, field_name: Optional[str] = None
    ) -> str:
        """
        Provide contextual help during model building.

        Args:
            current_step: Current step in model building
            user_inputs: Current user inputs
            field_name: Specific field being filled (optional)

        Returns:
            Helpful guidance text
        """
        prompt = f"""Provide contextual help for LBO model building.

CURRENT STEP: {current_step}
"""
        if field_name:
            prompt += f"CURRENT FIELD: {field_name}\n"

        prompt += f"CURRENT INPUTS:\n{json.dumps(user_inputs, indent=2)}\n"
        prompt += """
Provide:
1. Explanation of what this step/field means
2. Typical values or ranges
3. Common mistakes to avoid
4. Industry context if relevant
5. How this affects the model

Keep response concise and actionable (2-3 sentences).
"""

        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {
                        "role": "system",
                        "content": "You are a helpful financial modeling assistant providing real-time guidance.",
                    },
                    {"role": "user", "content": prompt},
                ],
                temperature=0.5,
            )

            return response.choices[0].message.content.strip()

        except openai.OpenAIError as e:
            logger.error(f"OpenAI API error providing help: {e}", exc_info=True)
            return f"Error providing help: {str(e)}"
        except Exception as e:
            logger.error(f"Unexpected error providing help: {e}", exc_info=True)
            return f"Error providing help: {str(e)}"


# Convenience functions
def validate_lbo_model(
    assumptions: Dict, industry: Optional[str] = None, api_key: Optional[str] = None
) -> ValidationResult:
    """Convenience function for model validation."""
    validator = LBOModelAIValidator(api_key=api_key)
    return validator.validate_model_quality(assumptions, industry)


def query_lbo_model(
    question: str,
    model_data: Dict,
    assumptions: Optional[Dict] = None,
    api_key: Optional[str] = None,
) -> str:
    """Convenience function for querying models."""
    validator = LBOModelAIValidator(api_key=api_key)
    return validator.query_model(question, model_data, assumptions)


if __name__ == "__main__":
    # Example usage
    print("LBO AI Validator Module")
    print("=" * 60)
    print("This module provides comprehensive AI-powered features for LBO models.")
    print("\nAvailable features:")
    print("1. Model validation and quality assurance")
    print("2. Output review and consistency checks")
    print("3. Scenario analysis generation")
    print("4. Natural language query interface")
    print("5. Market benchmarking")
    print("6. Documentation generation")
    print("7. Error diagnosis")
    print("8. Optimization suggestions")
    print("9. Prompt enhancement")
    print("10. Real-time guidance")
